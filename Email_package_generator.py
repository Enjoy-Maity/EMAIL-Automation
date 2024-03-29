import pandas as pd                                                 # Importing pandas with alias pd for reading the excel sheet and manipulating it freely.
from tkinter import messagebox                                      # Importing Messagebox from tkinter to display messages.
from openpyxl import load_workbook                                  # Importing load_workbook class from the openpyxl to load existing excel workbook.
from openpyxl.styles import Font,Border,Side,PatternFill,Alignment  # Importing classes from openpyxl to style the excel workbooks.
from openpyxl import Workbook                                       # Importing Workbook to Create Workbook using openpyxl.
from openpyxl.utils import get_column_letter,quote_sheetname        # Importing the get_column_letter from openpyxl to convert the column numbers to alphabet letter used in the excel sheet.
from datetime import datetime, timedelta                            # Importing the datetime and timedelta from datetime module, to filter out the excel sheet basd on today's maintenance date.
from openpyxl.worksheet.datavalidation import DataValidation        # Importing DataValidation from the openpyxl module to add data validation onto fields in email-package

flag = ""
workbook1 = ""

# Creating classes for custom made exceptions inheriting the default Exception class for raising and handling custom raised exceptions.
class CustomException(Exception):
    # Defining the Counstructor method for the CustomException Class
    def __init__(self,title,message):
        self.title      = title 
        self.message    = message
        
        #Calling the super(base) class and passing the arguments to the base class.
        super().__init__(self.title,self.message)
        
        # Displaying the message with the custom exception title passed to the object of the class for the CustomException.
        messagebox.showerror(self.title, self.message)

# Creating method for styling the worksheet.
def styling(workbook,sheetname):
    wb  =  load_workbook(workbook)                          # loading the workbook.
    ws  =  wb[sheetname]                                    # loading the worksheet.
    font_style  =  Font(color = "FFFFFF",bold = True)       # Setting the font style with color.
    col_widths = []                                         # Empty list for entering the max length of string in each column.

    # Iterating through the row values to find the max length of string in each column in the row and appending it to the col_widths list

    for row_values in ws.iter_rows(values_only = True):
        for j,value in enumerate(row_values):
            if len(col_widths)>j:
                if col_widths[j] < len(str(value)):
                    col_widths[j] = len(str(value))
            else:
                col_widths.insert(j,len(str(value)))

    # Standardising the length of each column in the sheet.

    for i,column_width in enumerate(col_widths,1):
        if column_width <= 47:
            ws.column_dimensions[get_column_letter(i)].width = column_width+3
        else:
            ws.column_dimensions[get_column_letter(i)].width = 50


    # Coloring the headers and alingning the headers text to center both horizontally and vertically.
    for column in range(1,ws.max_column+1):   # ws.max_column returns the total number of columns present
        col = get_column_letter(column)
        color_fill = PatternFill(start_color = '0033CC',end_color = '0033CC',fill_type = 'solid')
        ws[col+'1'].font = font_style
        ws[col+'1'].fill = color_fill
        ws[col+'1'].alignment = Alignment(horizontal = 'center',vertical = 'center')

    # Styling all the occupied cells in the sheet, by adding border to the cells, aligning the text in the center
    
    for row in ws:
        for cell in row:
            cell.alignment = Alignment(horizontal = 'center',vertical = 'center',wrap_text=False)
            cell.border = Border(top = Side(border_style = 'medium',color = '000000'),bottom = Side(border_style = 'medium',color = '000000'),left = Side(border_style = 'medium',color = '000000'),right = Side(border_style = 'medium',color = '000000'))

    # Saving the workbook with worksheet with all the changes.
    wb.save(workbook)
    wb.close()
    del wb
    
    objects = dir()
    for object in objects:
        if not object.startswith("__"):
            del object

# removing existing data validation
def removeExistingCellDataValidation(worksheet, cell):
    toRemove = []

    # Append all validation rules for cell to be removed.
    for validation in worksheet.data_validations.dataValidation:
        if validation.__contains__(cell):
            toRemove.append(validation)

    # Process all data validation rules set for removal.
    for rmValidation in toRemove:
        worksheet.data_validations.dataValidation.remove(rmValidation)

# Adding Data Validation
def validation_adder(workbook,worksheet):
    wb                          = load_workbook(workbook)
    ws                          = wb[worksheet]
    # getting max occupied rows
    maxrows                     = ws.max_row
    
    workbk                      = pd.ExcelFile(workbook)
    mail_id_sheet               = pd.read_excel(workbk,"Mail Id")

    for i in range(2,maxrows+1):
        removeExistingCellDataValidation(ws,ws[f'K{i}'])

    # Getting th unique change responsible from the mail id sheet
    unique_change_responsible   = list(mail_id_sheet['Change Responsible'].dropna().unique())
    unique_change_responsible.sort()
    len_of_list_of_unique_change_responsible = len(unique_change_responsible)

    # # Removing the Nan value if there's any such value in the unique change responsible list.
    # for i in unique_change_responsible:
    #     if (str(i).upper().strip() == 'NAN'):
    #         unique_change_responsible.remove(i)

    # unique_change_responsible   = f"{', '.join(unique_change_responsible)}"
    # unique_change_responsible   = f'"{unique_change_responsible}"'
    # print(len(unique_change_responsible))
    

    # Rules for data validation with their error message, title and prompt message, title.
    rule1               = DataValidation(type = "list", formula1=f"{quote_sheetname('Mail Id')}!$F$2:$F${len_of_list_of_unique_change_responsible+1}", allow_blank = True)
    rule1.error         = "Your Entry is Invalid!"
    rule1.errorTitle    = "Invalid Entry!"

    rule1.prompt        = "Please Select from the list"
    rule1.promptTitle   = "List Selection"
    
    # Adding the rules to the woksheet.
    ws.add_data_validation(rule1)
    
    
    
    # Setting the rows for the rules.
    range_setter_var_change_responsible    = f'K2:K{maxrows}'
    

    # Adding the ranges to the rules.
    rule1.add(range_setter_var_change_responsible)
    
    
    # Saving the Workbook
    wb.save(workbook)
    wb.close()
    workbk.close()
    del wb

    objects = dir()
    for object in objects:
        if not object.startswith("__"):
            del object

# Creating the main driver Method(Function) 
def email_package_sheet_creater(workbook):
    try:
        workbook1 = workbook
        # Checking if the workboook is not selected.
        if (len(workbook) == 0):
            raise CustomException(" File Not Selected!","Kindly select the MPBN Planning Workbook for Email-Package Creation!")
        
        # Reading the excel workbook for sheet in pandas.
        wb = pd.ExcelFile(workbook)
        sheets = wb.sheet_names

        temp_flag   = 0
        for sheet in sheets:
            if(sheet == "Planning Sheet"):
                temp_flag = 1
                break

        if (temp_flag == 0):
            del wb
            raise CustomException(" Planning Sheet Absent!","Kindly check the selected workbook for 'Planning Sheet' worksheet!")

        daily_plan_sheet  = pd.read_excel(wb,"Planning Sheet")

        # Filtering out the rows with planning status "Planned" and execution date for today's date
        todays_maintenance_date = datetime.now() + timedelta(1)
        tomorrow                = todays_maintenance_date
        todays_maintenance_date = todays_maintenance_date.strftime("%m/%d/%Y")
        try:
            daily_plan_sheet['Execution Date'] = pd.to_datetime(daily_plan_sheet['Execution Date'], format = "%d-%b-%Y")
            daily_plan_sheet['Execution Date'] = daily_plan_sheet['Execution Date'].dt.strftime("%m/%d/%Y")
            
        
        except:
            daily_plan_sheet['Execution Date'] = pd.to_datetime(daily_plan_sheet['Execution Date'], format = "%m/%d/%Y")
            daily_plan_sheet['Execution Date'] = daily_plan_sheet['Execution Date'].dt.strftime("%m/%d/%Y")
            

        # Checking if there's any data present or not.
        if (len(daily_plan_sheet) == 0):
            del daily_plan_sheet
            del wb
            raise CustomException(" Data Missing!","The Planning Sheet is empty!, Kindly Check!")
        
        difference = []
        
        for i in range(0,len(daily_plan_sheet)):
            if (daily_plan_sheet.iloc[i]['Execution Date'] != tomorrow.strftime('%m/%d/%Y')):
                difference.append(str(daily_plan_sheet.iloc[i]['S.NO']))

        daily_plan_sheet = daily_plan_sheet[daily_plan_sheet['Execution Date'] == todays_maintenance_date]

        if (len(daily_plan_sheet) == 0):
            del daily_plan_sheet
            del wb
            raise CustomException(" Today's Maintenance Data Missing!","Today's Maintenance Data is missing in the Planning Sheet! Kindly Check!")
        
        if (len(difference) > 0):
            del daily_plan_sheet
            del wb
            raise CustomException(" Maintenance Data for Other Date Present!",f"All the CR's present are not of Today's Maintenace Date for S.NO : {', '.join([str(num) for num in difference])}")
        
        else:
            # Filling the NA(Blank) cells with TempNA string.
            daily_plan_sheet.fillna("TempNA", inplace = True)

            # Getting the unique values of the planning status column of the excel sheet.
            planned_status_unique_values = list(daily_plan_sheet['Planning Status'].unique())
            
            for i in range(0,len(planned_status_unique_values)):
                # Changing the state of the unique inputs in the planning status column of the excel sheet.
                planned_status_unique_values[i] = planned_status_unique_values[i].strip().upper()
            
            if ((len(planned_status_unique_values) == 1) and (planned_status_unique_values[0] == "NA")):
                del daily_plan_sheet
                del wb
                
                # Raising custom exception for condition when there's no input in the planning status column of the Excel Sheet.
                raise CustomException(" Input Missing!","Kindly Enter the Planning Status input in uploaded sheet!")
            
            if ((len(planned_status_unique_values) > 1)):
                if ("NA" in planned_status_unique_values):
                    # Empty list for adding S.NO of rows with wrong input.
                    input_error = []
                    for i in range(0,len(daily_plan_sheet)):
                        if (daily_plan_sheet.iloc[i]['Planning Status'] == "NA"):
                            # Appending the S.NO of row with wrong input.
                            input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                    del daily_plan_sheet
                    del wb
                
                    # Raising the Exception for rows with no planning status.
                    raise CustomException(" Input Missing!",f"Planning Status Input is Missing for the below S.NO:\n{', '.join(str(num) for num in input_error)}")
                
                if ("PLANNED" in planned_status_unique_values) or ("SWAPPED" in planned_status_unique_values):
                    # Filtering the rows with planned crs
                    daily_plan_sheet['Planning Status'] = daily_plan_sheet['Planning Status'].str.strip()
                    daily_plan_sheet = daily_plan_sheet[daily_plan_sheet['Planning Status'].str.upper().str.contains('PLANNED|SWAPPED')]

                else:
                    del daily_plan_sheet
                    del wb
                
                    # Raising Custom Exception for not finding any dataframe row with Planning Status.
                    raise CustomException(" Incorrect Input","Kindly Check the Planning Status input in uploaded sheet!")
                
            Email_ID = pd.read_excel(workbook,"Mail Id")
        
            # Finding the Circles in the Mail ID worksheet of the MPBN Daily Planning workbook.
            circles = Email_ID['Circle'].tolist()

            # Creating empty lists for input error and circle not proper S.NO
            input_error         = []
            circle_not_proper   = []
            
            # Iterating (Looping) through the daily_plan_sheet dataframe index wise (index given by pandas to each row with data), to find out the serial 
            # numbers of the rows where the Circle input and the Change responsible is not properly entered by the user and any other fields that should be left unblank
            # by the user.
            for i in range(0,len(daily_plan_sheet)):
                if (daily_plan_sheet.iloc[i]['CR NO'] == "TempNA") or (daily_plan_sheet.iloc[i]['CR NO'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    continue
                if (daily_plan_sheet.iloc[i]['Circle'] not in circles):
                    circle_not_proper.append(daily_plan_sheet.iloc[i]['S.NO'])
                    continue
                if (daily_plan_sheet.iloc[i]['Activity Title'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Activity Title'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    continue
                if (daily_plan_sheet.iloc[i]['Circle'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Circle'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    continue
                if (daily_plan_sheet.iloc[i]['Risk'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Risk'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    continue
                if (daily_plan_sheet.iloc[i]['Location'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Location'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    continue
                if (daily_plan_sheet.iloc[i]['Impact'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Impact'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    continue
                if (daily_plan_sheet.iloc[i]['Technical Validator'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Technical Validator'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    continue
                if (daily_plan_sheet.iloc[i]['Activity-Type'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Activity-Type'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    continue
                if (daily_plan_sheet.iloc[i]['Vendor'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Vendor'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    continue
                if (daily_plan_sheet.iloc[i]['Protocol'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Protocol'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    continue
                if (daily_plan_sheet.iloc[i]['Execution Projection'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Execution Projection'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    continue
            
            if (len(input_error) > 0):
                # messagebox.showerror("  Input Errors",f"Input Details are missing for mandatory parameters for S.NO.: {','.join([str(num) for num in input_error])} in uploaded Planning sheet!")
                
                del daily_plan_sheet
                del wb

                # Deleting all the variables before returning the value for "Unsuccessful"
                objects = dir()
                for object in objects:
                    if not object.startswith("__"):
                        del object
                
                flag = 'Unsuccessful'

                raise CustomException("  Input Errors",f"Input Details are missing for mandatory parameters for S.NO.: {','.join([str(num) for num in input_error])} in uploaded Planning sheet!")
            
            if (len(circle_not_proper) > 0):
                # messagebox.showerror("  Circles Errors",f"Input Circles are wrong in Planning Sheet for S.NO. : {','.join([str(num) for num in circle_not_proper])}")
                
                del daily_plan_sheet
                del wb

                # Deleting all the variables before returning the value for "Unsuccessful"
                objects = dir()
                for object in objects:
                    if not object.startswith("__"):
                        del object
                
                flag = 'Unsuccessful'
                raise CustomException("  Circles Errors",f"Input Circles are wrong in Planning Sheet for S.NO. : {','.join([str(num) for num in circle_not_proper])}")
            
            # Writing into the Email-Package Sheet
            daily_plan_sheet_unique_cr = daily_plan_sheet['CR NO'].value_counts().index.to_list()
            
            # Creating empty lists for different columns.
            execution_date                                      = []
            maintenance_window                                  = []
            cr_no                                               = []
            activity_title                                      = []
            risk                                                = []
            location                                            = []
            circle                                              = []
            no_of_node_involved                                 = []
            cr_belongs_to_same_activity_of_previous_cr_yes_no   = []
            activity_checker                                    = []
            activity_initiator                                  = []
            impact                                              = []
            planning_status                                     = []
            domain                                              = []
            design_availability                                 = []
            technical_validator                                 = []
            complexity                                          = []
            activity_type                                       = []
            domain_kpi                                          = []
            impacted_node                                       = []
            kpi_details                                         = []
            oss_name                                            = []
            oss_IP                                              = []
            total_time_spent_on_planned_crs_mins                = []
            vendor                                              = []
            protocol                                            = []
            execution_projection                                = []

            for idx,cr in enumerate(daily_plan_sheet_unique_cr):
                # Creating the count variable to assign the number of occurences of the CR in the 
                count = daily_plan_sheet['CR NO'].value_counts()[idx]
                
                # Setting the counter to 0 to run a loop until the count value of the CR, to assess all the rows in the dataframe with the same CR Number.
                counter = 0

                # Creating temp variables to hold the temporary data until that temporary data is written onto the result dataframe.
                # Change responsible	(K)
                # Final Status		(Q)
                # Reason For Rollback / Cancel(R)
                # Inter-domain Name	(AF)
                # Second Level Validation Status(AG)	
                # Inter-domain KPI status	(AH)
                # MOP View Status		(AI)
                
                execution_date_temp                                     = daily_plan_sheet.iloc[0]['Execution Date']
                cr_no_temp                                              = cr
                maintenance_window_temp                                 =  ''
                activity_title_temp                                     =  ''
                risk_temp                                               =  ''
                location_temp                                           =  ''
                circle_temp                                             =  ''
                no_of_node_involved_temp                                =  ''
                cr_belongs_to_same_activity_of_previous_cr_yes_no_temp  =  ''
                activity_checker_temp                                   =  ''
                activity_initiator_temp                                 =  ''
                impact_temp                                             =  ''
                planning_status_temp                                    =  ''
                domain_temp                                             =  ''
                reason_for_rollback_cancel_temp                         =  ''
                design_availability_temp                                =  ''
                technical_validator_temp                                =  ''
                complexity_temp                                         =  ''
                activity_type_temp                                      =  ''
                domain_kpi_temp                                         =  ''
                impacted_node_temp                                      =  ''
                kpi_details_temp                                        =  ''
                oss_name_temp                                           =  ''
                oss_IP_temp                                             =  ''
                total_time_spent_on_planned_crs_mins_temp               =  ''
                vendor_temp                                             =  ''
                protocol_temp                                           =  ''
                execution_projection_temp                               =  ''

                # Starting another loop to collect all data of particular CR no. from the daily_plan_sheet dataframe to assess the data and manipulate it according to our needs.
                for i in range(0,len(daily_plan_sheet)):
                    if (daily_plan_sheet.iloc[i]['CR NO'] == cr):
                        if (counter<count):
                            if (count > 1):
                                # Data for the RAN should be written first for the CR Data, if there's any row with RAN domain KPI for the CR number. 
                                if (str(daily_plan_sheet.iloc[i]['Domain kpi']).upper().startswith('RAN')):
                                    if ((len(str(daily_plan_sheet.iloc[i]['IMPACTED NODE']).strip()) == 0) or (daily_plan_sheet.iloc[i]['IMPACTED NODE']) == "TempNA"):
                                        impacted_node_temp = impacted_node_temp
                                    else:
                                        if (len(str(impacted_node_temp)) == 0):
                                            if ((str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() != 'TempNA') and (str(daily_plan_sheet.iloc[i]['IMPACTED NODE']).strip() != 'TempNA')):
                                                impacted_node_temp = f"({str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()} ):- {str(daily_plan_sheet.iloc[i]['IMPACTED NODE'])}"
                                        else:
                                            if (str(impacted_node_temp).__contains__(str(daily_plan_sheet.iloc[i]['Domain kpi'])) == False):
                                                if ((str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() != 'TempNA') and (str(daily_plan_sheet.iloc[i]['IMPACTED NODE']).strip() != 'TempNA')):
                                                    impacted_node_temp = f"({str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()} ):- {str(daily_plan_sheet.iloc[i]['IMPACTED NODE'])} || {impacted_node_temp}"
                                    
                                    if (len(str(domain_kpi_temp)) == 0):
                                        if(str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() != 'TempNA'):
                                            domain_kpi_temp = f"{str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()}"
                                        else:
                                            domain_kpi_temp = daily_plan_sheet.iloc[i]['Domain kpi']
                                    
                                    elif ((len(str(domain_kpi_temp)) > 0) and (str(domain_kpi_temp).__contains__(str(daily_plan_sheet.iloc[i]['Domain kpi'])) == False)):
                                        if (str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() != 'TempNA'):
                                            domain_kpi_temp = f"{str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()} || {domain_kpi_temp}"
                                    
                                    if ((len(str(daily_plan_sheet.iloc[i]['KPI DETAILS']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['KPI DETAILS']).strip() == 'TempNA')):
                                        kpi_details_temp = kpi_details_temp
                                    
                                    else:
                                        if (len(str(kpi_details_temp)) == 0):
                                            if ((str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() != 'TempNA') and (str(daily_plan_sheet.iloc[i]['KPI DETAILS']).strip() != 'TempNA')):
                                                kpi_details_temp = f"({str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()} ):- {(daily_plan_sheet.iloc[i]['KPI DETAILS'])}"
                                        
                                        if ((len(str(kpi_details_temp))>0) and (str(kpi_details_temp).__contains__(str(daily_plan_sheet.iloc[i]['Domain kpi'])) == False)):
                                            if ((str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() != 'TempNA') and (str(daily_plan_sheet.iloc[i]['KPI DETAILS']).strip() != 'TempNA')):
                                                kpi_details_temp = f"({str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()} ):- {(daily_plan_sheet.iloc[i]['KPI DETAILS'])} || {kpi_details_temp}"
                                    
                                    if (len(str(daily_plan_sheet.iloc[i]['oss name']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['oss name']).strip() == 'TempNA') :
                                        oss_name_temp = oss_name_temp
                                    else: 
                                        oss_name_temp  =  daily_plan_sheet.iloc[i]['oss name']
                                    
                                    if (len(str(daily_plan_sheet.iloc[i]['oss ip']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['oss ip']).strip() == 'TempNA') :
                                        oss_IP_temp = oss_IP_temp
                                    else:
                                        oss_IP_temp  =  daily_plan_sheet.iloc[i]['oss ip']

                                    if (len(str(maintenance_window_temp))) == 0:
                                        if (len(str(daily_plan_sheet.iloc[i]['Maintenance Window']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Maintenance Window']).strip() == 'TempNA'):
                                            maintenance_window_temp = str(maintenance_window_temp)
                                        else:
                                            maintenance_window_temp  = str(daily_plan_sheet.iloc[i]['Maintenance Window'])

                                    if(len(str(activity_title_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Activity Title']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Activity Title']).strip() == 'TempNA'):
                                            activity_title_temp = str(activity_title_temp)
                                        else:
                                            activity_title_temp  =  str(daily_plan_sheet.iloc[i]['Activity Title'])

                                    if(len(str(risk_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Risk']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Risk']).strip() == 'TempNA'):
                                            risk_temp = str(risk_temp)
                                        else:
                                            risk_temp  =  str(daily_plan_sheet.iloc[i]['Risk'])

                                    if (len(str(location_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Location']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Location']).strip() == 'TempNA'):
                                            location_temp = str(location_temp)
                                        else:    
                                            location_temp  =  str(daily_plan_sheet.iloc[i]['Location'])

                                    if (len(str(circle_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Circle']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Circle']).strip() == 'TempNA')):
                                            circle_temp = str(circle_temp)
                                        else:
                                            circle_temp  =  str(daily_plan_sheet.iloc[i]['Circle'])
                                    
                                    if (len(str(no_of_node_involved_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['No. of Node Involved']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['No. of Node Involved']).strip() == 'TempNA'):
                                            no_of_node_involved_temp = str(no_of_node_involved_temp)
                                        else:
                                            no_of_node_involved_temp  =  daily_plan_sheet.iloc[i]['No. of Node Involved']
                                    
                                    if (len(str(cr_belongs_to_same_activity_of_previous_cr_yes_no_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['CR Belongs to Same Activity of Previous CR- Yes/NO']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['CR Belongs to Same Activity of Previous CR- Yes/NO']).strip() == 'TempNA'):
                                            cr_belongs_to_same_activity_of_previous_cr_yes_no_temp = str(cr_belongs_to_same_activity_of_previous_cr_yes_no_temp)
                                        else:
                                            cr_belongs_to_same_activity_of_previous_cr_yes_no_temp  =  str(daily_plan_sheet.iloc[i]['CR Belongs to Same Activity of Previous CR- Yes/NO'])
                        
                                    if (len(str(activity_checker_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Activity Checker']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Activity Checker']).strip() == 'TempNA'):
                                            activity_checker_temp = str(activity_checker_temp)
                                        else:
                                            activity_checker_temp  =  str(daily_plan_sheet.iloc[i]['Activity Checker'])


                                    if (len(str(activity_initiator_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Activity Initiator']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Activity Initiator']).strip() == 'TempNA'):
                                            activity_initiator_temp = str(activity_initiator_temp)
                                        else:
                                            activity_initiator_temp  =  str(daily_plan_sheet.iloc[i]['Activity Initiator'])

                                    if (len(str(impact_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Impact']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Impact']).strip() == 'TempNA'):
                                            impact_temp = str(impact_temp)
                                        else:
                                            impact_temp  = str(daily_plan_sheet.iloc[i]['Impact'])

                                    if (len(str(planning_status_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Planning Status']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Planning Status']).strip() == 'TempNA'):
                                            planning_status_temp = str(planning_status_temp)
                                        else:
                                            planning_status_temp  =  str(daily_plan_sheet.iloc[i]['Planning Status'])

                                    if (len(str(domain_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Domain']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Domain']).strip() == 'TempNA'):
                                            domain_temp = str(domain_temp)
                                        else:
                                            domain_temp  =  str(daily_plan_sheet.iloc[i]['Domain'])

                                    if (len(str(design_availability_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Design Availability']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Design Availability']).strip() == 'TempNA'):
                                            design_availability_temp  =  str(design_availability_temp)
                                        else:
                                            design_availability_temp  =  str(daily_plan_sheet.iloc[i]['Design Availability'])

                                    if (len(str(technical_validator_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Technical Validator']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Technical Validator']).strip() == 'TempNA'):
                                            technical_validator_temp  =  str(technical_validator_temp)
                                        else:
                                            technical_validator_temp  =  str(daily_plan_sheet.iloc[i]['Technical Validator'])

                                    if (len(complexity_temp) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Complexity']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Complexity']).strip() == 'TempNA'):
                                            complexity_temp  =  str(complexity_temp)
                                        else:
                                            complexity_temp  =  str(daily_plan_sheet.iloc[i]['Complexity'])

                                    if (len(str(activity_type_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Activity-Type']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Activity-Type']).strip() == 'TempNA'):
                                            activity_type_temp  =  str(activity_type_temp)
                                        else:
                                            activity_type_temp  =  str(daily_plan_sheet.iloc[i]['Activity-Type'])

                                    if (len(str(total_time_spent_on_planned_crs_mins_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Total Time spent on Planned CRs (Mins)']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Total Time spent on Planned CRs (Mins)']).strip() == 'TempNA'):
                                            total_time_spent_on_planned_crs_mins_temp  =  str(total_time_spent_on_planned_crs_mins_temp)
                                        else:
                                            total_time_spent_on_planned_crs_mins_temp  =  str(daily_plan_sheet.iloc[i]['Total Time spent on Planned CRs (Mins)'])

                                    if (len(str(vendor_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Vendor']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Vendor']).strip() == 'TempNA'):
                                            vendor_temp  =  str(vendor_temp)
                                        else:
                                            vendor_temp  =  str(daily_plan_sheet.iloc[i]['Vendor'])

                                    if (len(str(protocol_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Protocol']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Protocol']).strip() == 'TempNA'):
                                            protocol_temp  =  str(protocol_temp)
                                        else:
                                            protocol_temp  =  str(daily_plan_sheet.iloc[i]['Protocol'])

                                    if (len(str(execution_projection_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Execution Projection']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Execution Projection']).strip() == 'TempNA'):
                                            execution_projection_temp  =  str(execution_projection_temp)
                                        else:
                                            execution_projection_temp  =  str(daily_plan_sheet.iloc[i]['Execution Projection'])

                                    
                                else:
                                    if (len(str(daily_plan_sheet.iloc[i]['IMPACTED NODE']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['IMPACTED NODE']).__contains__('NA')):
                                        impacted_node_temp = impacted_node_temp
                                    else:
                                        if (len(impacted_node_temp) == 0):
                                            if ((str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() != 'TempNA') and (str(daily_plan_sheet.iloc[i]['IMPACTED NODE']).strip() != 'TempNA')):
                                                impacted_node_temp = f"({str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()} ):- {str(daily_plan_sheet.iloc[i]['IMPACTED NODE'])}"
                                        else:
                                            if (str(impacted_node_temp).__contains__(str(daily_plan_sheet.iloc[i]['Domain kpi'])) == False):
                                                if ((str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() != 'TempNA') and (str(daily_plan_sheet.iloc[i]['IMPACTED NODE']).strip() != 'TempNA')):
                                                    impacted_node_temp +=  f" || ({str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()} ):- {str(daily_plan_sheet.iloc[i]['IMPACTED NODE'])}"
                                    
                                    if (len(domain_kpi_temp) == 0):
                                        if(str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() != 'TempNA'):
                                            domain_kpi_temp = str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()
                                        
                                        else:
                                            domain_kpi_temp = daily_plan_sheet.iloc[i]['Domain kpi']
                                    
                                    elif ((len(domain_kpi_temp)>0) and (str(domain_kpi_temp).__contains__(str(daily_plan_sheet.iloc[i]['Domain kpi'])) == False)):
                                        if (str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() != 'TempNA'):
                                            domain_kpi_temp +=  f" || {str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()}"
                                    
                                    if (len(str(daily_plan_sheet.iloc[i]['KPI DETAILS']).strip()) == 0):
                                        kpi_details_temp = kpi_details_temp
                                    else:
                                        if (len(kpi_details_temp) == 0):
                                            if ((str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() != 'TempNA') and (str(daily_plan_sheet.iloc[i]['KPI DETAILS']).strip() != 'TempNA')):
                                                kpi_details_temp = f"({str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()} ):- {str(daily_plan_sheet.iloc[i]['KPI DETAILS'])}"
                                        
                                        elif ((len(kpi_details_temp) > 0) and (str(kpi_details_temp).__contains__(str(daily_plan_sheet.iloc[i]['Domain kpi'])) == False)):
                                            if ((str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() != 'TempNA') and (str(daily_plan_sheet.iloc[i]['KPI DETAILS']).strip() != 'TempNA')):
                                                kpi_details_temp +=  f" || ({str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()} ):- {str(daily_plan_sheet.iloc[i]['KPI DETAILS'])}"
                                    
                                    if ((len(str(daily_plan_sheet.iloc[i]['oss name']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['oss name']).strip() == 'TempNA')):
                                        oss_name_temp = oss_name_temp
                                    else: 
                                        oss_name_temp = oss_name_temp
                                    
                                    if ((len(str(daily_plan_sheet.iloc[i]['oss ip']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['oss ip']).strip() == 'TempNA')):
                                        oss_IP_temp = oss_IP_temp
                                    else:
                                        oss_IP_temp  =  daily_plan_sheet.iloc[i]['oss ip']

                                    if (len(maintenance_window_temp)) == 0:
                                        if ((len(str(daily_plan_sheet.iloc[i]['Maintenance Window']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Maintenance Window']).strip() == 'TempNA')):
                                            maintenance_window_temp = maintenance_window_temp
                                        else:
                                            maintenance_window_temp  =  daily_plan_sheet.iloc[i]['Maintenance Window']
                                    

                                    if(len(activity_title_temp) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Activity Title']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Activity Title']).strip() == 'TempNA')):
                                            activity_title_temp = activity_title_temp
                                        else:
                                            activity_title_temp  =  daily_plan_sheet.iloc[i]['Activity Title']

                                    if(len(risk_temp) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Risk']).strip()) == 0) or ((str(daily_plan_sheet.iloc[i]['Risk'])).strip() == 'TempNA')):
                                            risk_temp = str(risk_temp)
                                        else:
                                            risk_temp  =  str(daily_plan_sheet.iloc[i]['Risk'])

                                    if (len(str(location_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Location']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Location']).strip() == 'TempNA')):
                                            location_temp = str(location_temp)
                                        else:    
                                            location_temp  =  daily_plan_sheet.iloc[i]['Location']

                                    if (len(circle_temp) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Circle']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Circle']).strip() == 'TempNA')):
                                            circle_temp = str(circle_temp)
                                        else:
                                            circle_temp  =  str(daily_plan_sheet.iloc[i]['Circle'])
                                    
                                    if (len(str(no_of_node_involved_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['No. of Node Involved']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['No. of Node Involved']).strip() == 'TempNA')):
                                            no_of_node_involved_temp = no_of_node_involved_temp
                                        else:
                                            no_of_node_involved_temp  =  daily_plan_sheet.iloc[i]['No. of Node Involved']
                                    
                                    if (len(cr_belongs_to_same_activity_of_previous_cr_yes_no_temp) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['CR Belongs to Same Activity of Previous CR- Yes/NO']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['CR Belongs to Same Activity of Previous CR- Yes/NO']).strip() == 'TempNA')):
                                            cr_belongs_to_same_activity_of_previous_cr_yes_no_temp = cr_belongs_to_same_activity_of_previous_cr_yes_no_temp
                                        else:
                                            cr_belongs_to_same_activity_of_previous_cr_yes_no_temp  =  daily_plan_sheet.iloc[i]['CR Belongs to Same Activity of Previous CR- Yes/NO']
                        
                                    if (len(str(activity_checker_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Activity Checker']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Activity Checker']).strip() == 'TempNA'):
                                            activity_checker_temp = str(activity_checker_temp)
                                        else:
                                            activity_checker_temp  =  str(daily_plan_sheet.iloc[i]['Activity Checker'])

                                    if (len(str(activity_initiator_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Activity Initiator']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Activity Initiator']).strip() == 'TempNA')):
                                            activity_initiator_temp = str(activity_initiator_temp)
                                        else:
                                            activity_initiator_temp  =  str(daily_plan_sheet.iloc[i]['Activity Initiator'])

                                    if (len(str(impact_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Impact']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Impact']).strip() == 'TempNA')):
                                            impact_temp = str(impact_temp)
                                        else:
                                            impact_temp  =  str(daily_plan_sheet.iloc[i]['Impact'])

                                    if (len(str(planning_status_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Planning Status']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Planning Status']).strip() == 'TempNA')):
                                            planning_status_temp = planning_status_temp
                                        else:
                                            planning_status_temp  =  daily_plan_sheet.iloc[i]['Planning Status']

                                    if (len(str(domain_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Domain']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Domain']).strip() == 'TempNA')):
                                            domain_temp = domain_temp
                                        else:
                                            domain_temp  =  daily_plan_sheet.iloc[i]['Domain']

                                    if (len(str(design_availability_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Design Availability']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Design Availability']).strip() == 'TempNA')):
                                            design_availability_temp  =  str(design_availability_temp)
                                        else:
                                            design_availability_temp  =  str(daily_plan_sheet.iloc[i]['Design Availability'])

                                    if (len(str(technical_validator_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Technical Validator']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Technical Validator']).strip() == 'TempNA')):
                                            technical_validator_temp  =  str(technical_validator_temp)
                                        else:
                                            technical_validator_temp  =  str(daily_plan_sheet.iloc[i]['Technical Validator'])

                                    if (len(str(complexity_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Complexity']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Complexity']).strip() == 'TempNA')):
                                            complexity_temp  =  str(complexity_temp)
                                        else:
                                            complexity_temp  =  str(daily_plan_sheet.iloc[i]['Complexity'])

                                    if (len(str(activity_type_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Activity-Type']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Activity-Type']).strip() == 'TempNA')):
                                            activity_type_temp  =  str(activity_type_temp)
                                        else:
                                            activity_type_temp  =  str(daily_plan_sheet.iloc[i]['Activity-Type'])

                                    if (len(str(total_time_spent_on_planned_crs_mins_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Total Time spent on Planned CRs (Mins)']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Total Time spent on Planned CRs (Mins)']).strip() == 'TempNA')):
                                            total_time_spent_on_planned_crs_mins_temp  =  str(total_time_spent_on_planned_crs_mins_temp)
                                        else:
                                            total_time_spent_on_planned_crs_mins_temp  =  str(daily_plan_sheet.iloc[i]['Total Time spent on Planned CRs (Mins)'])

                                    if (len(str(vendor_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Vendor']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Vendor']).strip() == 'TempNA')):
                                            vendor_temp  =  str(vendor_temp)
                                        else:
                                            vendor_temp  =  str(daily_plan_sheet.iloc[i]['Vendor'])

                                    if (len(str(protocol_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Protocol']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Protocol']).strip() == 'TempNA')):
                                            protocol_temp  =  str(protocol_temp)
                                        else:
                                            protocol_temp  =  str(daily_plan_sheet.iloc[i]['Protocol'])

                                    if (len(str(execution_projection_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Execution Projection']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Execution Projection']).strip() == 'TempNA')):
                                            execution_projection_temp  =  str(execution_projection_temp)
                                        else:
                                            execution_projection_temp  =  str(daily_plan_sheet.iloc[i]['Execution Projection'])
                        
                            elif (count == 1):
                                if (str(daily_plan_sheet.iloc[i]['CR NO']) == cr):
                                    
                                    if ((len(str(daily_plan_sheet.iloc[i]['IMPACTED NODE']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['IMPACTED NODE']).strip() == 'TempNA')):
                                                impacted_node_temp = str(impacted_node_temp)
                                    else:
                                        if (len(str(impacted_node_temp)) == 0):
                                            impacted_node_temp = str(daily_plan_sheet.iloc[i]['IMPACTED NODE'])
                                    
                                    if (len(str(domain_kpi_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Domain kpi']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Domain kpi']).strip() == 'TempNA')):
                                            domain_kpi_temp = str(domain_kpi_temp)
                                        else:
                                            domain_kpi_temp = str(daily_plan_sheet.iloc[i]['Domain kpi']).upper()
                                    
                                    if ((len(str(daily_plan_sheet.iloc[i]['KPI DETAILS']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['KPI DETAILS']).strip() == 'TempNA')):
                                        kpi_details_temp = str(kpi_details_temp)
                                    else:
                                        if (len(str(kpi_details_temp)) == 0):
                                            kpi_details_temp = str(daily_plan_sheet.iloc[i]['KPI DETAILS'])
                                    
                                    if ((len(str(daily_plan_sheet.iloc[i]['oss name']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['oss name']).strip() == 'TempNA')):
                                        oss_name_temp = str(oss_name_temp)
                                    else: 
                                        oss_name_temp  =  str(daily_plan_sheet.iloc[i]['oss name'])
                                    
                                    if ((len(str(daily_plan_sheet.iloc[i]['oss ip']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['oss ip']).strip() == 'TempNA')):
                                        oss_IP_temp = oss_IP_temp
                                    else:
                                        oss_IP_temp  =  daily_plan_sheet.iloc[i]['oss ip']

                                    if (len(str(maintenance_window_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Maintenance Window']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Maintenance Window']).strip() == 'TempNA')):
                                            maintenance_window_temp = str(maintenance_window_temp)
                                        else:
                                            maintenance_window_temp  = str(daily_plan_sheet.iloc[i]['Maintenance Window'])

                                    if(len(str(activity_title_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Activity Title']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Activity Title']).strip() == 'TempNA')):
                                            activity_title_temp = str(activity_title_temp)
                                        else:
                                            activity_title_temp  =  str(daily_plan_sheet.iloc[i]['Activity Title'])

                                    if(len(str(risk_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Risk']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Risk']).strip() == 'TempNA')):
                                            risk_temp  = str(risk_temp)
                                        else:
                                            risk_temp  = str(daily_plan_sheet.iloc[i]['Risk'])

                                    if (len(str(location_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Location']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Location']).strip() == 'TempNA')):
                                            location_temp = str(location_temp)
                                        else:    
                                            location_temp  = str(daily_plan_sheet.iloc[i]['Location'])

                                    if (len(str(circle_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Circle']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Circle']).strip() == 'TempNA')):
                                            circle_temp = str(circle_temp)
                                        else:
                                            circle_temp  = str(daily_plan_sheet.iloc[i]['Circle'])
                                    
                                    if (len(str(no_of_node_involved_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['No. of Node Involved']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['No. of Node Involved']).strip() == 'TempNA')):
                                            no_of_node_involved_temp = str(no_of_node_involved_temp)
                                        else:
                                            no_of_node_involved_temp  =  daily_plan_sheet.iloc[i]['No. of Node Involved']
                                    
                                    if (len(str(cr_belongs_to_same_activity_of_previous_cr_yes_no_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['CR Belongs to Same Activity of Previous CR- Yes/NO']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['CR Belongs to Same Activity of Previous CR- Yes/NO']).strip() == 'TempNA')):
                                            cr_belongs_to_same_activity_of_previous_cr_yes_no_temp = str(cr_belongs_to_same_activity_of_previous_cr_yes_no_temp)
                                        else:
                                            cr_belongs_to_same_activity_of_previous_cr_yes_no_temp  =  str(daily_plan_sheet.iloc[i]['CR Belongs to Same Activity of Previous CR- Yes/NO'])

                                    if (len(str(activity_checker_temp)) == 0):
                                        if (len(str(daily_plan_sheet.iloc[i]['Activity Checker']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Activity Checker']).strip() == 'TempNA'):
                                            activity_checker_temp = str(activity_checker_temp)
                                        else:
                                            activity_checker_temp  =  str(daily_plan_sheet.iloc[i]['Activity Checker'])

                                    if (len(str(activity_initiator_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Activity Initiator']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Activity Initiator']).strip() == 'TempNA')):
                                            activity_initiator_temp = str(activity_initiator_temp)
                                        else:
                                            activity_initiator_temp  = str(daily_plan_sheet.iloc[i]['Activity Initiator'])

                                    if (len(str(impact_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Impact']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Impact']).strip() == 'TempNA')):
                                            impact_temp = str(impact_temp)
                                        else:
                                            impact_temp  = str(daily_plan_sheet.iloc[i]['Impact'])

                                    if (len(str(planning_status_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Planning Status']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Planning Status']).strip() == 'TempNA')):
                                            planning_status_temp = str(planning_status_temp)
                                        else:
                                            planning_status_temp  = str(daily_plan_sheet.iloc[i]['Planning Status'])

                                    if (len(str(domain_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Domain']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Domain']).strip() == 'TempNA')):
                                            domain_temp = str(domain_temp)
                                        else:
                                            domain_temp  = str(daily_plan_sheet.iloc[i]['Domain'])

                                    if (len(str(design_availability_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Design Availability']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Design Availability']).strip() == 'TempNA')):
                                            design_availability_temp  =  str(design_availability_temp)
                                        else:
                                            design_availability_temp  =  str(daily_plan_sheet.iloc[i]['Design Availability'])

                                    if (len(str(technical_validator_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Technical Validator']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Technical Validator']).strip() == 'TempNA')):
                                            technical_validator_temp  =  str(technical_validator_temp)
                                        else:
                                            technical_validator_temp  =  str(daily_plan_sheet.iloc[i]['Technical Validator'])

                                    if (len(str(complexity_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Complexity']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Complexity']).strip() == 'TempNA')):
                                            complexity_temp  =  str(complexity_temp)
                                        else:
                                            complexity_temp  =  str(daily_plan_sheet.iloc[i]['Complexity'])

                                    if (len(activity_type_temp) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Activity-Type']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Activity-Type']).strip() == 'TempNA')):
                                            activity_type_temp  =  str(activity_type_temp)
                                        else:
                                            activity_type_temp  =  str(daily_plan_sheet.iloc[i]['Activity-Type'])

                                    if ((len(str(vendor_temp).strip()) == 0) or (str(vendor_temp).strip() == 'TempNA')):
                                        if (len(str(daily_plan_sheet.iloc[i]['Vendor']).strip()) == 0):
                                            vendor_temp  =  str(vendor_temp)
                                        else:
                                            vendor_temp  =  str(daily_plan_sheet.iloc[i]['Vendor'])

                                    if (len(str(protocol_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Protocol']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Protocol']).strip() == 'TempNA')):
                                            protocol_temp  =  str(protocol_temp)
                                        else:
                                            protocol_temp  =  str(daily_plan_sheet.iloc[i]['Protocol'])

                                    if (len(str(execution_projection_temp)) == 0):
                                        if ((len(str(daily_plan_sheet.iloc[i]['Execution Projection']).strip()) == 0) or (str(daily_plan_sheet.iloc[i]['Execution Projection']).strip() == 'TempNA')):
                                            execution_projection_temp  =  str(execution_projection_temp)
                                        else:
                                            execution_projection_temp  =  str(daily_plan_sheet.iloc[i]['Execution Projection'])


                        # Incrementing the value of counter
                        counter +=  1
                

                # Creating the List for each column data by appending the temp variable data to respective column list of the particular selected CR Number.
                execution_date.append(execution_date_temp)
                maintenance_window.append(maintenance_window_temp)
                cr_no.append(cr_no_temp)
                activity_title.append(activity_title_temp)
                risk.append(risk_temp)
                location.append(location_temp)
                circle.append(circle_temp)
                no_of_node_involved.append(no_of_node_involved_temp)
                cr_belongs_to_same_activity_of_previous_cr_yes_no.append(cr_belongs_to_same_activity_of_previous_cr_yes_no_temp)
                activity_checker.append(activity_checker_temp)
                activity_initiator.append(activity_initiator_temp)
                impact.append(impact_temp)
                planning_status.append(planning_status_temp)
                domain.append(domain_temp)
                design_availability.append(design_availability_temp)
                technical_validator.append(technical_validator_temp)
                complexity.append(complexity_temp)
                activity_type.append(activity_type_temp)
                domain_kpi.append(domain_kpi_temp)
                impacted_node.append(impacted_node_temp)
                kpi_details.append(kpi_details_temp)
                oss_name.append(oss_name_temp)
                oss_IP.append(oss_IP_temp)
                total_time_spent_on_planned_crs_mins.append(total_time_spent_on_planned_crs_mins_temp)
                vendor.append(vendor_temp)
                protocol.append(protocol_temp)
                execution_projection.append(execution_projection_temp)
            
            # Creating the list for the columns to make a pandas dataframe to be written into the excel sheet.
            columns_for_the_email_package = ['Execution Date',
                                            'Maintenance Window',
                                            'CR NO',
                                            'Activity Title',
                                            'Risk',
                                            'Location',
                                            'Circle',
                                            'No. of Node Involved',
                                            'CR Belongs to Same Activity of Previous CR- Yes/NO',
                                            'Change Responsible',
                                            'Activity Checker',
                                            'Activity Initiator',
                                            'Impact',
                                            'Planning Status',
                                            'Domain',
                                            'Final Status',
                                            'Reason For Rollback / Cancel',
                                            'Design Availability',
                                            'Technical Validator',
                                            'Complexity',
                                            'Activity-Type',
                                            'Domain kpi',
                                            'IMPACTED NODE',
                                            'KPI DETAILS',
                                            'oss name',
                                            'oss ip',
                                            'Total Time spent on Planned CRs (Mins)',
                                            'Vendor',
                                            'Protocol',
                                            'Execution Projection',
                                            'Inter-domain Name',
                                            'Second Level Validation Status',
                                            'Inter-domain KPI status',
                                            'MOP View Status']
            
            df = pd.DataFrame(columns = columns_for_the_email_package)
            df['Execution Date']                                        = execution_date
            df['Maintenance Window']                                    = maintenance_window
            df['CR NO']                                                 = cr_no
            df['Activity Title']                                        = activity_title
            df['Risk']                                                  = risk
            df['Location']                                              = location
            df['Circle']                                                = circle
            df['No. of Node Involved']                                  = no_of_node_involved
            df['CR Belongs to Same Activity of Previous CR- Yes/NO']    = cr_belongs_to_same_activity_of_previous_cr_yes_no
            df['Change Responsible']                                    = ''
            df['Activity Checker']                                      = activity_checker
            df['Activity Initiator']                                    = activity_initiator
            df['Impact']                                                = impact
            df['Planning Status']                                       = planning_status
            df['Domain']                                                = domain
            df['Final Status']                                          = ''
            df['Reason For Rollback / Cancel']                          = ''
            df['Design Availability']                                   = design_availability
            df['Technical Validator']                                   = technical_validator
            df['Complexity']                                            = complexity
            df['Activity-Type']                                         = activity_type
            df['Domain kpi']                                            = domain_kpi
            df['IMPACTED NODE']                                         = impacted_node
            df['KPI DETAILS']                                           = kpi_details
            df['oss name']                                              = oss_name
            df['oss ip']                                                = oss_IP
            df['Total Time spent on Planned CRs (Mins)']                = total_time_spent_on_planned_crs_mins
            df['Vendor']                                                = vendor
            df['Protocol']                                              = protocol
            df['Execution Projection']                                  = execution_projection
            df['Inter-domain Name']                                     = ''
            df['Second Level Validation Status']                        = ''
            df['Inter-domain KPI status']                               = ''
            df['MOP View Status']                                       = ''
            
            df.replace('TempNA'," ",inplace = True)
            
            df['Execution Date'] = pd.to_datetime(df['Execution Date'])
            df['Execution Date'] = df['Execution Date'].dt.strftime('%m/%d/%Y')
            
            writer = pd.ExcelWriter(workbook,engine = 'openpyxl',mode = 'a',if_sheet_exists = 'replace')
            new_sheetname = 'Email-Package'
            df.index +=  1
            df.to_excel(writer,sheet_name = new_sheetname,index_label = 'S.NO')
            writer.close()
            
            del writer
            del df
            del daily_plan_sheet
            del wb
            
            # Calling the styling function to stylise the worksheet.
            styling(workbook,new_sheetname)
            validation_adder(workbook,new_sheetname)
            messagebox.showinfo("   Email Package Data Preparation Status",'Email-Package Sheet prepared!')

            # Deleting the variables, once it's use is finished.
            objects = dir()
            for object in objects:
                if not object.startswith("__"):
                    del object
            
            # Returning Successful status
            flag = "Successful"
    
    # Handling Exceptions
    except CustomException:
        # Deleting all the variables before returning the value for "Unsuccessful"
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"

    except PermissionError as e:
        e = str(e).split(":")
        e.remove(e[0])
        e = ':'.join(e)
        messagebox.showerror("  Permission Error!",f"Kindly close {e} as it's open in Excel!")

        # Deleting all the variables before returning the value for "Unsuccessful"
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"
    
    except Exception as error:
        import traceback
        messagebox.showerror("  Exception Occured!",f"{traceback.format_exc()}\n{error}")

        # Deleting all the variables before returning the value for "Unsuccessful"
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"
    
    finally:
        import gc
        # import win32com.client as win32

        # excel = win32.Dispatch("Excel.Application")
        
        # if(len(workbook1) > 0):
        #     excel_workbook = excel.Workbooks.Open(workbook1)
        #     excel_workbook.Close()
        
        # excel.Application.Quit()
        
        gc.collect()
        return flag

# email_package_sheet_creater(r"C:\Users\emaienj\Downloads\New_folder_(3)\MPBN Daily Planning Sheet.xlsx")