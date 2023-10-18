import sys                                                          # Importing the sys to run cmd commands from the script itself.
from openpyxl import load_workbook                                  # Importing load_workbook class from the openpyxl to load existing excel workbook.
from openpyxl.styles import Font,Border,Side,PatternFill,Alignment  # Importing classes from openpyxl to style the excel workbooks.
from openpyxl import Workbook                                       # Importing Workbook to Create Workbook using openpyxl.
from openpyxl.utils import get_column_letter                        # Importing the get_column_letter from openpyxl to convert the column numbers to alphabet letter used in the excel sheet.
import pandas as pd                                                 # Importing Pandas to manipulate the data from the excel sheet.
from datetime import datetime,timedelta                             # Importing datetime and timedelta to get today's maintenance date based on system's current date and time settings.
#from tkinter import *                                               # Importing all the classes from tkinter GUI Module of python.
from tkinter import messagebox                                      # Importing Messagebox to invoke messages where required.
from pathlib import Path                                            # Importing Path from pathlib to check the existence of a file.
from threading import Thread                                        # Importing Thread for creation of threads
import numpy as np                                                  # Importing numpy for operations on numpy arrays obtained from pandas.
import win32com.client as win32
import gc
import os

flag = ""
# workbook1 = ""
# workbook2 = ""
# workbook3 = ""

# # Creating class for Threads with returning value.
# class CustomThread(Thread):
#     def __init__(self, group=None, target=None, name=None, args=(), kwargs={}, Verbose=None):
#         Thread.__init__(self, group, target, name, args, kwargs)
#         self._return = None

#     def run(self):
#         if(self._target is not None):
#             self._return = self._target(*self._args, **self._kwargs)
    
#     def join(self):
#         Thread.join(self)
#         return self._return

# Creating Custom classes to handle custom defined Exceptions (Interuptions) to handle the flow of program.
class TomorrowDataNotFound(Exception):
    def __init__(self,msg):
        self.msg = msg

class CustomException(Exception):
    def __init__(self,title,msg):
        self.title = title
        self.msg = msg
        messagebox.showerror(self.title,self.msg)

def sheet_cleaner(workbook):
    # Loading the Workbook
    wb = load_workbook(workbook)

    # Creating a list of sheets to be removed
    sheets_to_be_removed = ["VAS-Inter Domain","RAN-Inter Domain","CS Core-Inter Domain","PS Core-Inter Domain"]

    # Getting the Sheet names
    sheets = wb.sheetnames

    # Iterating through each sheet to delete the sheets for interdomain.
    for sheet in sheets:
        if(sheet in sheets_to_be_removed):
            del wb[sheet]

    wb.save(workbook)
    wb.close()

    objects = dir()
    for object in objects:
        del object

#####################################################################
#########################  P1 P3 appender  ##########################
#####################################################################

# def p1_sheet_finder_and_loader(workbook_path):
#     outlook = win32.Dispatch("Outlook.Application")
#     mapi = outlook.GetNamespace("MAPI")
#     inbox = mapi.GetDefaultFolder(6)

#     messages = inbox.Items
#     messages.Sort("[ReceivedTime]",True)

#     subject_line_we_are_looking_for = "MPBN Planning Automation Tracker P1_Sheet"
#     subject_line_we_are_looking_for = subject_line_we_are_looking_for.lower()

#     last_date_for_mail_check = datetime.now() - timedelta(days = 3)
#     last_date_for_mail_check = last_date_for_mail_check.replace(hour=0,minute=0,second=0)

#     temp_flag = 0
    
#     for message in messages:
#         try:
#             dt = message.ReceivedTime
#             year,month,day,hour,minute,second = dt.year,dt.month,dt.day,dt.hour,dt.minute,dt.second
#             dt = datetime(year=year,month=month,day=day,hour=hour,minute=minute,second=second)

#             if(dt>=last_date_for_mail_check):
#                 mail_Subject = message.Subject

#                 if(mail_Subject.strip().lower() == subject_line_we_are_looking_for):
#                     temp_flag = 1
#                     attachment = message.Attachements.Item(1)
#                     attachment.SaveAsFile(workbook_path)
#                     break
#         except:
#             continue
    
#     if(temp_flag == 0):
#         folders = len(inbox.Folders)
#         if(folders > 0):
#             for i in range(folders):
#                 folder_messages = inbox.Folders[i].Items
#                 folder_messages.Sort("[ReceivedTime]",True)

#                 for message in folder_messages:
#                     try:
#                         dt = message.ReceivedTime
#                         year,month,day,hour,minute,second = dt.year,dt.month,dt.day,dt.hour,dt.minute,dt.second
#                         dt = datetime(year=year,month=month,day=day,hour=hour,minute=minute,second=second)

#                         if(dt>=last_date_for_mail_check):
#                             mail_Subject = message.Subject

#                             if(mail_Subject.strip().lower() == subject_line_we_are_looking_for):
#                                 temp_flag = 1
#                                 attachment = message.Attachements.Item(1)
#                                 attachment.SaveAsFile(workbook_path)
#                                 break
#                     except:
#                         continue
                
#                 if(temp_flag == 1):
#                     break
        
#     if(temp_flag == 0):
#         folders = len(inbox.Folders)
#         if(folders > 0):
#             for i in range(folders):
#                 sub_folders = len(inbox.Folders[i].Folders)

#                 if(sub_folders > 0):
#                     for j in range(sub_folders):
#                         sub_folder_messages = inbox.Folders[i].Folders[j].Items

#                         sub_folder_messages.Sort("[ReceivedTime]",True)

#                         for message in sub_folder_messages:
#                             try:
#                                 dt = message.ReceivedTime
#                                 year,month,day,hour,minute,second = dt.year,dt.month,dt.day,dt.hour,dt.minute,dt.second
#                                 dt = datetime(year=year,month=month,day=day,hour=hour,minute=minute,second=second)

#                                 if(dt>=last_date_for_mail_check):
#                                     mail_Subject = message.Subject

#                                     if(mail_Subject.strip().lower() == subject_line_we_are_looking_for):
#                                         temp_flag = 1
#                                         attachment = message.Attachements.Item(1)
#                                         attachment.SaveAsFile(workbook_path)
#                                         break
#                             except:
#                                 continue
                        
#                         if(temp_flag == 1):
#                             break
                
#                 if(temp_flag == 1):
#                     break
#     objects = dir()
#     for object in objects:
#         if not object.startswith("__"):
#             del object

def p1_mailer(workbook_path,sender):
    outlook = win32.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.Attachments.Add(workbook_path)
    mail.HTMLBody = f"<html>\
                        <body>\
                                <div>\
                                    <p>Hi Team,</p>\
                                    <p>Kindly find the updated MPBN Planning Automation Tracker File</p>\
                                    </div>\
                                        <br><div>\
                                            Regards,<br>{sender}<br>SRF MPBN | SDU Bharti<br>Ericsson India Global Services Pvt. Ltd.<br>\
                                            </div>\
                            </body>\
                    </html>"
    mail.To = "PDLPBNSRFP@pdl.internal.ericsson.com;"
    mail.Subject = f"MPBN Planning Automation Tracker P1_sheet: {datetime.today().__format__('%d-%b-%Y')}"
    # mail.CC = ""
    mail.Save()
    mail.Display()
    # mail.Send()


def p_one_p_three_appender(sender,workbook):
    # Getting the email-package-sheet
    wb              = pd.ExcelFile(workbook)
    email_package   = pd.read_excel(wb,"Email-Package")
    planning_sheet  = pd.read_excel(wb,"Planning Sheet")

    planning_sheet = planning_sheet[planning_sheet['Planning Status'] == 'Discussed']
    planning_sheet['Execution Date'] = pd.to_datetime(planning_sheet['Execution Date'], format="%d-%b-%Y")
    planning_sheet['Execution Date'] = planning_sheet['Execution Date'].dt.strftime("%m/%d/%Y")
    # print(planning_sheet)
    if(len(planning_sheet) > 0):
        # Concatenating the email package and the remaining rows from planning sheet with planning status 'Discussed'
        try:
            email_package['Execution Date'] = pd.to_datetime(email_package['Execution Date'], format="%d-%b-%Y")
        except:
            try:
                email_package['Execution Date'] = pd.to_datetime(email_package['Execution Date'], format="%m/%d/%Y")
            
            except Exception as e:
                import traceback
                messagebox.showerror("  Exception Occurred!",f"{traceback.format_exc()}\n\n{e}")
                flag = "Unsuccessful"
                return flag

        email_package['Execution Date'] = email_package['Execution Date'].dt.strftime("%m/%d/%Y")
        email_package = pd.concat([email_package,planning_sheet],ignore_index=True)
    
    # Getting the unique technical validator.
    unique_technical_validator_set = email_package['Technical Validator'].dropna().unique()
    # print(unique_technical_validator_set)
    unique_technical_validator_set = unique_technical_validator_set.astype(str)
    # unique_technical_validator_set = np.delete(unique_technical_validator_set,np.where((unique_technical_validator_set == 'nan')|(unique_technical_validator_set == 'Nan')))
    
    ''' 
        If the User is not a technical validator then we are throwing an Exception so that only the respective Technical Validator file 
        gets written out which are present in the Planning Sheet.
    '''
    
    if (sender not in unique_technical_validator_set):
        del planning_sheet
        del email_package
        del wb
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"
        messagebox.showerror(' Technical Validator not Found!','Technical Validator is not found in the Planning Sheet, Kindly Check!')
        return flag
    
    else:
        # Here we are trying to get the parent folder path of the Workbook containing the Email_package sheet.
        p1_workbook_file = os.path.join(os.path.dirname(workbook),"MPBN Planning Automation Tracker P1.xlsx")
        p1_sheet_name = 'MPBN Activity List'
        p1_dataframe = email_package
        

        p1_dataframe.drop("S.NO",axis = "columns",inplace = True)
        p1_columns = p1_dataframe.columns.to_list()

        # Finding out whether the file for MPBN Planning Automation Tracker P1.xlsx exists or not
        # If the File does not exists then in that case the file is created

        if (Path(p1_workbook_file).exists() == False):
            neo_response = messagebox.askokcancel("   P1 workbook doesn't exist!", f"'MPBN Planning Automation Tracker P1 Sheet ' doesn't exist. Download the same from outlook (latest mail) and then press ok!")
            
            if(not neo_response):
                messagebox.showwarning("    User Selected 'Cancel'!","You selected 'Cancel', so, dropping the work here!")
                flag = "Unsuccessful"
                return flag
            # p1_sheet_finder_and_loader(p1_workbook_file)
        
        else:
            response = messagebox.askyesno("P1 workbook already exists!",f"'MPBN Planning Automation Tracker P1' already exists.Hope you have downloaded the same from latest mail. If yes then (click 'Yes') or (click 'No') if you want to drop this task?")

            if(not response):
                # os.remove(p1_workbook_file)
                # p1_sheet_finder_and_loader(p1_workbook_file)
                flag = "Unsuccessful"
                return flag
        
        # Loading the workbook to find the number of rows occupied in the worksheet to continue the S.NO series in that worksheet.
        # print(p1_workbook_file)
        p1_workbook = load_workbook(p1_workbook_file)

        # Changing the Index of the dataframe to start from 1
        p1_dataframe.reset_index(drop = True, inplace = True)
        p1_dataframe.index += (p1_workbook[p1_sheet_name].max_row)
        p1_dataframe.insert(0,'S.NO',p1_dataframe.index)
                
        p1_workbook.close()

        # Reading the Excel file in pandas.
        p1_file_read = pd.ExcelFile(p1_workbook_file)
        p1_file_read = pd.read_excel(p1_file_read,p1_sheet_name)

        #Converting the execution date column values in the email_package to datetime datatype to execute the further operations
        p1_file_read['Execution Date'] = pd.to_datetime(p1_file_read['Execution Date'],format= "%m/%d/%Y")
        p1_file_read['Execution Date'] = p1_file_read['Execution Date'].dt.strftime("%m/%d/%Y")
        
        # Getting the unique Execution Date from the Execution Date Column of the MPBN Planning Automation Tracker
        p1_file_read_unique_execution_date = list(p1_file_read['Execution Date'].unique())
        # print(p1_file_read_unique_execution_date)
                
        # Assigning a Variable to get the today's maintenance date to check whether today's maintenance date's data is present in the MPBN Planning Automation Tracker
        todays_maintenance_date = email_package.iloc[1]['Execution Date']
        todays_maintenance_date = todays_maintenance_date.strftime("%m/%d/%Y")
        # print(todays_maintenance_date)

        ''' 
            In this condition we are trying to check whether today's maintenance date is present in the MPBN Planning Automation Tracker Workbook's 
            MPBN Activity List 
        '''
        if (todays_maintenance_date not in p1_file_read_unique_execution_date):
            p1_dataframe['Execution Date'] = pd.to_datetime(p1_dataframe['Execution Date'],format= "%m/%d/%Y")
            p1_dataframe['Execution Date'] = p1_dataframe['Execution Date'].dt.strftime("%m/%d/%Y")
            writer1 = pd.ExcelWriter(p1_workbook_file, engine = 'openpyxl', mode = 'a', if_sheet_exists = 'overlay')
            p1_dataframe.to_excel(writer1,p1_sheet_name,startrow = p1_workbook[p1_sheet_name].max_row, index = False,index_label = 'S.NO',header = False)
            writer1.close()
            del writer1
            del p1_dataframe
            del p1_file_read
            del p1_workbook
                    
                    
            # Styling the worksheet.
            styling(p1_workbook_file,p1_sheet_name)
                    
            # message showing MPBN Planning Automation Tracker Status is successfully edited.
            messagebox.showinfo("   MPBN Planning Automation Tracker Status",f"All planned CRs for mentioned Validators have been updated in MPBN Planning Automation Tracker!")
                    
            p1_mailer(p1_workbook_file,sender)
            messagebox.showinfo("   Mail Drafted!","Mail for MPBN Planning Automation Tracker has been drafted successfully!")

            objects = dir()
            for object in objects:
                if not object.startswith("__"):
                   del object
                
            
        else:
            del p1_workbook
            # Message showing that the data for today's maintenance date is already present in the MPBN Planning Automation Tracker Status Excel worksheet.
            messagebox.showinfo("   Data already present","Data for today's maintenance date is already present in the MPBN Planning Automation Tracker")

            objects = dir()
            for object in objects:
                if not object.startswith("__"):
                    del object


        flag = 'Successful'
        return flag

            


        


#####################################################################
#############################    Styling   ##########################
#####################################################################

# Method(Function) for styling the worksheet.
def styling(workbook,sheetname):
    wb  =  load_workbook(workbook)                          # loading the workbook.
    ws  =  wb[sheetname]                                    # loading the worksheet.
    font_style  =  Font(color = "FFFFFF",bold = True)       # Setting the font style with color.
    col_widths = []                                         # Empty list for entering the max length of string in each column.

    # Iterating through the row values to find the max length of string in each column in the row and appending it to the col_widths list

    for row_values in ws.iter_rows(values_only = True):
        for j,value in enumerate(row_values):
            if len(col_widths)>j:
                if col_widths[j] < len(str(value)):
                    col_widths[j] = len(str(value))
            else:
                col_widths.insert(j,len(str(value)))

    # Standardising the length of each column in the sheet.

    for i,column_width in enumerate(col_widths,1):
        if column_width <= 47:
            ws.column_dimensions[get_column_letter(i)].width = column_width+3
        else:
            ws.column_dimensions[get_column_letter(i)].width = 50


    # Coloring the headers and alingning the headers text to center both horizontally and vertically.
    for column in range(1,ws.max_column+1):   # ws.max_column returns the total number of columns present
        col = get_column_letter(column)
        color_fill = PatternFill(start_color = '0033CC',end_color = '0033CC',fill_type = 'solid')
        ws[col+'1'].font = font_style
        ws[col+'1'].fill = color_fill
        ws[col+'1'].alignment = Alignment(horizontal = 'center',vertical = 'center')

    # Styling all the occupied cells in the sheet, by adding border to the cells, aligning the text in the center
    
    for row in ws:
        for cell in row:    #type: ignore
            cell.alignment = Alignment(horizontal = 'center',vertical = 'center',wrap_text=False)
            cell.border = Border(top = Side(border_style = 'medium',color = '000000'),bottom = Side(border_style = 'medium',color = '000000'),left = Side(border_style = 'medium',color = '000000'),right = Side(border_style = 'medium',color = '000000'))

    # Saving the workbook with worksheet with all the changes.
    wb.save(workbook)
    wb.close()

    # Deleting all the variables before returning to main method.
    objects = dir()
    for object in objects:
        if not object.startswith("__"):
            del object

# Method(Function) for quitting the application.
def quit(event):
    sys.exit(0)


#####################################################################
#############################  Validation   #########################
#####################################################################

# Method(Function) for adding data validation like vlookup or index match
def validation_adder(workbook,worksheet):
    # Loading the parent workbook and the worksheet in the memory of the computer.
    wb = pd.read_excel(workbook,"Email-Package")

    blank_change_responsible_field = []
    
    wb['Change Responsible'].fillna("TempNA",inplace = True)

    # Checking the Change Responsible for each row to be non-blank.
    for i in range(0,len(wb)):
        if (wb.iloc[i]['Change Responsible'] == 'TempNA'):
            blank_change_responsible_field.append(wb.iloc[i]['S.NO'])
        
    
    if(len(blank_change_responsible_field) > 0):
        del planning_sheet
        del email_package
        del wb
        raise CustomException(" Blank Change Responsible Field!",f"Kindly Enter the Change Responsible detail for S.NO:\n{', '.join(str(num) for num in blank_change_responsible_field)}")
    
    else:
        dictionary_from_cr_to_change_responsible_email_package = dict(zip(wb['CR NO'],wb['Change Responsible']))

        daily_plan_sheet = pd.read_excel(workbook,worksheet)
        
        try:
            daily_plan_sheet['Execution Date'] = pd.to_datetime(daily_plan_sheet['Execution Date'])
        
        except:
            try:
                daily_plan_sheet['Execution Date'] = pd.to_datetime(daily_plan_sheet['Execution Date'], format = "%d-%b-%Y")
            
            except:
                daily_plan_sheet['Execution Date'] = pd.to_datetime(daily_plan_sheet['Execution Date'], format = "%m/%d/%Y")
        
        
        daily_plan_sheet['Execution Date'] = daily_plan_sheet['Execution Date'].dt.strftime('%d-%b-%Y')
        
        daily_plan_sheet['Change Responsible'] = daily_plan_sheet['CR NO'].map(dictionary_from_cr_to_change_responsible_email_package)

        daily_plan_sheet.reset_index(drop = True, inplace = True)
        writer = pd.ExcelWriter(workbook, engine = 'openpyxl', mode = 'a', if_sheet_exists = 'replace')
        daily_plan_sheet.to_excel(writer,sheet_name = worksheet,index = False)
        
        writer.close()
        
        styling(workbook,worksheet)

        # Deleting all the variables before returning to main method.
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object


    

#####################################################################
#############################  Paco_cscore  #########################
#####################################################################

# Driver Method(Function)
def paco_cscore(sender,workbook):   #type:ignore
    try:
        #user = subprocess.getoutput("echo %username%") # finding the Username of the user where the directory of the file is located 

        #workbook = r"C:/Daily/MPBN Daily Planning Sheet.xlsx" # system path from where the program will take the input
        global flag;
        flag_for_planning_sheet = 0
        flag_for_mail_id        = 0
        # global workbook1; workbook1 = workbook
        wb = load_workbook(workbook)
        sheets = wb.sheetnames
        for sheet in sheets:
            if (sheet == 'Planning Sheet'):
                flag_for_planning_sheet = 1
            
            if (sheet == 'Mail Id'):
                flag_for_mail_id = 1

        wb.close()
        
        if(flag_for_planning_sheet == 0):
            del wb
            raise TomorrowDataNotFound("The Planning Sheet is not present! Kindly Check!")
        
        if(flag_for_mail_id == 0):
            del wb
            raise TomorrowDataNotFound("The Mail Id  is empty! Kindly Check!")
        
        # Calling the method to add vlookup to the column of Change Responsible in Planning Sheet
        thread = Thread(target=validation_adder,args=(workbook,"Planning Sheet"))
        thread.daemon = True
        thread.start()
        thread.join()

        daily_plan_sheet = pd.read_excel(workbook,'Planning Sheet')
        tomorrow = datetime.today()+timedelta(1) # getting tomorrow date for data execution
        difference = []
        
        daily_plan_sheet['Execution Date'] = pd.to_datetime(daily_plan_sheet['Execution Date'])
        
        if (len(daily_plan_sheet) == 0):
            del daily_plan_sheet
            del wb
            raise TomorrowDataNotFound("The Planning Sheet is empty! Kindly Check!")

        for i in range(0,len(daily_plan_sheet)):
            if (daily_plan_sheet.iloc[i]['Execution Date'].strftime('%Y-%m-%d') != tomorrow.strftime('%Y-%m-%d')):
                difference.append(str(daily_plan_sheet.iloc[i]['S.NO']))
        
        daily_plan_sheet['Execution Date'] = daily_plan_sheet['Execution Date'].dt.strftime("%m/%d/%Y")
        daily_plan_sheet = daily_plan_sheet[daily_plan_sheet['Execution Date'] == tomorrow.strftime("%m/%d/%Y")]
        
        if (len(daily_plan_sheet) == 0):
            del daily_plan_sheet
            del wb
            raise TomorrowDataNotFound("Data for today's maintenance date is not present in the MPBN Daily Planning Sheet, kindly check!")
        
        if (len(difference) > 0):
            del daily_plan_sheet
            del wb
            raise TomorrowDataNotFound(f"All the CR's present are not of Today's Maintenace Date for S.NO : {', '.join([str(num) for num in difference])}")
        
        else:
            daily_plan_sheet = daily_plan_sheet[daily_plan_sheet['Planning Status'].str.upper() == 'PLANNED']
            Email_ID = pd.read_excel(workbook,"Mail Id")
            
            # Finding the Circles and Change Responsible available in the Mail ID worksheet of the MPBN Daily Planning workbook.
            circle = Email_ID['Circle'].tolist()
            original_change_responsible = Email_ID['Change Responsible'].unique()
            original_change_responsible = original_change_responsible.astype(str)
                
            # Removing the NAN value from the list.
            orignal_change_responsible = np.delete(original_change_responsible,np.where((original_change_responsible == 'nan')|(original_change_responsible == 'Nan')))


            # Creating an empty list and empty dataframe to append the S.NO. of rows with input errors and creating a new dataframe from the daily_plan_sheet dataframe with only required data(rows).
            input_error = []
            result_df = pd.DataFrame()
            
            # Replacing all the blank fields(excel cells) in the dataframe with 'NA'
            daily_plan_sheet.fillna("TempNA",inplace = True)
            
            # Creating empty list to find out the serial numbers of the rows where the Circle input and the Change responsible is not properly entered by the user.
            circle_not_proper = []
            change_responsible_not_proper = []

            # Iterating (Looping) through the daily_plan_sheet dataframe index wise (index given by pandas to each row with data), to find out the serial 
            # numbers of the rows where the Circle input and the Change responsible is not properly entered by the user and any other fields that should be left unblank
            # by the user.
            # print(daily_plan_sheet)
            for i in range(0,len(daily_plan_sheet)):
                # print(daily_plan_sheet.iloc[i]['CR NO'])
                if (daily_plan_sheet.iloc[i]['CR NO'] == "TempNA") or (daily_plan_sheet.iloc[i]['CR NO'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])

                if (not daily_plan_sheet.iloc[i]['Circle'] in circle):
                    circle_not_proper.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                if (not daily_plan_sheet.iloc[i]['Change Responsible'] in original_change_responsible):
                    change_responsible_not_proper.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                if (daily_plan_sheet.iloc[i]['Activity Title'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Activity Title'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                if (daily_plan_sheet.iloc[i]['Circle'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Circle'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                if (daily_plan_sheet.iloc[i]['Risk'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Risk'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                if (daily_plan_sheet.iloc[i]['Location'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Location'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                if (daily_plan_sheet.iloc[i]['Change Responsible'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Change Responsible'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                if (daily_plan_sheet.iloc[i]['Impact'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Impact'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                if (daily_plan_sheet.iloc[i]['Technical Validator'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Technical Validator'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                if (daily_plan_sheet.iloc[i]['Activity-Type'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Activity-Type'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                if (daily_plan_sheet.iloc[i]['Vendor'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Vendor'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                if (daily_plan_sheet.iloc[i]['Protocol'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Protocol'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])
                    
                if (daily_plan_sheet.iloc[i]['Execution Projection'] == 'TempNA') or (daily_plan_sheet.iloc[i]['Execution Projection'] == None):
                    input_error.append(daily_plan_sheet.iloc[i]['S.NO'])

                else:
                    result_df = pd.concat([result_df,daily_plan_sheet.iloc[i].to_frame().T], ignore_index= True)
                    
                    
            
            result_df.drop_duplicates(keep = 'first', inplace= True)
            
            # Deleting the old daily_plan_sheet dataframe which we won't be using, to free up memory space.
            del daily_plan_sheet
            daily_plan_sheet = result_df.copy(deep = True)
            # Deleting the result_df after creating a deep copy of it and assigning variable daily_plan_sheet to that deep copy.
            del result_df
            

            if (len(input_error) > 0):
                # messagebox.showerror("  Input Errors",f"Input Error in Planning Sheet! Check 'Location', 'Circle', 'Change Responsible', 'Impact', 'Vendor', 'Protocol' & 'Execution Projection' for S.NO.: {','.join([str(num) for num in input_error])}")
                flag = 'Unsuccessful'
                del daily_plan_sheet
                del Email_ID
                del wb

                input_error = list(set(input_error))
                # Sorting the input error indices to get the list of input error in ascending order.
                input_error.sort()
                # print(input_error)
                raise CustomException("  Input Errors",f"Input Error in Planning Sheet! Check 'Location', 'Circle', 'Change Responsible', 'Impact', 'Vendor', 'Protocol' & 'Execution Projection' for S.NO.: {','.join([str(num) for num in input_error])}")
            
            if (len(circle_not_proper) > 0):
                # messagebox.showerror("  Circles Errors",f"Input Circles are wrong in Planning Sheet for S.NO. : {','.join([str(num) for num in circle_not_proper])}")
                flag = 'Unsuccessful'
                del daily_plan_sheet
                del Email_ID
                del wb
                # print("circle")
                raise CustomException("  Circles Errors",f"Input Circles are wrong in Planning Sheet for S.NO. : {','.join([str(num) for num in circle_not_proper])}")
            
            if (len(change_responsible_not_proper) > 0):
                # messagebox.showerror("  Change Responsible Errors",f"Input Change Responsible are wrong in Planning Sheet for S.NO.: {','.join([str(num) for num in change_responsible_not_proper])}")
                flag = 'Unsuccessful'
                del daily_plan_sheet
                del Email_ID
                del wb
                # print("change_responsible_not_proper",change_responsible_not_proper)
                raise CustomException("  Change Responsible Errors",f"Input Change Responsible are wrong in Planning Sheet for S.NO.: {','.join([str(num) for num in change_responsible_not_proper])}")
            

            # print("Hello")
            thread = Thread(target = sheet_cleaner,args = (workbook,))
            thread.start()

            sheetname = "PS Core-Inter Domain"
            sheetname2 = "CS Core-Inter Domain"
            sheetname3 = "RAN-Inter Domain"
            sheetname4 = "VAS-Inter Domain"

            category = "MPBN-MS"
            owner_domain = "SRF MPBN"
            team_leader = "Karan Loomba"

            ####################################################### Entering details for ps core or paco circle ###########################################################
            execution_date = []
            maintenance_window = []
            mpbn_cr_no = []
            location = []
            mpbn_change_responsible_executor = []
            validator = []
            impact = []
            circle = []
            mpbn_activity_title = []
            cr_owner_domain = []
            inter_domain = []
            cr_category = []
            impacted_node_details = []
            Kpis_to_be_monitored = []
            # Execution Date	Maintenance Window	MPBN CR NO	CR Category	Impact	Location	Circle	MPBN Activity Title	CR Owner Domain	MPBN Change Responsible	Technical Validator/Team Lead	InterDomain	Impacted Node Details	KPI's to be monitored
            for i in range(0,len(daily_plan_sheet)):
                if ((daily_plan_sheet.iloc[i]['Domain kpi'].upper() == 'PS-CORE') or (daily_plan_sheet.iloc[i]['Domain kpi'].upper() == 'PS') or (daily_plan_sheet.iloc[i]['Domain kpi'].upper() == 'PS_CORE') or (daily_plan_sheet.iloc[i]['Domain kpi'].upper().startswith('PACO')) or (daily_plan_sheet.iloc[i]['Domain kpi'].upper().startswith("PS")) and (daily_plan_sheet.iloc[i]['Planning Status'].upper() == 'PLANNED')):
                    execution_date.append(daily_plan_sheet.iloc[i]['Execution Date'])
                    maintenance_window.append(daily_plan_sheet.iloc[i]['Maintenance Window'])
                    mpbn_cr_no.append(daily_plan_sheet.iloc[i]['CR NO'])
                    cr_category.append(category)
                    impact.append(daily_plan_sheet.iloc[i]['Impact'])
                    location.append(daily_plan_sheet.iloc[i]['Location'])
                    txt = str(daily_plan_sheet.iloc[i]['Circle'])
                    circle.append(txt.upper())
                    mpbn_activity_title.append(daily_plan_sheet.iloc[i]['Activity Title'])
                    cr_owner_domain.append(owner_domain)
                    mpbn_change_responsible_executor.append(daily_plan_sheet.iloc[i]['Change Responsible'])
                    technical_validator = daily_plan_sheet.iloc[i]['Technical Validator']
                    if technical_validator == team_leader:
                        validator.append(team_leader)
                    else:
                        tech_validator_team_leader = technical_validator+"/"+team_leader
                        validator.append(tech_validator_team_leader)
                    inter_domain.append(daily_plan_sheet.iloc[i]['Domain kpi'].upper())
                    impacted_node_details.append(daily_plan_sheet.iloc[i]['IMPACTED NODE'])
                    Kpis_to_be_monitored.append(daily_plan_sheet.iloc[i]['KPI DETAILS'])

            dictionary1 = {'CR':mpbn_cr_no,'Maintenance Window':maintenance_window,'CR Category':cr_category,'Impact':impact,'Location':location,'Circle':circle,'MPBN Activity Title':mpbn_activity_title,'CR Owner Domain':cr_owner_domain,'Change Responsible':mpbn_change_responsible_executor,'Technical Validator/Team Lead':validator,'InterDomain':inter_domain,'Impacted Node Details':impacted_node_details,'KPIs to be monitored':Kpis_to_be_monitored}
            df = pd.DataFrame(dictionary1)
            df.drop_duplicates(subset = 'CR',keep = "first", inplace = True)
            df.replace("TempNA","",inplace = True)
            
            ######################################################### Entering details for Cs core #######################################################################
            execution_date = []
            maintenance_window = []
            mpbn_cr_no = []
            location = []
            mpbn_change_responsible_executor = []
            validator = []
            impact = []
            circle = []
            mpbn_activity_title = []
            cr_owner_domain = []
            inter_domain = []
            cr_category = []
            impacted_node_details = []
            Kpis_to_be_monitored = []
            for i in range(0,len(daily_plan_sheet)):
                if ((daily_plan_sheet.iloc[i]['Domain kpi'].upper().startswith("CS")) or (daily_plan_sheet.iloc[i]['Domain kpi'].upper().startswith("STP")) or (daily_plan_sheet.iloc[i]['Domain kpi'].upper().startswith("CORE")) and (daily_plan_sheet.iloc[i]['Planning Status'].upper() == 'PLANNED')) :
                    execution_date.append((daily_plan_sheet.iloc[i]['Execution Date']))
                    maintenance_window.append(daily_plan_sheet.iloc[i]['Maintenance Window'])
                    mpbn_cr_no.append(daily_plan_sheet.iloc[i]['CR NO'])
                    cr_category.append(category)
                    impact.append(daily_plan_sheet.iloc[i]['Impact'])
                    location.append(daily_plan_sheet.iloc[i]['Location'])
                    txt = str(daily_plan_sheet.iloc[i]['Circle'])
                    circle.append(txt.upper())
                    mpbn_activity_title.append(daily_plan_sheet.iloc[i]['Activity Title'])
                    cr_owner_domain.append(owner_domain)
                    mpbn_change_responsible_executor.append(daily_plan_sheet.iloc[i]['Change Responsible'])
                    technical_validator = daily_plan_sheet.iloc[i]['Technical Validator']
                    if technical_validator == team_leader:
                        validator.append(team_leader)
                    else:
                        tech_validator_team_leader = technical_validator+"/"+team_leader
                        validator.append(tech_validator_team_leader)
                    inter_domain.append(daily_plan_sheet.iloc[i]['Domain kpi'].upper())
                    impacted_node_details.append(daily_plan_sheet.iloc[i]['IMPACTED NODE'])
                    Kpis_to_be_monitored.append(daily_plan_sheet.iloc[i]['KPI DETAILS'])
            dictionary2 = {'CR':mpbn_cr_no,'Maintenance Window':maintenance_window,'CR Category':cr_category,'Impact':impact,'Location':location,'Circle':circle,'MPBN Activity Title':mpbn_activity_title,'CR Owner Domain':cr_owner_domain,'Change Responsible':mpbn_change_responsible_executor,'Technical Validator/Team Lead':validator,'InterDomain':inter_domain,'Impacted Node Details':impacted_node_details,'KPIs to be monitored':Kpis_to_be_monitored}
            df2 = pd.DataFrame(dictionary2)
            df2.drop_duplicates(subset = 'CR',keep = "first", inplace = True)
            df2.replace("TempNA","",inplace = True)

            ##########################################################  Entering details for RAN  ########################################################################
            execution_date = []
            maintenance_window = []
            mpbn_cr_no = []
            location = []
            mpbn_change_responsible_executor = []
            validator = []
            impact = []
            circle = []
            mpbn_activity_title = []
            cr_owner_domain = []
            inter_domain = []
            cr_category = []
            impacted_node_details = []
            Kpis_to_be_monitored = []
            oss_name = []
            oss_IP = []
            # Execution Date	Maintenance Window	MPBN CR NO	CR Category	Impact	Location	Circle	MPBN Activity Title	CR Owner Domain	MPBN Change Responsible	Technical Validator/Team Lead	InterDomain	Impacted Node Details	KPI's to be monitored
            for i in range(0,len(daily_plan_sheet)):
                if ((daily_plan_sheet.iloc[i]['Domain kpi'].upper().startswith("RAN") and (daily_plan_sheet.iloc[i]['Planning Status'].upper() == 'PLANNED'))):
                    execution_date.append(daily_plan_sheet.iloc[i]['Execution Date'])
                    maintenance_window.append(daily_plan_sheet.iloc[i]['Maintenance Window'])
                    mpbn_cr_no.append(daily_plan_sheet.iloc[i]['CR NO'])
                    cr_category.append(category)
                    impact.append(daily_plan_sheet.iloc[i]['Impact'])
                    location.append(daily_plan_sheet.iloc[i]['Location'])
                    txt = str(daily_plan_sheet.iloc[i]['Circle'])
                    circle.append(txt.upper())
                    mpbn_activity_title.append(daily_plan_sheet.iloc[i]['Activity Title'])
                    cr_owner_domain.append(owner_domain)
                    mpbn_change_responsible_executor.append(daily_plan_sheet.iloc[i]['Change Responsible'])
                    technical_validator = daily_plan_sheet.iloc[i]['Technical Validator']
                    if technical_validator == team_leader:
                        validator.append(team_leader)
                    else:
                        tech_validator_team_leader = technical_validator+"/"+team_leader
                        validator.append(tech_validator_team_leader)
                    inter_domain.append(daily_plan_sheet.iloc[i]['Domain kpi'])
                    impacted_node_details.append(daily_plan_sheet.iloc[i]['IMPACTED NODE'])
                    Kpis_to_be_monitored.append(daily_plan_sheet.iloc[i]['KPI DETAILS'])
                    oss_name.append(daily_plan_sheet.iloc[i]['oss name'])
                    oss_IP.append(daily_plan_sheet.iloc[i]['oss ip'])

            dictionary3 = {'CR':mpbn_cr_no,'Maintenance Window':maintenance_window,'CR Category':cr_category,'Impact':impact,'Location':location,'Circle':circle,'MPBN Activity Title':mpbn_activity_title,'CR Owner Domain':cr_owner_domain,'Change Responsible':mpbn_change_responsible_executor,'Technical Validator/Team Lead':validator,'InterDomain':inter_domain,'Impacted Node Details':impacted_node_details,'KPIs to be monitored':Kpis_to_be_monitored,'OSS Name':oss_name,'OSS IP':oss_IP}
            df3 = pd.DataFrame(dictionary3)
            df3.drop_duplicates(subset = 'CR',keep = "first", inplace = True)
            df3.replace("TempNA","",inplace = True)

            ##########################################################  Entering details for VAS  ########################################################################
            
            execution_date = []
            maintenance_window = []
            mpbn_cr_no = []
            location = []
            mpbn_change_responsible_executor = []
            validator = []
            impact = []
            circle = []
            mpbn_activity_title = []
            cr_owner_domain = []
            inter_domain = []
            cr_category = []
            impacted_node_details = []
            Kpis_to_be_monitored = []
            oss_name = []
            oss_IP = []
            for i in range(0,len(daily_plan_sheet)):
                if ((daily_plan_sheet.iloc[i]['Domain kpi'].upper().startswith('VAS')) and (daily_plan_sheet.iloc[i]['Planning Status'].upper() == 'PLANNED')):
                    execution_date.append(daily_plan_sheet.iloc[i]['Execution Date'])
                    maintenance_window.append(daily_plan_sheet.iloc[i]['Maintenance Window'])
                    mpbn_cr_no.append(daily_plan_sheet.iloc[i]['CR NO'])
                    cr_category.append(category)
                    impact.append(daily_plan_sheet.iloc[i]['Impact'])
                    location.append(daily_plan_sheet.iloc[i]['Location'])
                    txt = str(daily_plan_sheet.iloc[i]['Circle'])
                    circle.append(txt.upper())
                    mpbn_activity_title.append(daily_plan_sheet.iloc[i]['Activity Title'])
                    cr_owner_domain.append(owner_domain)
                    mpbn_change_responsible_executor.append(daily_plan_sheet.iloc[i]['Change Responsible'])
                    technical_validator = daily_plan_sheet.iloc[i]['Technical Validator']
                    if technical_validator == team_leader:
                        validator.append(team_leader)
                    else:
                        tech_validator_team_leader = technical_validator+"/"+team_leader
                        validator.append(tech_validator_team_leader)
                    inter_domain.append(daily_plan_sheet.iloc[i]['Domain kpi'].upper())
                    impacted_node_details.append(daily_plan_sheet.iloc[i]['IMPACTED NODE'])
                    Kpis_to_be_monitored.append(daily_plan_sheet.iloc[i]['KPI DETAILS'])

            dictionary4 = {'CR':mpbn_cr_no,'Maintenance Window':maintenance_window,'CR Category':cr_category,'Impact':impact,'Location':location,'Circle':circle,'MPBN Activity Title':mpbn_activity_title,'CR Owner Domain':cr_owner_domain,'Change Responsible':mpbn_change_responsible_executor,'Technical Validator/Team Lead':validator,'InterDomain':inter_domain,'Impacted Node Details':impacted_node_details,'KPIs to be monitored':Kpis_to_be_monitored}
            df4 = pd.DataFrame(dictionary4)
            df4.drop_duplicates(subset = 'CR', keep = "first", inplace = True)
            df4.replace("TempNA","",inplace = True)

            # Dropping the Index of each Dataframe so that they're not written into the excel sheets.
            df.reset_index(drop = True,inplace = True)
            df2.reset_index(drop = True,inplace = True)
            df3.reset_index(drop = True,inplace = True)
            df4.reset_index(drop = True,inplace = True)

            thread.join()

            # writer = pd.ExcelWriter(workbook,engine = 'xlsxwriter')

            # daily_plan_sheet.to_excel(writer,sheet_name = 'Planning Sheet',index = False)
            # df.to_excel(writer,sheet_name = sheetname,index = False)
            # df2.to_excel(writer,sheet_name = sheetname2,index = False)
            # df3.to_excel(writer,sheet_name = sheetname3,index = False)
            # Email_Id.to_excel(writer,sheet_name = 'Mail Id',index = False)
            # print("Hello")
            # Writing the dataframes into the worksheets.
            # Checking whether the interdomain data are present in the sheet or not.
            if((len(df) > 0) or (len(df2) > 0) or (len(df3) > 0) or (len(df4) > 0)):
                writer = pd.ExcelWriter(workbook,engine = "openpyxl",mode = "a",if_sheet_exists = "replace")
                if (len(df) > 0):
                    df.to_excel(writer,sheet_name = sheetname,index = False)
                
                if (len(df2) > 0):
                    df2.to_excel(writer,sheet_name = sheetname2,index = False)
                
                if (len(df3) > 0):
                    df3.to_excel(writer,sheet_name = sheetname3,index = False)
                
                if (len(df4) > 0):
                    df4.to_excel(writer,sheet_name = sheetname4,index = False)

                writer.close()
                del writer
                
                # Styling the worksheets.
                if(len(df)):
                    styling(workbook,sheetname)
                    del df
                
                if(len(df2)):
                    styling(workbook,sheetname2)
                    del df2
                
                if(len(df3)):
                    styling(workbook,sheetname3)
                    del df3
                
                if(len(df4)):
                    styling(workbook,sheetname4)
                    del df4
                
                # Message showing that all the Interdomain Sheets have been written.
                messagebox.showinfo("   Interdomain Data Preparation Status","Interdomain KPIs Mail Data Preparation Task Completed!")
                

                objects = dir()
                for object in objects:
                    if not object.startswith("__"):
                        del object
                
                # Calling the Method(Function) that can write into the Automation tracker sheet.
                flag = p_one_p_three_appender(sender,workbook)
                del daily_plan_sheet
                # print(flag)
                # return flag
            
            else:

                # Message for the case when there's no Interdomain data present(pre-defined interdomains i.e. CS-Core, PS-Core, RAN, VAS, aren't present in the 
                # 'Domain Kpi' column of the sheet)
                response = messagebox.askokcancel("  No Interdomain Data Present!","No Interdomain detected in the 'Domain kpi' column of the uploaded planning sheet, Kindly check! if this is ok, Press Ok!")

                if(response):
                    flag = p_one_p_three_appender(sender,workbook)

                objects = dir()
                for object in objects:
                    if not object.startswith("__"):
                        del object
                
                flag = "Unsuccessful"


    # Exception for condition when Today's maintenance date is not present.
    except TomorrowDataNotFound as error:
        messagebox.showerror("  Data for today's maintenance not found",error)

        # Deleting all the variables before returning to main method.
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"
    
    # Handling Custom Exception
    except CustomException:
        # Deleting all the variables before returning to main method.
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"
    
    #Handling Key Error 
    except KeyError as e:
        messagebox.showerror("  Check for the below Header(KeyError) ",e)

        # Deleting all the variables before returning the value for "Unsuccessful"
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"
    
    #Handling Attribute Error 
    # except AttributeError as e:
    #     messagebox.showerror("  AttributeError Exception Occured",e)

    #     # Deleting all the variables before returning the value for "Unsuccessful"
    #     objects = dir()
    #     for object in objects:
    #         if not object.startswith("__"):
    #             del object
        
    #     flag = "Unsuccessful"
    
    # Handling Exception for permission error for opening/editing Workbook.
    except PermissionError as e:
        e = str(e).split(":")
        e.remove(e[0])
        e = ':'.join(e)
        messagebox.showerror("  Permission Error!",f"Kindly close {e} as it's open in Excel!")

        # Deleting all the variables before returning the value for "Unsuccessful"
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"

    # Handling any other Exception that has not been handled.
    except Exception as e:
        import traceback
        messagebox.showerror("  Exception Occured",f"{traceback.format_exc()}\n{e}")

        # Deleting all the variables before returning the value for "Unsuccessful"
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"
    
    finally:
        # import win32com.client as win32

        # excel = win32.Dispatch("Excel.Application")
        
        # if (len(workbook1) > 0):
        #     wb = excel.Workbooks.Open(workbook1)
        #     wb.Close()
        
        # if (len(workbook2) > 0):
        #     wb = excel.Workbooks.Open(workbook2)
        #     wb.Close()

        # if (len(workbook3) > 0):
        #     wb = excel.Workbooks.Open(workbook3)
        #     wb.Close()

        # excel.Application.Quit()
        
        gc.collect()
        # print(flag)
        return flag

# paco_cscore("Kartar Singh",r"C:/Users/emaienj/Downloads/MPBN Daily Planning Sheet.xlsx")