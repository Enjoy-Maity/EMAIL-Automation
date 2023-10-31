import pandas as pd
import numpy as np
import win32com.client as win32
from tkinter import messagebox
from Custom_Exception import CustomException
from datetime import datetime,timedelta
from openpyxl import load_workbook,Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Side,Border,PatternFill,Alignment
from Custom_Warning import CustomWarning
import os
import traceback

flag = ''

def mail_drafter(path,sender):
    outlook = win32.Dispatch('Outlook.Application')
    message = outlook.CreateItem(0)
    message.To      = 'karan.k.loomba@ericsson.com'
    message.CC      = 'vishal.kumar.garg@ericsson.com;PDLPBNSRFP@pdl.internal.ericsson.com'
    message.Subject = f"SRF-MPBN_Team_Availability_Tracker-{datetime.now().__format__('%Y')}"
    body        = f"<html>\
                        <body>\
                            <div>\
                            <p>Hi Sir,</p>\
                            <p>Please find the attached SRF-MPBN_Team_Availability_Tracker till date.</p>\
                                </div><br><br>\
                            <div>\
                                <p>Regards,<br>\
                                    {sender}<br>\
                                    SRF MPBN | SDU Bharti<br>\
                                    Ericsson India Global Services Pvt. Ltd.</p>\
                                </div>\
                            </body>\
                    </html>"
    message.HTMLBody = body
    message.Attachments.Add(path)
    message.Save()
    message.Display()

def styler(workbook_path, sheetname):
    workbook = load_workbook(workbook_path)
    worksheet = workbook[sheetname]
    font_style = Font(color = '000000', bold=True)
    yellowish_color_for_columns = 'FFE699'
    bluish_color_for_columns    = 'BDD7EE'
    red    = 'FF0000'
    orange = 'FF9933'
    green  = '00FF00'
    yellow = 'FFFF66'
    col_widths  = []
    # print("line 47 hello")
    
    if(sheetname == 'Summary'):
        BDD7EE_bluish_color_columns = ['A','B','C','D']
        # print("inside Summary styling, line 49")
        # Iterating through the row values to find max length of strings in each column, going row-wise
        for row_values in worksheet.iter_rows(values_only= True):
            for j,value in enumerate(row_values):
                value = str(value)
                if(len(col_widths) > j):
                    if(col_widths[j] < len(value)):
                        col_widths[j] = len(value)
                else:
                    col_widths.insert(j,len(value))
        
        # Standardising the length of each column in the sheet.

        i = 0
        while(i < len(col_widths)):
            column_width = col_widths[i]
            if(column_width <= 47):
                worksheet.column_dimensions[get_column_letter(i+1)].width = column_width+3
            
            else:
                worksheet.column_dimensions[get_column_letter(i+1)].width = 50
            i+=1
        
        # Coloring the header and aligning the headers text
        i = 1
        while(i<=worksheet.max_column):
            column_letter   = get_column_letter(i)
            if(column_letter in BDD7EE_bluish_color_columns):
                color_fill      = PatternFill(start_color = bluish_color_for_columns,
                                              end_color=bluish_color_for_columns,
                                              fill_type= 'solid')
            
            else:
                color_fill      = PatternFill(start_color = yellowish_color_for_columns,
                                              end_color=yellowish_color_for_columns,
                                              fill_type= 'solid')
            
            worksheet[f'{column_letter}1'].font = font_style
            worksheet[f'{column_letter}1'].fill = color_fill
            worksheet[f'{column_letter}1'].alignment = Alignment(horizontal='center',
                                                                 vertical='center')
            i+=1
        
        # Styling the occupied cells
        for row in worksheet:
            for cell in row:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center')
                cell.border = Border(top = Side(border_style = 'medium',color = '000000'),
                                     bottom = Side(border_style = 'medium',color = '000000'),
                                     left = Side(border_style = 'medium',color = '000000'),
                                     right = Side(border_style = 'medium',color = '000000'))
        
        # coloring different availability percentages with different color
        i = 1
        while(((5*i)+1) <= worksheet.max_row):
            index= ((5*i)+1)
            j = 5
            while(j <= worksheet.max_column):
                column_letter = get_column_letter(j)
                value = worksheet[f"{column_letter}{index}"].value
                if(isinstance(value,str)):
                    value = float(value.split("%")[0])

                    if(value == 100.00):
                        worksheet[f"{column_letter}{index}"].font = Font(color='000000')
                        worksheet[f"{column_letter}{index}"].fill = PatternFill(start_color= green,
                                                                                end_color=green,
                                                                                fill_type= 'solid')
                    
                    if((value >= 90.00) and (value <= 95.00)):
                        worksheet[f"{column_letter}{index}"].font = Font(color='000000')
                        worksheet[f"{column_letter}{index}"].fill = PatternFill(start_color= orange,
                                                                                end_color=orange,
                                                                                fill_type= 'solid')
                    if(value < 90.00):
                        worksheet[f"{column_letter}{index}"].font = Font(color='FFFFFF',bold=True)
                        worksheet[f"{column_letter}{index}"].fill = PatternFill(start_color= red,
                                                                                end_color=red,
                                                                                fill_type= 'solid')
                j+=1
            i+=1


    if(sheetname == 'Team Availability'):
        BDD7EE_bluish_color_columns = ['A','B','C','D','E','F']
        
        # Iterating through the row values to find max lenth of strings in each column, going row-wise
        for row_values in worksheet.iter_rows(values_only= True):
            for j,value in enumerate(row_values):
                value = str(value)
                if(len(col_widths) > j):
                    if(col_widths[j] < len(value)):
                        col_widths[j] = len(value)
                
                else:
                    col_widths.insert(j,len(value))
        # print(col_widths,'line 116')
        # Standardising the length of each column in the sheet.
        i = 0
        while(i < len(col_widths)):
            column_width = col_widths[i]
            if(column_width <= 47):
                worksheet.column_dimensions[get_column_letter(i+1)].width = column_width+3
            
            else:
                worksheet.column_dimensions[get_column_letter(i+1)].width = 50
            i+=1
        
        # Coloring the header and aligning the headers text
        i = 1
        while(i<=worksheet.max_column):
            column_letter   = get_column_letter(i)
            if(column_letter in BDD7EE_bluish_color_columns):
                color_fill      = PatternFill(start_color = bluish_color_for_columns,
                                              end_color=bluish_color_for_columns,
                                              fill_type= 'solid')
                # print('inside blue_ columns line 136')
            
            else:
                color_fill      = PatternFill(start_color = yellowish_color_for_columns,
                                              end_color=yellowish_color_for_columns,
                                              fill_type= 'solid')
            
            worksheet[f'{column_letter}1'].font = font_style
            worksheet[f'{column_letter}1'].fill = color_fill
            worksheet[f'{column_letter}1'].alignment = Alignment(horizontal='center',
                                                                 vertical='center')
            i+=1
        
        # Styling the occupied cells
        for row in worksheet:
            for cell in row:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center')
                cell.border = Border(top = Side(border_style = 'medium',color = '000000'),
                                     bottom = Side(border_style = 'medium',color = '000000'),
                                     left = Side(border_style = 'medium',color = '000000'),
                                     right = Side(border_style = 'medium',color = '000000'))
        
        # Coloring the cells with value as 'Not Available' red
        i = 2
        while(i <= worksheet.max_row):
            j = 6
            while(j <= worksheet.max_column):
                column_letter = get_column_letter(j)
                if(str(worksheet[f'{column_letter}{i}'].value) == 'Not Available'):
                    worksheet[f'{column_letter}{i}'].font = Font(color= 'FFFFFF', bold=True)
                    worksheet[f'{column_letter}{i}'].fill = PatternFill(start_color= red,
                                                                        end_color= red,
                                                                        fill_type= 'solid')
                j+=1
            i+=1
        
        
        i = 2
        while(i<= worksheet.max_row):
            day_type_value = worksheet[f'E{i}'].value
            leave_count_value = worksheet[f'F{i}'].value
            if((day_type_value == 'Normal') and (leave_count_value > 2)):
                worksheet[f'F{i}'].font = Font(color= red, bold=True)
            
            if(day_type_value == 'Weekend'):
                j = 1
                while(j<=worksheet.max_column):
                    column_letter = get_column_letter(j)
                    worksheet[f'{column_letter}{i}'].fill = PatternFill(start_color= orange,
                                                                        end_color= orange,
                                                                        fill_type= 'solid')
                    j+=1
            
            if(day_type_value == 'Holiday Support'):
                j = 1
                while(j<=worksheet.max_column):
                    column_letter = get_column_letter(j)
                    worksheet[f'{column_letter}{i}'].fill = PatternFill(start_color= yellow,
                                                                        end_color= yellow,
                                                                        fill_type= 'solid')
                    j+=1
            i+=1
        
    
    # Saving the workbook
    workbook.save(workbook_path)
    workbook.close()
    del workbook

    # Deleting all the variables used here
    objects = dir()
    for object in objects:
        del object

def not_available_writer(**kwargs):
    team_availability_sheet = kwargs['team_availability_sheet']
    remaining_change_responsible_that_are_absent = kwargs['remaining_change_responsible']
    index = kwargs['index']
    i =0
    while(i < remaining_change_responsible_that_are_absent.size):
        team_availability_sheet.loc[index,remaining_change_responsible_that_are_absent[i]] = 'Not Available'
        i += 1
    
    # print(team_availability_sheet,'\ninside not available writer \n\n')
    return team_availability_sheet

def available_writer(array_of_resources_involved, team_availability_sheet,index):
    i = 0
    while(i < array_of_resources_involved.size):
        team_availability_sheet.loc[index,array_of_resources_involved[i]] = 'Available'
        i += 1
    # print(team_availability_sheet,'\ninside available writer\n\n')
    return team_availability_sheet

def empty_space_writer(remaining_change_responsible, team_availability_sheet,index):
    i = 0
    while(i < remaining_change_responsible.size):
        team_availability_sheet.loc[index,remaining_change_responsible[i]] = ''
        i+=1
    # print(team_availability_sheet,"\ninside empty space writer\n\n")
    return team_availability_sheet

def data_filler(**kwargs):
    month = kwargs['current_month']
    day = kwargs['day']
    day_type = kwargs['day_type']
    array_of_resources_involved = kwargs['array_of_resources_involved']
    array_of_resources_involved = np.unique(array_of_resources_involved)
    acceptable_change_responsible = kwargs['acceptable_change_responsible']
    team_availability_sheet_required_columns_for_change_responsible_lists = kwargs['team_availability_sheet_required_columns_for_change_responsible_lists']
    
    # Changing the acceptable_change_responsible list into a numpy array
    acceptable_change_responsible = np.array(acceptable_change_responsible)
    acceptable_change_responsible = np.append(acceptable_change_responsible,['Karan Loomba'])
    new_change_responsible_added    = np.setdiff1d(acceptable_change_responsible,team_availability_sheet_required_columns_for_change_responsible_lists)
    change_responsible_exited       = np.setdiff1d(team_availability_sheet_required_columns_for_change_responsible_lists,acceptable_change_responsible)


    # Getting the names that are not present in the email-package and other resources not assigned
    remaining_change_responsible = np.setdiff1d(acceptable_change_responsible,array_of_resources_involved)

    # Checking that the names in remainin_change_responsible are really on leave or not
    if(day_type == 'Normal'):
        new_remaining_change_responsible_list = []
        if(remaining_change_responsible.size > 0):
            loop_index = 0
            while(loop_index < remaining_change_responsible.size):
                response_for_confirmation = messagebox.askyesno("    Resource Availability Confirmation", f"Please confirm if {remaining_change_responsible[loop_index]} is on leave today?")
                
                if(not response_for_confirmation):
                    new_remaining_change_responsible_list.append(remaining_change_responsible[loop_index])
                
                loop_index+=1
        
        new_remaining_change_responsible_list = np.array(new_remaining_change_responsible_list)
        remaining_change_responsible = np.setdiff1d(remaining_change_responsible,new_remaining_change_responsible_list)

        array_of_resources_involved = np.append(array_of_resources_involved,new_remaining_change_responsible_list)

    # print(remaining_change_responsible)
    # Getting the dataframes
    team_availability_sheet = kwargs['team_availability_sheet']
    summary_sheet = kwargs['summary_sheet']

    # print(new_change_responsible_added)
    if(len(team_availability_sheet) > 0):
        if(new_change_responsible_added.size > 0):
            i  = 0
            while(i < new_change_responsible_added.size):
                change_responsible_selected = new_change_responsible_added[i]
                j = 0
                while(j < len(team_availability_sheet)):
                    if((team_availability_sheet.iloc[j]['Leave Count'] != 'Weekend')):
                        team_availability_sheet.loc[j,change_responsible_selected] = 'Not in Team'
                    
                    if((team_availability_sheet.iloc[j]['Leave Count'] == 'Weekend')):
                        team_availability_sheet.loc[j,change_responsible_selected] = 'Weekend'
                    j+=1
                i+=1
        
    
    # filling the team_availability_sheet
    index = len(team_availability_sheet)
    team_availability_sheet.loc[index,'S No'] = index+1
    team_availability_sheet.loc[index,'Month'] = month
    team_availability_sheet.loc[index,'Date'] = datetime.now().strftime("%d-%b-%y")
    team_availability_sheet.loc[index,'Day'] = day
    team_availability_sheet.loc[index,'Day Type'] = day_type
    team_availability_sheet.loc[index,'Leave Count'] = ''
    # print(team_availability_sheet.columns)
    # print(team_availability_sheet, "line 223")
    # print(f"{day_type=}")
    if(day_type == 'Normal'):
        if(remaining_change_responsible.size > 0):
            team_availability_sheet = not_available_writer(remaining_change_responsible = remaining_change_responsible,
                                                           team_availability_sheet = team_availability_sheet,
                                                           index = index)
        # print(remaining_change_responsible)
        team_availability_sheet.loc[index,'Leave Count'] = remaining_change_responsible.size
        # print(team_availability_sheet.iloc[index]['Leave Count'])

        if(array_of_resources_involved.size > 0):
            team_availability_sheet = available_writer(array_of_resources_involved = array_of_resources_involved,
                                                      team_availability_sheet = team_availability_sheet,
                                                      index=index)
        # print("hello line 287")
    if(day_type == 'Weekend'):
        if(remaining_change_responsible.size > 0):
            team_availability_sheet = empty_space_writer(remaining_change_responsible=remaining_change_responsible,
                                                         team_availability_sheet=team_availability_sheet,
                                                         index=index)
        
        if(array_of_resources_involved.size > 0):
            team_availability_sheet = available_writer(array_of_resources_involved= array_of_resources_involved,
                                                       team_availability_sheet=team_availability_sheet,
                                                       index=index)
        
    if(day_type == 'Holiday Support'):
        if(remaining_change_responsible.size > 0):
            # print('inside Holiday Support line 301')
            team_availability_sheet = empty_space_writer(remaining_change_responsible=remaining_change_responsible,
                                                         team_availability_sheet=team_availability_sheet,
                                                         index=index)
            # print('line 305')
        if(array_of_resources_involved.size > 0):
            team_availability_sheet = available_writer(array_of_resources_involved= array_of_resources_involved,
                                                       team_availability_sheet=team_availability_sheet,
                                                       index=index)
    # print(team_availability_sheet,'line 246')

    index = len(team_availability_sheet) - 1
    
    i = 0
    while(i < change_responsible_exited.size):
        team_availability_sheet.loc[index,change_responsible_exited] = 'Not in Team'
        i+=1

    team_availability_sheet['Date'] = pd.to_datetime(team_availability_sheet['Date'],errors='ignore')
    team_availability_sheet['Date'] = team_availability_sheet['Date'].dt.strftime("%d-%b-%y")
    path = kwargs['path']
    writer = pd.ExcelWriter(path=path,
                            engine='openpyxl',
                            mode = 'a',
                            if_sheet_exists='replace')
    # print(,'line 256')
    team_availability_sheet.to_excel(writer,
                sheet_name = 'Team Availability',
                index = False)
    writer.close()
    del writer

    # Styling the worksheet
    styler(workbook_path=path,
           sheetname='Team Availability')
    
    # messagebox.showinfo("    Attendance Workbook Updated!","'Team Availability' worksheet updated!!")

    availability_sheet = team_availability_sheet[team_availability_sheet['Month'] == month]
    normal_working_days_dataframe = availability_sheet[availability_sheet['Day Type'] == 'Normal']
    weekend_working_days_dataframe = availability_sheet[availability_sheet['Day Type'] == 'Weekend']
    holiday_working_days_dataframe = availability_sheet[availability_sheet['Day Type'] == 'Holiday Support']
    total_leave_counts         = normal_working_days_dataframe['Leave Count'].sum()

    # print(holiday_working_days_dataframe,'line 281 ')

    # Getting the length of the respective dataframes
    total_normal_working_days = len(normal_working_days_dataframe)
    total_weekend_working_days = len(weekend_working_days_dataframe)
    total_holiday_working_days = len(holiday_working_days_dataframe)
    

    # Checking if the month name is present in the summary sheet
    month_name_present_in_summary_sheet = False
    
    if(month in summary_sheet['Month'].unique()):
        month_name_present_in_summary_sheet = True
    
    # If month_name_present_in_summary_sheet is False, then creating the rows for the month
    index = len(summary_sheet)
    if(not month_name_present_in_summary_sheet):
        summary_sheet.loc[index,'S No']                    = index+1
        summary_sheet.loc[index,'Month']                   = month
        summary_sheet.loc[index,'Total Count']         = total_normal_working_days
        summary_sheet.loc[index,'Day Type']                = 'Normal'
        
        summary_sheet.loc[index+1,'S No']                    = index+2
        summary_sheet.loc[index+1,'Month']                   = month
        summary_sheet.loc[index+1,'Total Count']         = total_weekend_working_days
        summary_sheet.loc[index+1,'Day Type']                = 'Weekend'
        
        summary_sheet.loc[index+2,'S No']                    = index+3
        summary_sheet.loc[index+2,'Month']                   = month
        summary_sheet.loc[index+2,'Total Count']         = total_holiday_working_days
        summary_sheet.loc[index+2,'Day Type']                = 'Holiday Support'
        
        summary_sheet.loc[index+3,'S No']                    = index+4
        summary_sheet.loc[index+3,'Month']                   = month
        summary_sheet.loc[index+3,'Total Count']         = total_leave_counts
        summary_sheet.loc[index+3,'Day Type']                = 'Leave Count'
        
        summary_sheet.loc[index+4,'S No']                    = index+5
        summary_sheet.loc[index+4,'Month']                   = month
        summary_sheet.loc[index+4,'Total Count']         = total_normal_working_days
        summary_sheet.loc[index+4,'Day Type']                = 'Availability'
    
    if(month_name_present_in_summary_sheet):
        loop_index = 0
        while(loop_index < len(summary_sheet)):
            if(summary_sheet.iloc[loop_index]['Month'] == month):
                break
            else:
                loop_index+=1
        index = loop_index

        summary_sheet.loc[index,'S No']                    = index+1
        summary_sheet.loc[index,'Month']                   = month
        summary_sheet.loc[index,'Total Count']         = total_normal_working_days
        summary_sheet.loc[index,'Day Type']                = 'Normal'
        
        summary_sheet.loc[index+1,'S No']                    = index+2
        summary_sheet.loc[index+1,'Month']                   = month
        summary_sheet.loc[index+1,'Total Count']         = total_weekend_working_days
        summary_sheet.loc[index+1,'Day Type']                = 'Weekend'
        
        summary_sheet.loc[index+2,'S No']                    = index+3
        summary_sheet.loc[index+2,'Month']                   = month
        summary_sheet.loc[index+2,'Total Count']         = total_holiday_working_days
        summary_sheet.loc[index+2,'Day Type']                = 'Holiday Support'
        
        summary_sheet.loc[index+3,'S No']                    = index+4
        summary_sheet.loc[index+3,'Month']                   = month
        summary_sheet.loc[index+3,'Total Count']         = total_leave_counts
        summary_sheet.loc[index+3,'Day Type']                = 'Leave Count'
        
        summary_sheet.loc[index+4,'S No']                    = index+5
        summary_sheet.loc[index+4,'Month']                   = month
        summary_sheet.loc[index+4,'Total Count']         = total_normal_working_days
        summary_sheet.loc[index+4,'Day Type']                = 'Availability'

    # fixing the indices for different month rows
    # length_of_summary_sheet     = len(summary_sheet)
    normal_day_row_index            = index
    weekend_day_row_index           = index+1
    holiday_day_row_index           = index+2
    leave_count_row_index           = index+3
    availability_row_index          = index+4


    # Starting the loop for adding the data in summary sheet
    
    i = 0
    while(i < acceptable_change_responsible.size):
        change_responsible_selected = acceptable_change_responsible[i]
        # print(change_responsible_selected)
        if(not change_responsible_selected in new_change_responsible_added):
            summary_sheet.loc[normal_day_row_index,change_responsible_selected]     = len(normal_working_days_dataframe[normal_working_days_dataframe[change_responsible_selected] == 'Available'])
            # print(f"{summary_sheet.iloc[normal_day_row_index][change_responsible_selected]=}")
        
        if(change_responsible_selected in new_change_responsible_added):
            sum = len(normal_working_days_dataframe[normal_working_days_dataframe[change_responsible_selected] == 'Available']) + len(normal_working_days_dataframe[normal_working_days_dataframe[change_responsible_selected] == 'Not in Team'])
            summary_sheet.loc[normal_day_row_index,change_responsible_selected]     = normal_working_days_dataframe[normal_working_days_dataframe[change_responsible_selected] == 'Available']
        
        summary_sheet.loc[weekend_day_row_index,change_responsible_selected]    = len(weekend_working_days_dataframe[weekend_working_days_dataframe[change_responsible_selected] == 'Available'])
        # print(f"{summary_sheet.iloc[weekend_day_row_index][change_responsible_selected]=}")
        
        summary_sheet.loc[holiday_day_row_index,change_responsible_selected]    = len(holiday_working_days_dataframe[holiday_working_days_dataframe[change_responsible_selected] == 'Available'])
        # print(f"{summary_sheet.iloc[holiday_day_row_index][change_responsible_selected]=}")

        summary_sheet.loc[leave_count_row_index,change_responsible_selected]    = len(normal_working_days_dataframe[normal_working_days_dataframe[change_responsible_selected] == 'Not Available'])
        # print(summary_sheet.iloc[normal_day_row_index][change_responsible_selected])
        # print(total_normal_working_days)

        if((total_normal_working_days > 0) and (summary_sheet.iloc[normal_day_row_index][change_responsible_selected] != np.nan)):
            percentage = (sum/total_normal_working_days)*100
            summary_sheet.loc[availability_row_index,change_responsible_selected]   = f"{round(percentage,2)}%"
            # print(percentage)
        
        else:
            percentage = 0.0
            summary_sheet.loc[availability_row_index,change_responsible_selected]   = f"{round(percentage,2)}%"
        # print(f"{summary_sheet.iloc[availability_row_index][change_responsible_selected]=}")
        
        i+=1
    
    list_of_non_required_columns = ['S No','Month','Total Count','Day Type']
    summary_columns_array = np.array(summary_sheet.columns.tolist())
    summary_columns_array_required = np.setdiff1d(summary_columns_array,list_of_non_required_columns)
    summary_columns_array_required = np.setdiff1d(summary_columns_array_required,acceptable_change_responsible)

    if(summary_columns_array_required.size > 0):
        i = 0
        while(i < summary_columns_array_required.size):
            change_responsible_selected = summary_columns_array_required[i]
            summary_sheet.loc[normal_day_row_index,change_responsible_selected]     = len(normal_working_days_dataframe[normal_working_days_dataframe[change_responsible_selected] == 'Available'])
            # print(f"{summary_sheet.iloc[normal_day_row_index][change_responsible_selected]=}")
            
            summary_sheet.loc[weekend_day_row_index,change_responsible_selected]    = len(weekend_working_days_dataframe[weekend_working_days_dataframe[change_responsible_selected] == 'Available'])
            # print(f"{summary_sheet.iloc[weekend_day_row_index][change_responsible_selected]=}")
            
            summary_sheet.loc[holiday_day_row_index,change_responsible_selected]    = len(holiday_working_days_dataframe[holiday_working_days_dataframe[change_responsible_selected] == 'Available'])
            # print(f"{summary_sheet.iloc[holiday_day_row_index][change_responsible_selected]=}")

            summary_sheet.loc[leave_count_row_index,change_responsible_selected]    = len(normal_working_days_dataframe[normal_working_days_dataframe[change_responsible_selected] == 'Not Available'])
            # print(summary_sheet.iloc[normal_day_row_index][change_responsible_selected])
            # print(total_normal_working_days)

            if((total_normal_working_days > 0) and (summary_sheet.iloc[normal_day_row_index][change_responsible_selected] != np.nan)):
                percentage = (summary_sheet.iloc[normal_day_row_index][change_responsible_selected]/total_normal_working_days)*100
                summary_sheet.loc[availability_row_index,change_responsible_selected]   = f"{round(percentage,2)}%"
                # print(percentage)
            
            else:
                percentage = 0.0
                summary_sheet.loc[availability_row_index,change_responsible_selected]   = f"{round(percentage,2)}%"
            
            i+=1

    writer = pd.ExcelWriter(path=path,
                            engine='openpyxl',
                            mode = 'a',
                            if_sheet_exists='replace')
    # print(summary_sheet,'line 338')
    summary_sheet.to_excel(writer,
                sheet_name = 'Summary',
                index=False)
    writer.close()
    del writer

    # Styling the worksheet
    styler(workbook_path=path,
           sheetname='Summary')
    # messagebox.showinfo("    Attendance Workbook Updated!","'Summary' worksheet updated!!")


def attendance_workbook_creater(**kwargs):
    path = kwargs['path']
    workbook_to_be_created = Workbook()
    workbook_to_be_created.create_sheet(title = 'Summary',index = 0)
    workbook_to_be_created.create_sheet(title= 'Team Availability',index = 1)
    
    accepted_sheet_list = ['Team Availability','Summary']
    sheets = workbook_to_be_created.sheetnames
    
    i = 0
    while(i < len(sheets)):
        selected_sheet = sheets[i]
        if(not selected_sheet in accepted_sheet_list):
            del workbook_to_be_created[selected_sheet]
        i+=1

    columns_for_team_availability = ['S No','Month','Date','Day','Day Type','Leave Count']
    columns_for_team_availability.extend(kwargs['acceptable_change_responsible'])


    columns_for_summary = ['S No','Month','Total Count','Day Type']
    columns_for_summary.extend(kwargs['acceptable_change_responsible'])

    summary_sheet = workbook_to_be_created['Summary']
    team_availability_sheet = workbook_to_be_created['Team Availability']

    i = 1
    while(i <= len(columns_for_summary)):
        summary_sheet[f'{get_column_letter(i)}1'].value = columns_for_summary[i-1]
        i+=1
    
    i = 1
    while(i <= len(columns_for_team_availability)):
        team_availability_sheet[f'{get_column_letter(i)}1'] = columns_for_team_availability[i-1]
        i+=1

    workbook_to_be_created.save(path)
    workbook_to_be_created.close()
    del workbook_to_be_created
    # print('line 368 Hello')
    # print(path)
    # styler(path,'Summary')
    # styler(path,'Team Availability')


def main_function(workbook,**kwargs):
    try:
        global flag;
        if(len(workbook) == 0):
            raise CustomException("    Workbook Not Selected!","Email-Package Workbook not selected ")
        
        # creating a ExcelFile object to read the excel file
        email_package_reader = pd.ExcelFile(workbook,engine= 'openpyxl')
        email_package_reader = pd.read_excel(email_package_reader,'Email-Package')
        
        # Getting the important arrays
        night_shift_lead = kwargs['night_shift_lead']
        buffer_auditor_trainer = kwargs['buffer_auditor_trainer']
        resource_on_automation = kwargs['resource_on_automation']
        acceptable_change_responsible = kwargs['acceptable_change_responsible']     # notepad list
        
        array_of_unique_technical_validator = email_package_reader['Technical Validator'].dropna().unique()

        # deleting the entry for Karan Loomba
        if(not 'Karan Loomba' in array_of_unique_technical_validator):
            acceptable_change_responsible.remove('Karan Loomba')

        acceptable_change_responsible = np.array(acceptable_change_responsible)

        if(str(resource_on_automation).__contains__(',')):
            resource_on_automation = resource_on_automation.split(',')
            resource_on_automation = [resource.strip() for resource in resource_on_automation]

        if(str(buffer_auditor_trainer).__contains__(',')):
            buffer_auditor_trainer = buffer_auditor_trainer(',')
            buffer_auditor_trainer = [resource_on_automation.strip() for resource in buffer_auditor_trainer]
        
        array_of_unique_change_responsible = email_package_reader['Change Responsible'].dropna().unique()
        
        strings_to_be_deleted = ['Select Your Name!','No']
        mask = np.where(~np.in1d(acceptable_change_responsible,strings_to_be_deleted))
        # print(mask)
        acceptable_change_responsible = acceptable_change_responsible[mask]
        # print(acceptable_change_responsible)

        # masks_for_checks_in_acceptable_change_responsible_and_array_of_unique_change_responsible = np.isin(array_of_unique_change_responsible,
        #                                                                                                    acceptable_change_responsible,
        #                                                                                                    assume_unique=True)

        # if(False in masks_for_checks_in_acceptable_change_responsible_and_array_of_unique_change_responsible):
        #     raise CustomException("    Executor Name Missing!",
        #                           f"{', '.join(np.setdiff1d(array_of_unique_change_responsible,acceptable_change_responsible))} executors are not present in your uploaded Change Responsible list text file, Please Check!")
        

        array_of_resources_involved = np.append(array_of_unique_change_responsible,array_of_unique_technical_validator)

        # Appending night shift lead, buffer_auditor_trainer and resource on automation in array_of_resources_involved.
        if(night_shift_lead.upper().strip() != 'NA'):
            array_of_resources_involved = np.append(array_of_resources_involved,np.array([night_shift_lead]))
        
        if((buffer_auditor_trainer.upper().strip() != 'NA')):
            array_of_resources_involved = np.append(array_of_resources_involved,np.array([buffer_auditor_trainer]))
        
        if((not isinstance(resource_on_automation,list)) and (resource_on_automation.upper().strip() != 'NA')):
            array_of_resources_involved = np.append(array_of_resources_involved,np.array([resource_on_automation]))
        
        if(isinstance(resource_on_automation,list)):
            array_of_resources_involved = np.append(array_of_resources_involved,np.array([resource_on_automation]))
        
        if(not kwargs['sender'] in array_of_resources_involved):
            array_of_resources_involved = np.append(array_of_resources_involved,np.array([kwargs['sender']]))
        
        #Checking the day type
        response = messagebox.askyesno("    Day-Type Query","Is today a 'Normal' Working Day?")
        day_type = ''
        if(response):
            day_type = 'Normal'
        
        else:
            neo_response = messagebox.askyesno("    Day-Type Query","Is today a 'Weekend' Day?")
            if (neo_response):
                day_type = 'Weekend'
            
            else:
                neo_response = messagebox.askyesno("    Day-Type Query","Is today a 'Holiday'?")
                if(neo_response):
                    day_type = 'Holiday Support'
                else:
                    raise CustomException("    Wrong Selection","You have made the wrong day-type selection!!")

        # Determining the existence of the attendance workbook.
        attendance_workbook = "SRF_MPBN_Team_Availability.xlsx"

        # Folder where the email-package is located
        dirname = os.path.dirname(workbook)

        # condition for checking that attendance workbook is present or not
        if(not os.path.exists(os.path.join(dirname, attendance_workbook))):
            attendance_workbook_creater(path = os.path.join(dirname,attendance_workbook), 
                                        acceptable_change_responsible = acceptable_change_responsible)
        
        
        attendance_workbook_reader      = pd.ExcelFile(os.path.join(dirname,attendance_workbook))
        team_availability_sheet         = pd.read_excel(attendance_workbook_reader,sheet_name= 'Team Availability')
        
        team_availability_sheet_columns = team_availability_sheet.columns.tolist()
        non_required_columns            = ['S No','Month','Date','Day','Day Type','Leave Count']
        team_availability_sheet_required_columns_for_change_responsible_lists = np.setdiff1d(team_availability_sheet_columns,non_required_columns)
        
        
        if(len(team_availability_sheet) > 0):
            team_availability_sheet['Date'] = pd.to_datetime(team_availability_sheet['Date'],dayfirst=True,format='%d-%b-%y')
            # print(team_availability_sheet.iloc[len(team_availability_sheet)-1]['Date'])
            
            today_date = datetime.now().replace(hour=0,minute=0,second=0,microsecond=0)
            # print(today_date)
            # print(team_availability_sheet.iloc[len(team_availability_sheet)-1]['Date'] == today_date)
            
            if(team_availability_sheet.iloc[len(team_availability_sheet)-1]['Date'] == today_date):
                raise CustomWarning("    Date Already Present","Team Availability Data already present in Tracker for Today's Date!")
            
        # print(team_availability_sheet)
        summary_sheet = pd.read_excel(attendance_workbook_reader,sheet_name='Summary')

        # getting the current month and day
        current_month = str(datetime.now().__format__("%b'%y"))
        day = str(datetime.now().__format__('%A'))

        if(len(team_availability_sheet) > 0):
            if((day == 'Monday') and (team_availability_sheet.iloc[len(team_availability_sheet) - 1]['Day'] == 'Friday')):
                team_availability_sheet['Date'] = pd.to_datetime(team_availability_sheet['Date'],dayfirst=True,format='%d-%b-%y')
                if(not (datetime.now() - timedelta(days =2)).replace(hour = 0, minute = 0,second = 0,microsecond =0) in team_availability_sheet['Date'].unique()):
                    index = len(team_availability_sheet)
                    team_availability_sheet.loc[index,'S No'] = index+1
                    team_availability_sheet.loc[index,'Month'] = str((datetime.now() - timedelta(days = 2)).__format__("%b'%y"))
                    team_availability_sheet['Date'] = team_availability_sheet['Date'].dt.strftime('%d-%b-%y')
                    team_availability_sheet.loc[index,'Date'] = str((datetime.now() - timedelta(days =2)).__format__('%d-%b-%y'))
                    team_availability_sheet.loc[index,'Day'] = str((datetime.now() - timedelta(days =2)).__format__('%A'))
                    team_availability_sheet.loc[index,'Day Type'] = 'Weekend'
                    irregular_columns = {'S No','Month','Date','Day','Day Type'}
                    columns = set(team_availability_sheet.columns.to_list())
                    remaining_columns = (columns - irregular_columns)
                    
                    for i in remaining_columns:
                        team_availability_sheet.loc[index,i] = 'Weekend'

                team_availability_sheet['Date'] = pd.to_datetime(team_availability_sheet['Date'],dayfirst=True,format='%d-%b-%y')
                if(not (datetime.now() - timedelta(days =1)).replace(hour = 0, minute = 0,second = 0,microsecond =0) in team_availability_sheet['Date'].unique()):
                    index = len(team_availability_sheet)
                    team_availability_sheet.loc[index,'S No'] = index+1
                    team_availability_sheet.loc[index,'Month'] = str((datetime.now() - timedelta(days = 1)).__format__("%b'%y"))
                    team_availability_sheet['Date'] = team_availability_sheet['Date'].dt.strftime('%d-%b-%y')
                    team_availability_sheet.loc[index,'Date'] = str((datetime.now() - timedelta(days =1)).__format__('%d-%b-%y'))
                    team_availability_sheet.loc[index,'Day'] = str((datetime.now() - timedelta(days =1)).__format__('%A'))
                    team_availability_sheet.loc[index,'Day Type'] = 'Weekend'

                    for i in remaining_columns:
                        team_availability_sheet.loc[index,i] = 'Weekend'
                
                else:
                    team_availability_sheet['Date'] = team_availability_sheet['Date'].dt.strftime('%d-%b-%y')
        
        # print("Hello, line 450")
        data_filler(current_month=current_month,
                    day=day,
                    team_availability_sheet = team_availability_sheet,
                    summary_sheet = summary_sheet,
                    day_type = day_type,
                    array_of_resources_involved = array_of_resources_involved,
                    acceptable_change_responsible = acceptable_change_responsible,
                    path = os.path.join(dirname,attendance_workbook),
                    team_availability_sheet_required_columns_for_change_responsible_lists = team_availability_sheet_required_columns_for_change_responsible_lists)
        
        mail_drafter(path=os.path.join(dirname,attendance_workbook),
                     sender = kwargs['sender'])
        
        messagebox.showinfo("    Team Availability tracker updated!","SRF_MPBN_Team_Availability tracker has been updated successfully and mail drafted!")

        attendance_workbook_reader.close()
        del email_package_reader
        flag = 'Successful'
        
    except CustomException:
        flag = 'Unsuccessful'
    
    except CustomWarning:
        flag = 'Successful'
    
    except Exception as error:
        messagebox.showerror("    Exception Occurred!",f"{traceback.format_exc()}\n\n{error}")
    
    finally:
        return flag

# main_function(workbook= r"C:/Users/emaienj/Downloads/Daily Work/MPBN_Email_Package_28th Sep 2023.xlsx",
#               night_shift_lead = "Aswini Kumar Behera",
#               buffer_auditor_trainer = "NA",
#               resource_on_automation = "NA",
#               acceptable_change_responsible = ["Select Your Name!",
#                                                "Arka Maiti",
#                                                "Rupesh Mudgil",
#                                                "Karan Loomba",
#                                                "Manoj Kumar",
#                                                "Bharat Ji",
#                                                "Sachin Sharma",
#                                                "Pulluru Sreeramulu",
#                                                "Paras",
#                                                "Ramesh",
#                                                "Pulluru",
#                                                "Kaushal Kumar",
#                                                "Aswini Kumar Behera",
#                                                "Amit Tandon",
#                                                "Kartar Singh",
#                                                "Enjoy Maity",
#                                                "Ashwani Kumar I",
#                                                "Afsar Azizi",
#                                                "Subham Chitranshi",
#                                                "Prakash",
#                                                "No"],
#               sender = "Manoj Kumar")