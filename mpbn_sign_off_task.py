import win32com.client as win32
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import numpy as np
from datetime import datetime, timedelta
from tkinter import messagebox
from threading import Thread

class CustomThread(Thread):
    def __init__(self, group= None, target = None ,name = None, args=(), kwargs = {}, Verbose = None):
        Thread.__init__(self,group, target, name, args, kwargs)
        self._return = None
    
    def run(self):
        if self._target is not None:
            self._return = self._target(*self._args,**self._kwargs)
    
    def join(self):
        Thread.join(self)
        return self._return

class CustomException(Exception):
    def __init__(self,title,msg):
        self.msg = msg
        self.title = title
        super().__init__(self.title, self.msg)
        messagebox.showerror(self.title,self.msg)

def styling(workbook,sheetname):
    wb  =  load_workbook(workbook)
    ws  =  wb[sheetname]
    font_style  =  Font(color = "FFFFFF",bold = True)
    col_widths = []

    for row in ws.iter_rows(values_only = True):
        for j,value in enumerate(row):
            if len(col_widths)>j:
                if col_widths[j] < len(str(value)):
                    col_widths[j] = len(str(value))
            else:
                col_widths.insert(j,len(str(value)))

    for i,column_width in enumerate(col_widths,1):
        if column_width <= 47:
            ws.column_dimensions[get_column_letter(i)].width = column_width+3
        else:
            ws.column_dimensions[get_column_letter(i)].width = 50


    for column in range(1,ws.max_column+1):   # ws.max_column returns the total number of columns present
        col = get_column_letter(column)
        color_fill = PatternFill(start_color = '0033CC',end_color = '0033CC',fill_type = 'solid')
        ws[col+'1'].font = font_style
        ws[col+'1'].fill = color_fill
        ws[col+'1'].alignment = Alignment(horizontal = 'center',vertical = 'center')

    border = Border(top = Side(border_style = 'thick',color = '000000'),bottom = Side(border_style = 'thick',color = '000000'),left = Side(border_style = 'thick',color = '000000'),right = Side(border_style = 'thick',color = '000000'))

    for row in ws:
        for cell in row:
            cell.alignment = Alignment(horizontal = 'center',vertical = 'center',wrap_text=True)
            cell.border = Border(top = Side(border_style = 'medium',color = '000000'),bottom = Side(border_style = 'medium',color = '000000'),left = Side(border_style = 'medium',color = '000000'),right = Side(border_style = 'medium',color = '000000'))

    #rows = ws.max_row
    
    
    wb.save(workbook)

def email_package_writer(workbook,mail_bodies,email_package):
    cr_row_index_dictionary_in_email_package = dict()
    CRs_in_email_package = email_package['CR NO'].to_list()
    for row in range(0,len(email_package)):
        cr_row_index_dictionary_in_email_package[email_package.at[row,'CR NO']] = row
    
    for row in range(0,len(mail_bodies)):
        cr = mail_bodies.at[row,'CR No']
        if (mail_bodies.at[row,'CR No'] in CRs_in_email_package):
            index = int(cr_row_index_dictionary_in_email_package[cr])
            email_package.at[index,'MOP View Status'] = mail_bodies.at[row,'MOP View Status']
            email_package.at[index,'Final Status'] = mail_bodies.at[row,'Final Status  - Completed / Cancelled / Rollback']
            email_package.at[index,'Reason For Rollback / Cancel'] = mail_bodies.at[row,'Reason For Cancellation / Revert']
            email_package.at[index,'Inter-domain Name'] = mail_bodies.at[row,'Inter-Domain Name - If Applicable']
            email_package.at[index,'Second Level Validation Status'] = mail_bodies.at[row,'Second Level Validation Status']
            email_package.at[index,'Inter-domain KPI status'] = mail_bodies.at[row,'Inter-Domain KPIs confirmation taken from all respective domains by executor - If Applicable']
    
    email_package['Execution Date'] = pd.to_datetime(email_package['Execution Date'], format = "%m/%d/%Y")
    email_package = email_package[['S.NO','Execution Date','Maintenance Window','CR NO','Activity Title','Risk','Location','Circle','No. of Node Involved','CR Belongs to Same Activity of Previous CR- Yes/NO','Change Responsible','Activity Checker','Activity Initiator','Impact','Planning Status','Domain','Final Status','Reason For Rollback / Cancel','Design Availability','Technical Validator','Complexity','Activity-Type','Domain kpi','IMPACTED NODE','KPI DETAILS','oss name','oss ip','Total Time spent on Planned CRs (Mins)','Vendor','Protocol','Execution Projection','Inter-domain Name','Second Level Validation Status','Inter-domain KPI status','MOP View Status']]
    email_package['Execution Date'] = email_package['Execution Date'].dt.strftime("%m/%d/%Y")
    email_package['Second Level Validation Status'] = email_package['Second Level Validation Status'].fillna("NA")
    email_package.reset_index(drop = True, inplace = True)
    writer = pd.ExcelWriter(workbook,engine='openpyxl',mode='a',if_sheet_exists='replace')
    email_package.to_excel(writer,"Email-Package",index = False)
    writer.close()
        
    styling(workbook,"Email-Package")

def sendmail(dataframe):
    today = datetime.now()
    today = today.strftime("%m/%d/%Y")
    
    dataframe.fillna(" ")
    outlook_mailer=win32.Dispatch('Outlook.Application')
    msg=outlook_mailer.CreateItem(0)
    html_body="""
    <html>
        <body>
            <div>
                <p>Hi team,<br></p>

                <p>Please find below status of last night planned activities.</p>
            </div>
            <div>
            {}
            </div>
    </body>
    </html>
    """
    sender = "Enjoy Maity"
    msg.Subject=f"Team Consolidate Sign-Off_{today}"
    msg.To= "enjoy.maity@ericsson.com"
    msg.CC= "karan.k.loomba@ericsson.com"
    dataframe=dataframe.style.set_table_styles([
        {'selector':'th','props':'border:1px solid black; border-collapse : collapse; color:white;padding: 10px; background-color:rgb(0, 51, 204);text-align:center;'},
        {'selector':'tr','props':'border:1px solid black; border-collapse : collapse;padding: 10px;text-align:center;'},
        {'selector':'td','props':'border:1px solid black; border-collapse : collapse;padding: 10px;text-align:center;'},
        {'selector':'tr:nth-child(even)','props':'border:1px solid black; border-collapse : collapse;padding: 10px;text-align:center;'}])
    dataframe=dataframe.hide(axis='index')
    msg.HTMLBody=html_body.format(dataframe.to_html(classes = 'table table-stripped',index=False),sender)
    msg.Save()
    msg.Send()


def dfizer(body):
    if (isinstance(body,list)):
        body = body[0]
        columns = body.columns
        
        if (np.dtype(columns) == "int64"):
            new_body = pd.DataFrame(body.values[1:],columns = body.iloc[0])
            new_body.reset_index(drop = True, inplace = True)
            del body
            del columns
            return new_body
        
        else:
            return body
    elif (isinstance(body,pd.DataFrame) is True):
        columns = body.columns
        
        if (np.dtype(columns) == "int64"):
            new_body = pd.DataFrame(body.values[1:],columns = body.iloc[0])
            new_body.reset_index(drop = True, inplace = True)
            del body
            del columns
            return new_body
        
        else:
            return body

def remainder_change_responsible_dict_getter(remainder_cr,email_package):
    remainder_change_responsible = dict()
    
    for row in range(0,len(email_package)):
        if ((email_package.at[row,'CR NO'] in remainder_cr) or (email_package.at[row,'CR No'] in remainder_cr)):
            if (email_package.at[row,'Change Responsible'] not in remainder_change_responsible):
                remainder_change_responsible[email_package.at[row,'Change Responsible']] = []
                remainder_change_responsible[email_package.at[row,'Change Responsible']].append(email_package.at[row,'CR NO'])
            else:
                remainder_change_responsible[email_package.at[row,'Change Responsible']].append(email_package.at[row,'CR NO'])

    return remainder_change_responsible

def reversed_remainder_change_responsible_dict_getter(remainder_cr,email_package):
    remainder_change_responsible = dict()
    
    for row in range(0,len(email_package)):
        if (email_package.at[row,'CR No'] in remainder_cr):
            if (email_package.at[row,'Change Responsible'] not in remainder_change_responsible):
                remainder_change_responsible[email_package.at[row,'Change Responsible']] = []
                remainder_change_responsible[email_package.at[row,'Change Responsible']].append(email_package.at[row,'CR No'])
            else:
                remainder_change_responsible[email_package.at[row,'Change Responsible']].append(email_package.at[row,'CR No'])

    return remainder_change_responsible

def mail_bodies_generator():
    outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")    # MAPI is an API for messaging to do functions like fetching, and manipulation of mails in outlook
    inbox = outlook.GetDefaultFolder(6)
    flag_for_search = 0
    today = datetime.now()
    yesterday = today - timedelta(days = 1)
    yesterday = yesterday.replace(hour=23, minute=0, second=0).strftime('%Y-%m-%d %H:%M %p')
    today = today.strftime("%d-%m-%Y")
    messages = inbox.Items
    messages = messages.Restrict("[ReceivedTime] >= '"+ yesterday + "'")
    subject_we_are_looking_for = f"MPBN Activity Validation Sign Off"
    subject_we_are_looking_for = subject_we_are_looking_for.lower()
    mail_bodies = pd.DataFrame()
    
    for message in messages:
        mail_subject = message.Subject
        
        if mail_subject.lower().__contains__(subject_we_are_looking_for):
            flag_for_search = 1
            body = message.HTMLBody
            body = pd.read_html(body)
            body = dfizer(body)
            mail_bodies =  pd.concat([mail_bodies,body],ignore_index=True)

    if (flag_for_search == 0):
        messages = inbox.Folders["Karan Loomba Sir"].Items
        messages = messages.Restrict("[ReceivedTime] >= '" + yesterday + "'")
        for message in messages:
            mail_subject = message.Subject

            if (mail_subject.lower().__contains__(subject_we_are_looking_for)):
                flag_for_search = 1
                body = message.HTMLBody
                body = pd.read_html(body)
                body = dfizer(body)
                body.loc[:,'Execution Date'] = today
                mail_bodies =  pd.concat([mail_bodies,body],ignore_index=True)
                
    
    if (flag_for_search == 0):
            raise CustomException(" Sign-Off Mails not Found"," User sign off mails not present in mailbox.")
    
    if (flag_for_search == 1):
        mail_bodies.drop_duplicates(subset="CR No",keep = "first", inplace = True)
    return mail_bodies
    
def mpbn_signoff_main_task(workbook,required_worksheet):
    try:
        mail_bodies = mail_bodies_generator()
        email_package = pd.read_excel(workbook,required_worksheet)
        mail_bodies_change_responsible_set = set(mail_bodies['Change Responsible'].unique())
        email_package_change_responsible_set = set(email_package['Change Responsible'].unique())

        remaining_change_responsible = list(email_package_change_responsible_set - mail_bodies_change_responsible_set)
        remaining_change_responsible.sort()
        
        if (len(remaining_change_responsible) > 0):
            raise CustomException(" Signoff Missing!",f" Signoff for below User/Users Missing :/n{', '.join(remaining_change_responsible)}")
        
        else:
            email_package_cr_list = set(email_package['CR NO'].to_list())
            mail_bodies_cr_list = set(mail_bodies['CR No'].to_list())
            remainder_cr = list(email_package_cr_list - mail_bodies_cr_list)
            reversed_remainder_cr = list(mail_bodies_cr_list - email_package_cr_list)

            if (len(remainder_cr) > 0):
                
                thread_for_remainder_change_responsible_dict_getter = CustomThread(target = remainder_change_responsible_dict_getter,args = (remainder_cr,email_package))
                thread_for_remainder_change_responsible_dict_getter.daemon = True
                thread_for_remainder_change_responsible_dict_getter.start()
                remainder_change_responsible_dict = thread_for_remainder_change_responsible_dict_getter.join()

                temp_str = ""

                for change_responsible_key,cr_list in remainder_change_responsible_dict.items():
                    temp_str += f"{change_responsible_key} : {', '.join(cr_list)}\n"

                thread_for_writer = Thread(target = email_package_writer, args = (workbook,mail_bodies,email_package))
                thread_for_writer.daemon = True
                thread_for_writer.start()
                thread_for_writer.join()

                messagebox.showwarning("    Sign Off Missing!",f"The Sign Off is missing for Change Responsible/s are given below:\n{temp_str}")
                return "Unsuccessful"
            
            
            if (len(remainder_cr) == 0):
                if (len(reversed_remainder_cr) > 0):
                    thread_for_writer = Thread(target = email_package_writer, args = (workbook,mail_bodies,email_package))
                    thread_for_writer.daemon = True
                    thread_for_writer.start()
                    thread_for_writer.join()

                    temp_str = ""
                    
                    thread_for_reversed_remainder_change_responsible_dict_getter_for_mail_bodies = CustomThread(target = reversed_remainder_change_responsible_dict_getter, args = (reversed_remainder_cr,mail_bodies))
                    thread_for_reversed_remainder_change_responsible_dict_getter_for_mail_bodies.start()
                    reversed_change_responsible_to_cr_dict = thread_for_reversed_remainder_change_responsible_dict_getter_for_mail_bodies.join()
                    
                    for change_responsible_key,cr_list in reversed_change_responsible_to_cr_dict.items():
                        temp_str += f"{change_responsible_key} : {', '.join(cr_list)}\n\n"

                    messagebox.showwarning("    Extra CR Encountered!", f"Below mentioned CR Details are missing in Planning Sheet, Kindly Check!\n\n{temp_str}")
                    return"Unsuccessful"
                
                if (len(reversed_remainder_cr) == 0):
                    mail_bodies.reset_index(drop = True, inplace = True)
                    email_package_writer(workbook,mail_bodies,email_package)
                    messagebox.showinfo("   MPBN Sign Off Status","Sign Off Status has been updated successfully in Email Package for all tonight CRs!")
                    mail_status = messagebox.askyesno("   Consolidate Sign Off Mail Confirmation","Do you want to proceed for consolidate sign off mail communication?")
                    
                    if (mail_status):
                        mail_bodies['Execution Date'] = pd.to_datetime(mail_bodies["Execution Date"])
                        mail_bodies['Execution Date'] = mail_bodies['Execution Date'].dt.strftime("%d-%m-%Y")
                        mail_bodies['Second Level Validation Status'] = mail_bodies['Second Level Validation Status'].fillna("NA")
                        mail_bodies.fillna(" ",inplace = True)
                        mail_bodies.drop(['S.No.'],axis = 1,inplace=True)

                        mail_bodies.index += 1
                        mail_bodies.insert(0,'S.NO', mail_bodies.index)
                        sendmail(mail_bodies)
                        messagebox.showinfo("   Mail Successfully Sent","   Consolidate Sign Off Mail Communication Successful!")
                    else:
                        pass

                    return "Successful"


                
    except CustomException:
        return "Unsuccessful"
    
def mpbn_signoff(workbook):
    try:
        if(len(workbook) == 0):
            raise CustomException(" File Not Selected"," Please Select the MPBN Planning Workbbok first!")
        
        else:
            workbook = pd.ExcelFile(workbook)
            worksheets_in_workbook = workbook.sheet_names
            required_worksheet = ""
            for worksheet in worksheets_in_workbook:
                if (worksheet == "Email-Package"):
                    if (len(worksheet) > 0):
                        required_worksheet = worksheet
                        break
            
            if (len(required_worksheet) == 0):
                raise CustomException(' Email-Package Worksheet not Present','Kindly Click the Button for Interdomain Kpi Data Prep First!')
            
            status = mpbn_signoff_main_task(workbook,required_worksheet)
            return status

    except CustomException:
        return "Unsuccessful"

    # except Exception as e:
    #     messagebox.showerror("  Exception Occured",e)
    #     return "Unsuccessful"

#mpbn_signoff(r"C:\Daily\MPBN Daily Planning Sheet new copy.xlsx")