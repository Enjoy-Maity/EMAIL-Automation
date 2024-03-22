import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Side, Border
# from openpyxl.writer import save_workbook
from openpyxl.utils import get_column_letter
import os
import numpy as np
import win32com.client as win32
from tkinter import messagebox
from pathlib import Path
from datetime import datetime, timedelta
from Custom_Exception import CustomException

flag = ''
dictionary = dict()

class CustomWarning(Exception):
    def __init__(self, title, message):
        self.title = title
        self.message = message
        super().__init__(self.title, self.message)
        messagebox.showwarning(self.title, self.message)

def suffix_adder(date_for_suffix):
    suffix = ''
    if((int(date_for_suffix.__format__('%d')) <= 10) or (int(date_for_suffix.__format__('%d')) >= 20)):
        match int(date_for_suffix.__format__('%d'))%10:
            case 1 : suffix = 'st'
            case 2 : suffix = 'nd'
            case 3 : suffix = 'rd'
            case _ : suffix = 'th'
    else:
        suffix = 'th'
    return suffix


def email_parser(body):
    new_body_list = body.splitlines()
    result    = [[]]
    to        = []
    cc        = []
    from_mail = ""

    i = 0
    while(i < len(new_body_list)):
        new_body_list[i] = new_body_list[i].strip()

        if (new_body_list[i].startswith("From:")):
            from_mail = new_body_list[i].split(":")[1].split("<")[1].strip(">")

        if (new_body_list[i].startswith("To")):
            to = new_body_list[i].split(":")[1].split(">;")

        if (new_body_list[i].startswith("Cc:")):
            cc = new_body_list[i].split(":")[1].split(">;")

        if (new_body_list[i].startswith("Subject")):
            break
        i += 1

    i = 0
    while(i < len(to)):
        if to[i].__contains__("<"):
            to[i] = (to[i].split("<")[1]).strip()

        if to[i].__contains__(">"):
            to[i] = (to[i].split(">")[0]).strip()

        i += 1

    i = 0
    while(i < len(cc)):
        if cc[i].__contains__("<"):
            cc[i] = (cc[i].split("<")[1]).strip()

        if cc[i].__contains__(">"):
            cc[i] = (cc[i].split(">")[0]).strip()
        i += 1

    to.append(from_mail)
    result = [to,cc]

    del i
    del to
    del cc
    del new_body_list

    return result


def table_creater(dictionary: dict):
    # # 'Date' : datetime.now().strftime('%d/%m/%Y'),
    #                                 'Domain': 'SRF MPBN',
    #                                 'Total CR': total_Crs,
    #                                 'Total Picked CR': total_picked_Crs,
    #                                 'Total CR executed': total_crs_executed,
    #                                 'Cisco': dictionary_for_vendor_row_number['Cisco'],
    #                                 'Nokia': dictionary_for_vendor_row_number['Nokia'],
    #                                 'Ericsson': dictionary_for_vendor_row_number['Ericsson'],
    #                                 'Huawei': dictionary_for_vendor_row_number['Huawei'],
    #                                 'Cisco SDN': input_details_from_user['number_of_cisco_sdn_nodes'],
    #                                 'Nokia SDN': input_details_from_user['number_of_nokia_sdn_nodes'],
    #                                 'Others': dictionary_for_vendor_row_number['Extreme'],
    #                                 'Total Nodes': total_nodes,
    #                                 'CR cancelled due to other reason': total_crs_cancelled,
    #                                 'MPBN pre-post check performed': input_details_from_user['mpbn_pre_post_check_performed'],
    #                                 'Inter-domain KPI monitored': input_details_from_user['interdomain_kpi_monitored'],
    #                                 'Deviation found in KPI': input_details_from_user['deviation_found_in_kpi'],
    #                                 'Rollback': total_crs_rollback,
    #                                 'CR executed through automation': cr_executed_through_automation,
    #                                 'Night Executors Count' : unique_change_reponsible,
    #                                 'Day planners count': day_planners_count,
    #                                 'Resources on comp-off': '',
    #                                 'Resources on leaves': '',
    #                                 'Automation Support': 'NA'
    table = ("<table style = 'width:100%;'>" +
                "<tr style = 'background-color: rgb(255,255,0); text-align: center; width: 100%; border : 1px solid black; border-collapse : collapse;'>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Date</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Domain</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Total Picked CR</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Total CR</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Total CR executed</th>" +
                    "<th rowspan=1 colspan=10 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height: 50px; border: 2px solid black; border-collapse: collapse;'>Node Touches</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>CR cancelled in technical validation</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>CR cancelled due to other reason</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>MPBN pre-post check performed</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Inter-domain KPI monitored</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Deviation found in KPI</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Rollback</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>CR executed through automation</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Night executors count</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Day planners count</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Resources on comp-off</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Resources on leaves</th>" +
                    "<th rowspan=2 style='backround-color: rgb(255,255,0); text-align: center; width:auto; height:auto; border: 2px solid black; border-collapse: collapse;'>Automation Support</th>" +
                "</tr>" +
                "<tr style = 'background-color: rgb(255,255,0); text-align: center; width: 100%; height: auto; border : 1px solid black; border-collapse : collapse;'>" +
                    "<th rowspan=1 style='backround-color: rgb(255,255,0); text-align: center; width:auto; border: 2px solid black; border-collapse: collapse;'>Cisco</th>" +
                    "<th rowspan=1 style='backround-color: rgb(255,255,0); text-align: center; width:auto; border: 2px solid black; border-collapse: collapse;'>Nokia</th>" +
                    "<th rowspan=1 style='backround-color: rgb(255,255,0); text-align: center; width:auto; border: 2px solid black; border-collapse: collapse;'>Ericsson</th>" +
                    "<th rowspan=1 style='backround-color: rgb(255,255,0); text-align: center; width:auto; border: 2px solid black; border-collapse: collapse;'>Extreme/NWIE</th>" +
                    "<th rowspan=1 style='backround-color: rgb(255,255,0); text-align: center; width:auto; border: 2px solid black; border-collapse: collapse;'>Huawei</th>" +
                    "<th rowspan=1 style='backround-color: rgb(255,255,0); text-align: center; width:auto; border: 2px solid black; border-collapse: collapse;'>Cisco SDN</th>" +
                    "<th rowspan=1 style='backround-color: rgb(255,255,0); text-align: center; width:auto; border: 2px solid black; border-collapse: collapse;'>Nokia SDN</th>" +
                    "<th rowspan=1 style='backround-color: rgb(255,255,0); text-align: center; width:auto; border: 2px solid black; border-collapse: collapse;'>Nokia-IXR</th>" +
                    "<th rowspan=1 style='backround-color: rgb(255,255,0); text-align: center; width:auto; border: 2px solid black; border-collapse: collapse;'>Others</th>" +
                    "<th rowspan=1 style='backround-color: rgb(255,255,0); text-align: center; width:auto; border: 2px solid black; border-collapse: collapse;'>Total Nodes</th>" +
                "</tr>" +
                "<tr style = 'background-color: rgb(255, 255, 255); text-align: center; width: 100%; height: auto; border : 1px solid black; border-collapse : collapse;'>"
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'>{(datetime.now() + timedelta(days=1)).strftime('%d-%m-%Y')}</td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'>{dictionary['Domain']}</td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'>{dictionary['Total Picked CR']}</td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'>{dictionary['Total Planned CR']}</td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'></td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'>{dictionary['Night executors count']}</td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'>{dictionary['Day Planners count']}</td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'>{dictionary['Resources on comp-off']}</td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'>{dictionary['Resources on leaves']}</td>" +
                    f"<td rowspan=1 style='backround-color: rgb(255, 255, 255); text-align: center; width:auto;  height: auto; border: 2px solid black; border-collapse: collapse;'>NA</td>" +
                "</tr>" +
            "</table>")

    return table


def srf_mpbn_dashboard_tracker_file_writer(path: str):
    global dictionary

    # if os.path.exists(path):
    #     os.remove(path)
    #     from time import sleep
    #     sleep(2)

    wkbk = load_workbook(path)
    wkbk.active = wkbk['DASHBOARD']
    worksheet = wkbk['DASHBOARD']

    alignment_var = Alignment(horizontal = 'center', vertical = 'center')
    border = Border(left=Side(border_style = 'thin', color = '000000'),
                    right=Side(border_style = 'thin', color = '000000'),
                    top=Side(border_style = 'thin', color = '000000'),
                    bottom=Side(border_style = 'thin', color = '000000'))

    max_row = worksheet.max_row
    # print(f"{(worksheet[f'A{max_row}'].value == datetime.now().strftime('%m/%d/%Y')) = }")
    if worksheet[f'A{max_row}'].value == datetime.now().strftime("%m/%d/%Y"):
        wkbk.close()
        del wkbk
        raise CustomWarning("Dashboard already updated!!", "Dashboard is already updated for today's date.")
    else:
        worksheet[f'A{max_row + 1}'] = datetime.now().strftime("%m/%d/%Y")
        worksheet[f'B{max_row + 1}'] = "SRF MPBN"
        worksheet[f'C{max_row + 1}'] = str(dictionary['Total Picked CR'])
        worksheet[f'D{max_row + 1}'] = str(dictionary['Total Planned CR'])
        worksheet[f'W{max_row + 1}']  = str(dictionary['Night executors count'])
        worksheet[f'X{max_row + 1}'] = str(dictionary['Day Planners count'])
        worksheet[f'Y{max_row + 1}'] = str(dictionary['Resources on comp-off'])
        worksheet[f'Z{max_row + 1}'] = str(dictionary['Resources on leaves'])
        worksheet[f'AA{max_row + 1}'] = 'NA'

        # # Hiding the unwanted rows
        # i = 3
        # while i < max_row:
        #     worksheet.row_dimensions[i].hidden = True
        #     i += 1

        # Aligning the table cells and setting the border.
        i = 1
        while get_column_letter(i) != "AB":
            worksheet[get_column_letter(i) + str(max_row + 1)].alignment = alignment_var
            worksheet[get_column_letter(i) + str(max_row + 1)].border = border
            i += 1

        wkbk.save(path)
        wkbk.close()
        del wkbk

def srf_mpbn_dashboard_tracker_file_getter_and_mail_drafter(path: str, sender:str):
    outlook = win32.Dispatch('Outlook.Application')
    mapi = outlook.GetNamespace("MAPI")
    inbox = mapi.GetDefaultFolder(6)
    messages = inbox.Items
    messages.Sort('[ReceivedTime]', True)
    subject_we_are_looking_for = "RE: SRF MPBN Dashboard Report"
    acceptable_delivered_time = datetime.now() - timedelta(days=7)
    message_found = False
    message_to_be_used = None
    # #print(messages[0].ReceivedTime)

    i = 0
    while i < len(messages):
        try:
            message = messages[i]
            dt = message.ReceivedTime
            #print(dt)
            year, month, day, hour, minute = dt.year, dt.month, dt.day, dt.hour, dt.minute
            dt = datetime(year=year,
                          month=month,
                          day=day,
                          hour=hour,
                          minute=minute)
            #print(dt)

            if dt >= acceptable_delivered_time:
                # print(message.Subject.upper())
                # print(message.Subject.upper().__contains__(subject_we_are_looking_for.upper()))
                if message.Subject.upper().__contains__(subject_we_are_looking_for.upper()):
                    # print(f"{dt =}")
                    # print("message found and downloading the file")
                    message_found = True
                    # print("189")
                    # print("Attachments: - ", [message.Attachments.Item(i).FileName for i in range(1, len(message.Attachments)+1)])
                    l = 1
                    while l <= len(message.Attachments):
                        if ((str(message.Attachments.Item(l).FileName).strip().upper().startswith("SRF-DASHBOARD")) and
                                (str(message.Attachments.Item(l).FileName).endswith(".xlsx"))):
                            attachment = message.Attachments.Item(l)
                            break
                        l += 1
                    path = os.path.join(os.path.dirname(str(path)), str(attachment.FileName))
                    attachment.SaveAsFile(os.path.join(path))
                    message_to_be_used = message
                    break

                else:
                    i += 1
                    continue
            else:
                break
        except:
            i += 1
            continue

    if not message_found:
        sub_folders = len(inbox.Folders)

        j = 0
        while j < sub_folders:
            neo_messages = inbox.Folders[j].Items
            neo_messages.Sort('[ReceivedTime]', True)

            k = 0
            while k < len(neo_messages):
                try:
                    message = neo_messages[k]
                    dt = message.ReceivedTime
                    year, month, day, hour, minute = dt.year, dt.month, dt.day, dt.hour, dt.minute
                    dt = datetime(year=year,
                                month=month,
                                day=day,
                                hour=hour,
                                minute=minute)

                    if dt >= acceptable_delivered_time:
                        if message.Subject.upper().__contains__(subject_we_are_looking_for.upper()):
                            message_found = True
                            #print("223")
                            l = 1
                            while l <= len(message.Attachments):
                                if ((str(message.Attachments.Item(l).FileName).strip().upper().startswith("SRF-DASHBOARD")) and
                                        (str(message.Attachments.Item(l).FileName).endswith(".xlsx"))):
                                    attachment = message.Attachments.Item(l)
                                    break
                                l += 1
                            path = os.path.join(os.path.dirname(str(path)), str(attachment.FileName))
                            attachment.SaveAsFile(path)
                            # mail_drafter(message, sender, path)
                            message_to_be_used = message
                            break

                        else:
                            k += 1
                            continue
                    else:
                        break
                except:
                    k += 1
                    continue

            if message_found:
                break

            else:
                j += 1
                continue

    if not message_found:
        sub_folders = len(inbox.Folders)

        i = 0
        while i < sub_folders:
            sub_sub_folders = len(inbox.Folders[i].Folders)

            j = 0
            while j < sub_sub_folders:
                neo_neo_messages = inbox.Folders[i].Folders[j].Items
                neo_neo_messages.Sort("[ReceivedTime]", True)

                k = 0
                while k < len(neo_neo_messages):
                    try:
                        message = neo_messages[k]
                        dt = message.ReceivedTime
                        year, month, day, hour, minute = dt.year, dt.month, dt.day, dt.hour, dt.minute
                        dt = datetime(year=year,
                                    month=month,
                                    day=day,
                                    hour=hour,
                                    minute=minute)

                        if dt >= acceptable_delivered_time:
                            # if message.Subject.startswith(subject_we_are_looking_for):
                            if message.Subject.upper().__contains__(subject_we_are_looking_for.upper()):
                                message_found = True
                                #print("267")

                                l = 1
                                while l <= len(message.Attachments):
                                    if ((str(message.Attachments.Item(l).FileName).strip().upper().startswith("SRF-DASHBOARD")) and
                                            (str(message.Attachments.Item(l).FileName).endswith(".xlsx"))):
                                        attachment = message.Attachments.Item(l)
                                        break
                                    l += 1
                                path = os.path.join(os.path.dirname(str(path)), str(attachment.FileName))
                                attachment.SaveAsFile(path)
                                message_to_be_used = message
                                # mail_drafter(message, sender, path)
                                break

                            else:
                                k += 1
                                continue
                        else:
                            break

                    except:
                        k += 1
                        continue

                if message_found:
                    break

                else:
                    j += 1
                    continue

            if message_found:
                break

            else:
                i += 1
                continue

    if message_found:
        # print(path)
        # print(f"{os.path.exists(path) = }")
        srf_mpbn_dashboard_tracker_file_writer(str(path))
        mail_drafter(message_to_be_used, sender, path)

        #print(message_found)


def mail_drafter(message: str, sender: str, path: str):
    global dictionary
    date_to_be_added = datetime.now() + timedelta(1)
    date_to_be_added = f"{date_to_be_added.__format__('%d')}{suffix_adder(date_to_be_added)} {date_to_be_added.__format__('%b')} {date_to_be_added.__format__('%Y')}"
    subject = f"RE: SRF MPBN Dashboard Report {date_to_be_added}"
    mail_draft = message.ReplyAll()
    # with open(r"C:/Users/emaienj/Downloads/New_testing_of_day/mail_body.txt", "w") as f:
    #     f.write(mail_draft.Body)
    #     f.close()
    # del f
    result = email_parser(mail_draft.Body)

    to = f"{';'.join(result[0])}"
    # print(f"{to =}")
    cc = f"{';'.join(result[1])}"
    # print(f"{cc =}")
    message_to_be_sent = ("<html>" +
                                "<body>" +
                                      "<div>" +
                                          "<p>Hi Team,</p><br>" +
                                          "<p>Please find the status of MPBN dashboard for tonight MW.</p><br><br>" +
                                      "</div>" +
                                      "<div>" +
                                            f"<p>{table_creater(dictionary= dictionary)}<br></p>" +
                                      "</div>" +
                                      "<div>" +
                                            f"<p>Regards,<br>" +
                                                    f"<br>{sender}<br>" +
                                                    "SDU Bharti | SRF-MPBN<br>" +
                                                    "Ericsson India Global Services Pvt. Ltd.<br>" +
                                            "</p>" +
                                      "</div>" +
                                "</body>" +
                            "</html>")
    mail_draft.HTMLBody = message_to_be_sent + mail_draft.HTMLBody
    mail_draft.Subject = subject
    mail_draft.To = to
    mail_draft.CC = cc
    mail_draft.Attachments.Add(str(path))
    mail_draft.Save()
    #print("mail draft saved")
    mail_draft.Display()


def main_dashboard_func(workbook, sender, dictionary_for_mail):
    global flag
    flag = ""
    main_folder = os.path.dirname(workbook)
    # print(f"{main_folder = }")
    path_for_srf_mpbn_fni_tracker_dashboard_file = os.path.join(main_folder, 'SRF-DASHBOARD-TRACKER.xlsx')

    try:
        if os.path.exists(path_for_srf_mpbn_fni_tracker_dashboard_file):
            os.remove(path_for_srf_mpbn_fni_tracker_dashboard_file)
            from time import sleep
            sleep(2)

        global dictionary
        dictionary = dictionary_for_mail

        srf_mpbn_dashboard_tracker_file_getter_and_mail_drafter(path=path_for_srf_mpbn_fni_tracker_dashboard_file,
                                                                sender=sender)
        if flag != "Unsuccessful":
            flag = "Successful"

    except CustomException as e:
        flag = 'Unsuccessful'

    except CustomWarning as e:
        flag = 'Successful'

    except Exception as e:
        flag = "Unsuccessful"
        title =  e.__class__.__name__
        import traceback
        messagebox.showerror(
            title= title,
            message= (str(e) + "\n\n" + str(traceback.format_exc()))
        )

    finally:
        return flag

# main_dashboard_func(r"C:\Users\emaienj\Downloads\New_folder_(3)\MPBN Daily Planning Sheet.xlsx", "Enjoy Maity",{"Resources on leaves" : 0,
#                                                                                                                 "Resources on comp-off": 0,
#                                                                                                                 "Domain":"SRF MPBN",
#                                                                                                                 "Night executors count": 15,
#                                                                                                                 "Total Picked CR": 29,
#                                                                                                                 "Total Planned CR": 29,
#                                                                                                                 "Day Planners count": 3})
