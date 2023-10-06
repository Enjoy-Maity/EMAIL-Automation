from subprocess import Popen                                        # Importing for opening applications like notebook through cmd line command.
import pandas as pd                                                 # Importing for reading Excel Sheet data and Manipulating it.
from tkinter import messagebox                                      # Impoprting for showing messages.
from pathlib import Path                                            # Importing Path from pathlib module to check validate the path of a file if exists or not.
from openpyxl import Workbook                                       # Importing the Workbook from openpyxl module, so that we can create new workbook and sheets if required.
from openpyxl import load_workbook                                  # Importing the Workbook from openpyxl module, to load and check whether extra sheets are created or not.
import win32com.client as win32                                     # Importing the win32com.client for creating the COM Object of win32
from openpyxl.styles import Font,Border,Side,PatternFill,Alignment  # Importing classes from openpyxl to style the excel workbooks.
from openpyxl.utils import get_column_letter                        # Importing the get_column_letter from openpyxl to convert the column numbers to alphabet letter used in the excel sheet.
from datetime import datetime, timedelta                            # Importing the datetime and timedelta from datetime module, to filter out the excel sheet basd on today's maintenance date.
from openpyxl.worksheet.datavalidation import DataValidation        # Importing DataValidation from the openpyxl module to add data validation onto fields in email-package
import numpy as np
import shutil
from Custom_Warning import CustomWarning

flag = ""
workbook1 = ""
workbook2 = ""
workbook3 = ""

# Creating classes for custom made exceptions inheriting the default Exception class for raising and handling custom raised exceptions.
class CustomException(Exception):
    # Defining the Counstructor method for the CustomException Class
    def __init__(self,title,message):
        self.title      = title 
        self.message    = message
        
        #Calling the super(base) class and passing the arguments to the base class.
        super().__init__(self.title,self.message)
        
        # Displaying the message with the custom exception title passed to the object of the class for the CustomException.
        messagebox.showerror(self.title, self.message)

# Method for hiding sheet
def sheet_hider(workbook,worksheet):
    wb = load_workbook(workbook)
    ws = wb[worksheet]
    ws.sheet_state = "hidden"
    wb.save(workbook)
    wb.close()


# Creating method for styling the worksheet.
def styling(workbook,sheetname):
    wb  =  load_workbook(workbook)                          # loading the workbook.
    ws  =  wb[sheetname]                                    # loading the worksheet.
    font_style  =  Font(color = "FFFFFF",bold = True)       # Setting the font style with color.
    col_widths = []                                         # Empty list for entering the max length of string in each column.

    # Iterating through the row values to find the max length of string in each column in the row and appending it to the col_widths list

    for row_values in ws.iter_rows(values_only = True):
        for j,value in enumerate(row_values):
            if len(col_widths)>j:
                if col_widths[j] < len(str(value)):
                    col_widths[j] = len(str(value))
            else:
                col_widths.insert(j,len(str(value)))

    # Standardising the length of each column in the sheet.

    for i,column_width in enumerate(col_widths,1):
        if column_width <= 47:
            ws.column_dimensions[get_column_letter(i)].width = column_width+3
        else:
            ws.column_dimensions[get_column_letter(i)].width = 50


    # Coloring the headers and alingning the headers text to center both horizontally and vertically.
    for column in range(1,ws.max_column+1):   # ws.max_column returns the total number of columns present
        col = get_column_letter(column)
        color_fill = PatternFill(start_color = '0033CC',end_color = '0033CC',fill_type = 'solid')
        ws[col+'1'].font = font_style
        ws[col+'1'].fill = color_fill
        ws[col+'1'].alignment = Alignment(horizontal = 'center',vertical = 'center')

    # Styling all the occupied cells in the sheet, by adding border to the cells, aligning the text in the center
    
    for row in ws:
        for cell in row:
            cell.alignment = Alignment(horizontal = 'center',vertical = 'center',wrap_text=False)
            cell.border = Border(top = Side(border_style = 'medium',color = '000000'),bottom = Side(border_style = 'medium',color = '000000'),left = Side(border_style = 'medium',color = '000000'),right = Side(border_style = 'medium',color = '000000'))

    # Saving the workbook with worksheet with all the changes.
    wb.save(workbook)
    wb.close()
    del wb
    
    objects = dir()
    for object in objects:
        if not object.startswith("__"):
            del object

# Method(Function) for Drafting Mail.
def mail_drafter(dataframe,dataframe_for_top_table,html_body,sender,execution_date,email_package_workbook,maintenance_window):
    mail_draft          = win32.Dispatch('Outlook.Application')
    mail_draft          = mail_draft.CreateItem(0)
    mail_draft.To       = 'PDLMPBNSRF@pdl.internal.ericsson.com;'
    mail_draft.CC       = 'rohit.mahajan@ericsson.com;karan.k.loomba@ericsson.com;PDLPBNSRFP@pdl.internal.ericsson.com;vishal.kumar.garg@ericsson.com'
    # mail_draft.To       = 'enjoy.maity@ericsson.com;'
    # mail_draft.CC       = 'enjoy.maity@ericsson.com'
    mail_draft.Subject  = f"MPBN CRs For Tonight Maintenance Window - {execution_date} {maintenance_window}"

    dataframe = dataframe.style.set_table_styles([
        {'selector':'th','props':'border:1px solid black; border-collapse : collapse; color:white;padding: 10px; background-color:rgb(0, 51, 204);text-align:center;'},
        {'selector':'tr','props':'border:1px solid black; border-collapse : collapse;padding: 10px;text-align:center;'},
        {'selector':'td','props':'border:1px solid black; border-collapse : collapse;padding: 10px;text-align:center;'},
        {'selector':'tr:nth-child(even)','props':'border:1px solid black; border-collapse : collapse;padding: 10px;text-align:center;'}])

    dataframe = dataframe.hide(axis='index') # hiding the index coloumn
    
    # Stylising the top table which include the information for the leaves, comp-off, etc.
    dataframe_for_top_table = dataframe_for_top_table.style.set_table_styles([
        {'selector':'th','props':'border:2px solid black; border-collapse : collapse; color:black;padding: 10px; background-color:#6092B6;text-align:center;font-weight:bold;'},
        {'selector':'tr','props':'border:2px solid black; border-collapse : collapse;padding: 10px;text-align:center;font-weight:bold;'},
        {'selector':'td','props':'border:2px solid black; border-collapse : collapse;padding: 10px;text-align:center;font-weight:bold;'},
        {'selector':'tr:nth-child(even)','props':'border:1px solid black; border-collapse : collapse;padding: 10px;text-align:center;font-weight:bold;'}])

    dataframe_for_top_table = dataframe_for_top_table.hide(axis='index') # hiding the index coloumn

    mail_draft.HTMLBody = html_body.format(execution_date,maintenance_window,dataframe_for_top_table.to_html(index = False,headers = False),dataframe.to_html(index = False),sender)
    mail_draft.Attachments.Add(email_package_workbook)
    
    # Saving the mail draft
    mail_draft.Save()
    # Displaying the mail draft
    mail_draft.Display()

    objects = dir()
    for object in objects:
        if not object.startswith("__"):
            del object

# Method(Function) for creating email package workbook and mail draft.
def email_package_workbook_generator(sender,worksheet,mail_id_sheet,folder,execution_date,evening_message_workbook_message,maintenance_window):
    # Creating Workbook File Path 
    workbook = rf"{folder}\\MPBN_Email_Package_{execution_date}.xlsx"
    # global workbook3; workbook3 = workbook
    
    # Checking if the Email_Package workbook is created or not.
    if(Path(workbook).exists() == False):
        wb = Workbook()
        wb.create_sheet("Email-Package", index = 0)
        wb.save(workbook)
        wb.close()
        del wb

        # Loading the workbook
        wb = load_workbook(workbook)

        # Getting the sheetnames
        wb_sheets = wb.sheetnames

        # Iterating through the loop to see whether there'sother sheet present in the workbook, if yes then removing then.
        for sheet in wb_sheets:
            if(sheet != "Email-Package"):
                del wb[sheet]
        wb.save(workbook)
        wb.close()
        del wb

    # Changing the NA to blank space.
    worksheet.replace("TempNA"," ",inplace = True)

    # Creating the Writer Variable to write into the new excel workbook
    new_workbook = pd.ExcelFile(workbook)
    writer = pd.ExcelWriter(new_workbook,engine = 'openpyxl', mode = 'a', if_sheet_exists = 'replace')
    worksheet.to_excel(writer,"Email-Package",index = False)
    mail_id_sheet.to_excel(writer,"Circle Mail Id",index= False)
    writer.close()
    del writer

    styling(workbook, "Email-Package")
    styling(workbook,"Circle Mail Id")
    sheet_hider(workbook,"Circle Mail Id")

    # Creating the Html Body for the Body
    html_body = "<html>\
                        <body>\
                            <div>\
                                <p>Hi Team,</p>\
                                <p>Please find the list of MPBN planned activities for tonight Maintenance Window <strong>{} {}</strong>.<br><br>\
                                   Also, Please find the attached e-mail package and filter by your name - you will get the activity detail assigned to you.</p>\
                                <p><span style = 'background-color:#FFFF00'><strong>Note: Final CR Status may vary post CAB discussion.</strong></span></p>\
                            </div>\
                            <div>\
                                <br>\
                                    {}\
                                <br><br>\
                            </div>\
                            <div>\
                                {}\
                                <br>\
                                <br>\
                            </div>\
                            <div>\
                                <p>Thanks & Regards,<br>{}<br>SRF MPBN | SDU Bharti<br>Ericsson India Global Services Pvt. Ltd.</p>\
                            </div>\
                        </body>\
                    </html>"
    
    # Filtering out required columns
    worksheet = worksheet[['S.NO',
                           'Execution Date',
                           'Maintenance Window',
                           'CR NO',
                           'Activity Title',
                           'Risk',
                           'Location',
                           'Circle',
                           'No. of Node Involved',
                           'CR Belongs to Same Activity of Previous CR- Yes/NO',
                           'Change Responsible',
                           'Activity Checker']]
    
    worksheet.reset_index(drop = True, inplace = True)

    # Calling the Mail Drafter Method for drafting the mail but not send it.
    mail_drafter(worksheet,evening_message_workbook_message,html_body,sender,execution_date,workbook,maintenance_window)

    del new_workbook

    # Deleting the objects before the returning back to the main function.
    objects = dir()
    for object in objects:
        if not object.startswith("__"):
            del object


# Method(Function) for creating the evening message text.
def evening_task (sender,night_shift_lead,buffer_auditor_trainer,resource_on_automation,workbook,acceptable_change_responsible):
    try:
        wb=pd.ExcelFile(workbook)
        ws=wb.sheet_names
        worksheet=''
        mail_id_sheet = ''

        # Finding the Email package worksheet.
        if('Email-Package' in ws):
            worksheet='Email-Package'
        
        if('Mail Id' in ws):
            mail_id_sheet = 'Mail Id'
        

        if (len(worksheet) == 0):
            # messagebox.showwarning(' Email-Package Worksheet not Present','Kindly Click the Button for Interdomain Kpi Data Prep First!')
            del mail_id_sheet
            del worksheet
            del wb
            # Deleting all the variables before returning the value "Unsuccessful"
            objects = dir()
            for object in objects:
                if not object.startswith("__"):
                    del object

            flag = 'Unsuccessful'
            raise CustomWarning(' Email-Package Worksheet not Present',"Kindly Click the Button for 'Email Package Preparation' First!")
        
        if (len(mail_id_sheet) == 0):
            # messagebox.showwarning(' Email-Package Worksheet not Present','Kindly Click the Button for Interdomain Kpi Data Prep First!')
            del mail_id_sheet
            del worksheet
            del wb
            # Deleting all the variables before returning the value "Unsuccessful"
            objects = dir()
            for object in objects:
                if not object.startswith("__"):
                    del object

            flag = 'Unsuccessful'
            raise CustomWarning(' Mail Id Worksheet not Present','Kindly Check the selected input workbook for Mail Id sheet!')
        
        else:
        # Reading relevant sheet.
            worksheet=pd.read_excel(wb,worksheet)
            mail_id_sheet= pd.read_excel(wb,mail_id_sheet)

            # Checking Condition for the data pertaining to today's maintenance date being non-existent.
            if (len(worksheet) == 0):
                # messagebox.showwarning(' Email-Package Worksheet Empty','Kindly Click the Button for Interdomain Kpi Data Prep First!')
                del worksheet
                del wb
                # Deleting all the variables before returning the value "Unsuccessful"
                objects = dir()
                for object in objects:
                    if not object.startswith("__"):
                        del object

                flag = 'Unsuccessful'
                raise CustomWarning(' Email-Package Worksheet Empty','Kindly Click the Button for Interdomain Kpi Data Prep First!')
            
            strings_to_be_deleted = ['Select Your Name!','No']
            array_of_unique_change_responsible = worksheet.dropna().unique()
            new_acceptable_change_responsible = np.array(acceptable_change_responsible)
            new_acceptable_change_responsible = np.setdiff1d(new_acceptable_change_responsible,strings_to_be_deleted)
            masks_for_checks_in_acceptable_change_responsible_and_array_of_unique_change_responsible = np.isin(array_of_unique_change_responsible,
                                                                                                           new_acceptable_change_responsible,
                                                                                                           assume_unique=True)
            if(False in masks_for_checks_in_acceptable_change_responsible_and_array_of_unique_change_responsible):
                raise CustomException("    Executor Name Missing!",
                                  f"{', '.join(np.setdiff1d(array_of_unique_change_responsible,acceptable_change_responsible))} executors are not present in your uploaded Change Responsible list text file, Please Check!")
            
            total_no_of_crs=len(worksheet)      # getting the total number of Crs
            total_no_of_resource = 16           
            
            # Initializing the variables for counting number of CRs with Critical, Major Risk levels for all the circles along with Delhi's seperate 
            # count of Risk levels.
            critical = 0                        # Risk Level 1
            delhi_critical = 0
            major = 0                           # Risk Level 2 
            delhi_major = 0

            # Initializing the variables for counting the manual, create, enable and partially automated CR's.
            manual = 0
            create = 0
            enable = 0
            partially_automation = 0

            # Getting the maintenance windo from the excel sheet.
            maintenance_window = f"({worksheet.at[0,'Maintenance Window']})"
            
            # Creating a dictionary to give the subscipt of the month name based on the month number.
            month_dictionary = {
                '01' : 'Jan',
                '02' : 'Feb',
                '03' : 'Mar',
                '04' : 'Apr',
                '05' : 'May',
                '06' : 'Jun',
                '07' : 'Jul',
                '08' : 'Aug',
                '09' : 'Sep',
                '10' : 'Oct',
                '11' : 'Nov',
                '12' : 'Dec'
            }

            # Splitting the date to format it to be sent through the message.
            worksheet['Execution Date'] = pd.to_datetime(worksheet['Execution Date'])
            worksheet['Execution Date'] = worksheet['Execution Date'].dt.strftime('%m/%d/%Y')
            exec_date = worksheet.at[0,'Execution Date']
            exec_date = exec_date.strip().split('/')
            
            # Adding suffices to the date.
            suffixes = { 1: 'st' , 2: 'nd' , 3: 'rd'}
            day = ''    
            if (3 < int(exec_date[1]) < 21) or (23 < int(exec_date[1]) < 31):
                day = f'{int(exec_date[1]):02d}th'
            else:
                day = f'{int(exec_date[1]):02d}{suffixes[int(exec_date[1])%10]}'

            execution_date= f'{day} {month_dictionary[exec_date[0]]} {exec_date[2]}'

            
            resources_occupied_in_night_activities = worksheet['Change Responsible'].unique()
            resources_occupied_in_night_activities = resources_occupied_in_night_activities.astype(str)

            resources_occupied_in_night_activities = np.delete(resources_occupied_in_night_activities,np.where((resources_occupied_in_night_activities == 'nan')|(resources_occupied_in_night_activities == 'Nan')))
            

            # Filling the blank fields in the dataframe with 'NA'.
            worksheet.fillna("TempNA", inplace = True)

            '''
                Iterating (Looping) over the dataframe for finding the number of critical and major risk level along with the number of CR's done with 
                Automation, Partial-Automation and Manually.
            '''
            for row in range(0,len(worksheet)):
                if worksheet.at[row,'Risk'].strip() == 'Level 1':
                    critical+=1
                    if worksheet.at[row,'Circle'].strip() == 'DL':
                        delhi_critical+=1
                    if (worksheet.at[row,'Execution Projection'].strip().upper() == 'MANUAL') or (worksheet.at[row,'Execution Projection'].strip().upper() == 'MANNUAL'):
                        manual+=1
                    if (worksheet.at[row,'Execution Projection'].strip().upper() == 'CREATE') or (worksheet.at[row,'Execution Projection'].strip().upper() == 'CRETA'):
                        create+=1
                    if (worksheet.at[row,'Execution Projection'].strip().upper() == 'ENABLE'):
                        enable+=1
                    if (worksheet.at[row,'Execution Projection'].upper().__contains__('ENABLE')) and (worksheet.at[row,'Execution Projection'].upper().__contains__('MANUAL')):
                        partially_automation+=1
                    if (worksheet.at[row,'Execution Projection'].upper().__contains__('CREATE')) and (worksheet.at[row,'Execution Projection'].upper().__contains__('MANUAL')):
                        partially_automation+=1
                if worksheet.at[row,'Risk'].strip() == 'Level 2':
                    major+=1
                    if worksheet.at[row,'Circle'].strip() == 'DL':
                        delhi_major+=1
                    if (worksheet.at[row,'Execution Projection'].strip().upper() == 'MANUAL') or (worksheet.at[row,'Execution Projection'].strip().upper() == 'MANNUAL'):
                        manual+=1
                    if (worksheet.at[row,'Execution Projection'].strip().upper() == 'CREATE') or (worksheet.at[row,'Execution Projection'].strip().upper() == 'CRETA'):
                        create+=1
                    if (worksheet.at[row,'Execution Projection'].strip().upper() == 'ENABLE'):
                        enable+=1
                    if (worksheet.at[row,'Execution Projection'].upper().__contains__('ENABLE')) and (worksheet.at[row,'Execution Projection'].upper().__contains__('MANUAL')):
                        partially_automation+=1
                    if (worksheet.at[row,'Execution Projection'].upper().__contains__('CREATE')) and (worksheet.at[row,'Execution Projection'].upper().__contains__('MANUAL')):
                        partially_automation+=1

            # Finding the Number of resources that are on leave.
            if(night_shift_lead in resources_occupied_in_night_activities):
                resource_on_leave = total_no_of_resource - (3 + len(resources_occupied_in_night_activities))
            
            else:
                resource_on_leave = total_no_of_resource - (3 + len(resources_occupied_in_night_activities) + 1)

            if ((resource_on_leave) < 0):
                resource_on_leave = 0       # If the value of Number of resources falls below zero and becomes negative, which isn't possible, setting it to zero.
            
            # 02d ensures that the integer is printed in double digit format
            # Creating the message text that is going to be sent via telegram.
            message = """
    Dear Sir,

    <<Pre Notification Critical & Major CRs>>
    Date : {}
    {}
    Total : {} ({:02d} Critical {:02d} Major )
    | Team : MPBN
    • Bharti-I ->
    Critical {:02d} ({:02d} Delhi)
    Major    {:02d} ({:02d} Delhi)
    Resource's occupied in night activities : {:02d}
    Resource in Day/Planning :: 3
    Resource on Comp off/Leave :- {}
    Night Shift Lead :: {}
    Resource engaged in CLI Preparation :: N/A
    Resource on Buffer/Auditor/Training : {}
    Resource on Automation : {}
    Rollback CR re-attempt count : 0
    Partially completed CR re-attempt count : 0
    Updated automation CR count
    ======================
    Total CRs                       :{}
    CR Planned Manually             :{}
    CR Planned via Enable Tool      :{}
    CR Planned via CREATE Tool      :{}
    CR Planned Partial Automation   :{}
    ======================

    Regards,
    {}
            """
            # Adding all the other relevant data to the message text by formatting it.
            message = message.format(execution_date,
                                     maintenance_window,
                                     total_no_of_crs,
                                     critical,
                                     major,
                                     critical,
                                     delhi_critical,
                                     major,
                                     delhi_major,
                                     len(resources_occupied_in_night_activities),
                                     resource_on_leave,night_shift_lead,
                                     buffer_auditor_trainer,
                                     resource_on_automation,
                                     total_no_of_crs,
                                     manual,
                                     enable,
                                     create,
                                     partially_automation,
                                     sender)
            
            # Creating the file path where the text file for the message is being saved.
            file_path = workbook.split("/")
            file_path.remove(file_path[-1])
            file_path = '\\'.join(file_path)
            folder    = file_path
            file_path = rf'{file_path}\\evening message.txt'
            
            # Writing the text into the file defined by the file path.
            with open(file_path,'w') as f:
                f.write(message)
                f.close()
            messagebox.showinfo("   Task Completed Successfully",f"Evening Message generated successfully at {file_path}")
            del f
            # Asking for response, whether thr user wants to check the message being created.
            response = messagebox.askyesno("   Evening Message","Do You want to open the Evening Message text?")
            
            # If the response is positive, then the created text message is opened in notebook via the use of Popen from subprocess module.
            if (response):
                Popen(['notepad.exe',file_path])
            
            else:
                pass
            
            # Setting the folder in the folder path for creation of the email package workbook.
            temp_folder = rf"{folder}\\temp"

            if(Path(temp_folder).exists()):
                try:
                    shutil.rmtree(temp_folder)
                
                except:
                    import os
                    for file in os.listdir(temp_folder):
                        os.remove(os.path.join("\\",temp_folder,file))
                    
                    os.rmdir(temp_folder)

            # creating a new temp folder with temp excel file for the data to write into.
            Path(rf"{folder}\\temp").mkdir(exist_ok=True)

            # Writing into a temp xlsx file named temp.xlsx
            temp_file_for_the_evening_message = rf"{temp_folder}\\tmp.xlsx"
            # global workbook2; workbook2 = temp_file_for_the_evening_message

            # Checking if the temp_file_for_evening_message is existent or not
            if (Path(temp_file_for_the_evening_message).exists() == False):
                
                # Creating a openpyxl.Workbook variable for creation and manipulation of the workbook.
                wb = Workbook()

                # Creating sheet for the file named evening message
                wb.create_sheet("evening_message", index = 0)
                
                # Opening the sheet to write into it.
                ws = wb["evening_message"]
                
                # Loading required data and info field.
                ws['A1'].value = "Resource's occupied in night activities"
                ws['A2'].value = "Resource in Day/Planning"
                ws['A3'].value = "Resource on Comp off"
                ws['A4'].value = "Resource on Leave"
                ws['A5'].value = "Night Shift Lead "
                ws['A6'].value = "Resource occupied in 2nd Level Validation/Buffer/Training"
                ws['A7'].value = "Resource on Training "
                ws['A8'].value = "Total CR’s"

                ws['B1'].value = 11
                ws['B2'].value = 3
                ws['B3'].value = 0
                ws['B4'].value = resource_on_leave
                ws['B5'].value = 1
                ws['B6'].value = 0
                ws['B7'].value = 0
                ws['B8'].value = total_no_of_crs

                ws['C1'].value = " "
                ws['C5'].value = night_shift_lead
                # Closing the openpyxl.Workbook variable.
                wb.save(temp_file_for_the_evening_message)
                wb.close()
                del wb
            
            else:
                # Loading the xlsx file in openpyxl module of python
                wb = load_workbook(temp_file_for_the_evening_message)

                # Reading the required sheet.
                ws = wb["evening_message"]
                
                ws['B1'].value = 11
                ws['B2'].value = 3
                ws['B3'].value = 0
                ws['B4'].value = resource_on_leave
                ws['B5'].value = 1
                ws['B6'].value = 0
                ws['B7'].value = 0
                ws['B8'].value = total_no_of_crs

                ws['C1'].value = " "
                ws['C5'].value = night_shift_lead

                # Closing and saving the openpyxl.Workbook variable.
                wb.save(temp_file_for_the_evening_message)
                wb.close()
                del wb
            
            evening_message_workbook = pd.ExcelFile(temp_file_for_the_evening_message)
            evening_message_workbook_message = pd.read_excel(evening_message_workbook,'evening_message')
            evening_message_workbook_message.fillna(" ",inplace = True)

            # Calling the Email Package Workbook generator and mail drafter.
            email_package_workbook_generator(sender,worksheet,mail_id_sheet,temp_folder,execution_date,evening_message_workbook_message,maintenance_window)

            response = messagebox.showinfo("    SRF MPBN Team Availability Tracker","We are going to update SRF_MPBN_Team_Availability tracker. So, Please ensure to download latest tracker before proceeding!")

            if(response.lower() == 'ok'):
                import attendance
                flag = attendance.main_function(workbook=workbook,
                                                night_shift_lead = night_shift_lead,
                                                buffer_auditor_trainer = buffer_auditor_trainer,
                                                resource_on_automation = resource_on_automation,
                                                acceptable_change_responsible = acceptable_change_responsible,
                                                sender = sender)
            else:
                flag = "Unsuccessful"

            # Deleting all the local variables 
            objects = dir()
            for object in objects:
                if not object.startswith("__"):
                    del object
            
            # flag = 'Successful'

    # Handling Exceptions 
    except CustomException:
        # Delelting all the local variables before returning the value "Unsuccessful"
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"
    
    except Exception as e:
        import traceback
        messagebox.showerror("  Exception Occured!",f"{traceback.format_exc()}\n{e}")

        # Delelting all the local variables before returning the value "Unsuccessful"
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"
    
    finally:
        import gc
        # excel = win32.Dispatch("Excel.Application")

        # if(len(workbook1) > 0):
        #     wb = excel.Workbooks.Open(workbook1)
        #     wb.Close()
        
        # if(len(workbook2) > 0):
        #     wb = excel.Workbooks.Open(workbook2)
        #     wb.Close()
        
        # if(len(workbook3) > 0):
        #     wb = excel.Workbooks.Open(workbook3)
        #     wb.Close()

        # excel.Application.Quit()
        gc.collect()
        return flag

    

# evening_task('Manoj Kumar','Sachin Sharma','NA','Kartar Singh',"C:/Users/emaienj/Downloads/MPBN Daily Planning Sheet-1.xlsx")