import pandas as pd                                                 # Importing Pandas for reading csv and excel files and  manipulating it.
from tkinter import messagebox                                      # Importing Messagebox from tkinter for displaying messages.
from openpyxl import load_workbook                                  # Importing load_workbook class from the openpyxl to load existing excel workbook.
from openpyxl.styles import Font,Border,Side,PatternFill,Alignment  # Importing classes from openpyxl to style the excel workbooks.
from openpyxl.utils import get_column_letter                        # Importing the get_column_letter from openpyxl to convert the column numbers to alphabet letter used in the excel sheet.
from openpyxl.worksheet.datavalidation import DataValidation        # Importing DataValidation from the openpyxl module to add data validation onto fields in planning sheet.

flag = ""
workbook1 = ""

# Creating Custom Exception inheriting base default Exception class for raising, handling and custom exceptions.
class CustomException(Exception):
    def __init__(self,title,message):
        self.title = title
        self.message = message
        super().__init__(title,message)
        messagebox.showerror(self.title, self.message)

# Method(Function) for styling the worksheets.
def styling(workbook,sheetname):
    wb  =  load_workbook(workbook)
    ws  =  wb[sheetname]
    font_style  =  Font(color = "FFFFFF",bold = True)
    col_widths = []

    # Iterating through the row values to find the length of string in each column in the row and appending it to the col_widths list

    for row_values in ws.iter_rows(values_only = True):
        for j,value in enumerate(row_values):
            if len(col_widths)>j:
                if col_widths[j] < len(str(value)):
                    col_widths[j] = len(str(value))
            else:
                col_widths.insert(j,len(str(value)))

    # Standardising the length of each column in the sheet.

    for i,column_width in enumerate(col_widths,1):
        if column_width <= 47:
            ws.column_dimensions[get_column_letter(i)].width = column_width+3
        else:
            ws.column_dimensions[get_column_letter(i)].width = 50


    # Coloring the headers and alingning the headers text to center both horizontally and vertically.
    for column in range(1,ws.max_column+1):   # ws.max_column returns the total number of columns present
        col = get_column_letter(column)
        color_fill = PatternFill(start_color = '0033CC',end_color = '0033CC',fill_type = 'solid')
        ws[col+'1'].font = font_style
        ws[col+'1'].fill = color_fill
        ws[col+'1'].alignment = Alignment(horizontal = 'center',vertical = 'center')

    # Styling all the occupied cells in the sheet, by adding border to the cells, aligning the text in the center
    
    for row in ws:
        for cell in row:
            cell.alignment = Alignment(horizontal = 'center',vertical = 'center',wrap_text=False)
            cell.border = Border(top = Side(border_style = 'medium',color = '000000'),bottom = Side(border_style = 'medium',color = '000000'),left = Side(border_style = 'medium',color = '000000'),right = Side(border_style = 'medium',color = '000000'))

    #rows = ws.max_row
    wb.save(workbook)
    wb.close()
    del wb

    # Deleting all the variables before returning to main method.
    objects = dir()
    for object in objects:
        if not object.startswith("__"):
            del object

# Adding Data Validation
def validation_adder(workbook,worksheet):
    wb                          = load_workbook(workbook)
    ws                          = wb[worksheet]
    # getting max occupied rows
    maxrows                     = ws.max_row

    # Rule for data validation with their error message, title and prompt message, title.
    rule1               = DataValidation(type = "list", formula1 = '"Planned, Discussed"',allow_blank = True)
    rule1.error         = "Your Entry is Invalid!"
    rule1.errorTitle    = "Invalid Entry!"

    rule1.prompt        = "Please Select from the list"
    rule1.promptTitle   = "List Selection"

    rule3               = DataValidation(type = "list", formula1 = '"Create, Enable, Manual, Manual/Enable, Create/Manual"',allow_blank = True)
    rule3.error         = "Your Entry is Invalid!"
    rule3.errorTitle    = "Invalid Entry!"

    rule3.prompt        = "Please Select from the list"
    rule3.promptTitle   = "List Selection"

    # Adding the rules to the woksheet.
    ws.add_data_validation(rule1)
    ws.add_data_validation(rule3)

    # Setting the rows for the rules.
    range_setter_var_planning_status       = f'P2:P{maxrows}'
    range_setter_var_execution_projection  = f'AF2:AF{maxrows}'

    # Adding the ranges to the rules.
    rule1.add(range_setter_var_planning_status)
    rule3.add(range_setter_var_execution_projection)

    # Saving the Workbook
    wb.save(workbook)
    wb.close()
    del wb

    # Deleting all the variables before returning to main method.
    objects = dir()
    for object in objects:
        if not object.startswith("__"):
            del object

# Creating the main function
def planning_sheet_creater(report_path,planning_workbook,sender):
    try:
        # Checking if the length of report_path is given or not
        if (len(report_path) == 0):
            # Raising the custom exception incase the report csv file is not selected.
            raise CustomException(" Report Not Selected!","ITSM Report not selected! Kindly Select the ITSM report!")
        
        # Checking if the planning workbook is selected or not.
        if(len(planning_workbook) == 0):
            # Raising the custom exception in case the Planning workbook is not selected.
            raise CustomException(" Planning Workbook Not Selected!","Kindly Select the MPBN Planning Sheet workbook!")
        
        else:
            # Reading the report csv file. 
            # By default the encoding of the report csv file is in ANSI encoding, so to read it in pandas we have to set encoding "mbcs" or "cp1252"
            try:
                report = pd.read_csv(report_path, encoding = "mbcs")
            
            except:
                try:
                    report = pd.read_csv(report_path, encoding = "cp1252")
                
                except Exception as error:
                    messagebox.showerror("  Exception Occured!",error)
                    
                    # Deleting all the variables before returning the value for "Unsuccessful"
                    objects = dir()
                    for object in objects:
                        if not object.startswith("__"):
                            del object
                    
                    flag = "Unsuccessful"
            
            # Columns for the planning sheet in the planning workbook.
            columns_for_planning_sheet =    ["Execution Date",
                                            "Maintenance Window",
                                            "CR NO",
                                            "Activity Title",
                                            "Risk",
                                            "Location",
                                            "Circle",
                                            "Region",
                                            "No. of Node Involved",
                                            "CR Belongs to Same Activity of Previous CR- Yes/NO",
                                            "Change Responsible",
                                            "Activity Checker",
                                            "Activity Initiator",
                                            "Impact",
                                            "Planning Status",
                                            "Domain",
                                            "Final Status",
                                            "Reason For Rollback / Cancel",
                                            "Design Availability",
                                            "Technical Validator",
                                            "Complexity",
                                            "Activity-Type",
                                            "Domain kpi",
                                            "IMPACTED NODE",
                                            "KPI DETAILS",
                                            "oss name",
                                            "oss ip",
                                            "Total Time spent on Planned CRs (Mins)",
                                            "Vendor",
                                            "Protocol",
                                            "Execution Projection",
                                            "Inter-domain Name",
                                            "Second Level Validation Status",
                                            "Inter-domain KPI status",
                                            "MOP View Status"]
            
            global workbook1;workbook1 = planning_workbook

            # Creating the dataframe for daily_planning_sheet to write into the planning sheet of the planning workbook.
            daily_planning_sheet = pd.DataFrame(columns= columns_for_planning_sheet)
            
            # Filtering out the rows from the report excluding rows where report[Status*] = Draft
            report = report[(report["Status*"].str.lower() != "draft") & (report["Status*"].str.lower() != "rejected")]

            # Resetting the index values of report
            report.reset_index(drop = True,inplace = True)
            # Taking out only required data from the report
            report = report[['Scheduled Start Date+','Change ID*+','Summary*','Impact*','Site Group','Submitter*','Operational Categorization Tier 1+','Operational Categorization Tier 3']]

            # Reading the Mail ID for getting the region from the sheet corresponding to circle or site group.
            mail_id_sheet_region = pd.ExcelFile(planning_workbook)
            mail_id_sheet_region = pd.read_excel(mail_id_sheet_region,"Mail Id")

            # Filtering only needed columns.
            mail_id_sheet_region = mail_id_sheet_region[['Circle','Region']]
            
            # Making a dictionary object from the two columns of the circle vs region.
            mail_id_sheet_region_dictionary_from_circle_to_region = dict(zip(mail_id_sheet_region['Circle'],mail_id_sheet_region['Region']))

            # Formatting the date of the report['Scheduled Start Date+'].
            report["Scheduled Start Date+"] = pd.to_datetime(report["Scheduled Start Date+"])
            report["Scheduled Start Date+"] = report["Scheduled Start Date+"].dt.strftime("%d-%b-%Y")
            
            # Selecting the data from the raw report and entering it in the daily_planning_sheet
            daily_planning_sheet['Execution Date']                                      =   report["Scheduled Start Date+"]
            daily_planning_sheet['Maintenance Window']                                  =   "00:00 To 06:00 Hrs"
            daily_planning_sheet['CR NO']                                               =   report["Change ID*+"]
            daily_planning_sheet['Activity Title']                                      =   report["Summary*"]
            daily_planning_sheet['Risk']                                                =   ""
            daily_planning_sheet['Location']                                            =   ""
            daily_planning_sheet['Circle']                                              =   report["Site Group"]
            daily_planning_sheet['Region']                                              =   report["Site Group"].map(mail_id_sheet_region_dictionary_from_circle_to_region)
            daily_planning_sheet['No. of Node Involved']                                =   ""
            daily_planning_sheet['CR Belongs to Same Activity of Previous CR- Yes/NO']  =   ""
            daily_planning_sheet['Change Responsible']                                  =   ""
            daily_planning_sheet['Activity Checker']                                    =   ""
            daily_planning_sheet['Activity Initiator']                                  =   report["Submitter*"]
            daily_planning_sheet['Impact']                                              =   ""
            daily_planning_sheet['Planning Status']                                     =   ""
            daily_planning_sheet['Domain']                                              =   ""
            daily_planning_sheet['Final Status']                                        =   ""
            daily_planning_sheet['Reason For Rollback / Cancel']                        =   ""
            daily_planning_sheet['Design Availability']                                 =   ""
            daily_planning_sheet['Technical Validator']                                 =   sender
            daily_planning_sheet['Complexity']                                          =   ""
            daily_planning_sheet['Activity-Type']                                       =   report["Operational Categorization Tier 3"]
            daily_planning_sheet['Domain kpi']                                          =   ""
            daily_planning_sheet['IMPACTED NODE']                                       =   ""
            daily_planning_sheet['KPI DETAILS']                                         =   ""
            daily_planning_sheet['oss name']                                            =   ""
            daily_planning_sheet['oss ip']                                              =   ""
            daily_planning_sheet['Total Time spent on Planned CRs (Mins)']              =   ""
            daily_planning_sheet['Vendor']                                              =   ""
            daily_planning_sheet['Protocol']                                            =   ""
            daily_planning_sheet['Execution Projection']                                =   ""
            daily_planning_sheet['Inter-domain Name']                                   =   ""
            daily_planning_sheet['Second Level Validation Status']                      =   ""
            daily_planning_sheet['Inter-domain KPI status']                             =   ""
            daily_planning_sheet['MOP View Status']                                     =   ""
            
            daily_planning_sheet.reset_index(drop = True, inplace = True)
            daily_planning_sheet.index += 1
            daily_planning_sheet.insert(0,"S.NO",daily_planning_sheet.index)
            daily_planning_sheet.reset_index(drop = True, inplace = True)
            
            # Iterating through the report dataframe for writing into the planning sheet
            for i in range(0, len(report)):          
                if(report.iloc[i]['Impact*'].__contains__("1-Extensive/Widespread")):
                    daily_planning_sheet.at[i,"Risk"] = "Level 1"
                
                if(report.iloc[i]['Impact*'].__contains__("2-Significant/Large")):
                    daily_planning_sheet.at[i,"Risk"] = "Level 2"
                
                if ((report.iloc[i]['Impact*'].strip() != "2-Significant/Large") and (report.iloc[i]['Impact*'].strip() != "1-Extensive/Widespread")):
                    daily_planning_sheet.at[i,"Risk"] = report.iloc[i]['Impact*']
                
                if(report.iloc[i]['Operational Categorization Tier 1+'].__contains__("MPBN")):
                    daily_planning_sheet.at[i,"Domain"] = "MPBN-MS"

            # Creating the writer for writing into the planning sheet.
            writer = pd.ExcelWriter(planning_workbook,engine = "openpyxl", mode = "a", if_sheet_exists = "replace")
            daily_planning_sheet.to_excel(writer,"Planning Sheet",index = False)    # writing daily_planning_sheet into the planning sheet.
            writer.close()
            del writer

            # styling the worksheet.
            styling(planning_workbook,"Planning Sheet")
            
            # adding the data validation.
            validation_adder(planning_workbook,"Planning Sheet")

            # Message shown after successful task running.
            messagebox.showinfo("   Sheet Creation Successful!","Tonight CRs Parameter Copied in MPBN Planning Sheet!")
            del daily_planning_sheet
            del report
            
            # Deleting all the variables before returning the value for "Successful"
            objects = dir()
            for object in objects:
                if not object.startswith("__"):
                    del object

            flag = "Successful"

    except CustomException:
        # Deleting all the variables before returning the value for "Unsuccessful"
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"
    
    except PermissionError as e:
        e = str(e).split(":")
        e.remove(e[0])
        e = ':'.join(e)
        messagebox.showerror("    Permission Error!",f"Kindly Close the selected {e} if opened in Excel!")

        # Deleting all the variables before returning the value for "Unsuccessful"
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"

    except Exception as error:
        import traceback
        messagebox.showerror("  Exception Occured!",f"{traceback.format_exc()}\n{error}")
        
        # Deleting all the variables before returning the value for "Unsuccessful"
        objects = dir()
        for object in objects:
            if not object.startswith("__"):
                del object

        flag = "Unsuccessful"

    finally:
        import gc
        # import win32com.client as win32

        # excel = win32.Dispatch("Excel.Application")
        # if(len(workbook1) > 0):
        #     wb = excel.Workbooks.Open(workbook1)
        #     wb.Close()

        # excel.Application.Quit()
        gc.collect()
        return flag

#planning_sheet_creater(r"C:/Users/emaienj/OneDrive - Ericsson/Documents/Report.csv",r"C:\Users\emaienj\Downloads\MPBN Daily Planning Sheet - Copy.xlsx","Enjoy Maity")